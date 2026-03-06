#!/usr/bin/env python3
"""Deterministic data validation for jobs-data-v3.xlsx"""

import openpyxl
import math
from collections import defaultdict, Counter

wb = openpyxl.load_workbook('jobs-data-v3.xlsx', data_only=True)
issues = []

def issue(severity, sheet, description):
    issues.append((severity, sheet, description))

# ========================================
# 1. TASKS SHEET VALIDATION
# ========================================
print("Checking 3 Tasks...")
ws_tasks = wb['3 Tasks']

VALID_AUT_SCORES = {0.0, 0.25, 0.5, 0.75, 1.0}
task_jobs = set()  # (SOC_Code, Job_Title)
task_soc_codes = set()
task_ids_seen = set()
time_shares_by_job = defaultdict(float)
task_count = 0
aut_issues = 0

for r in range(2, ws_tasks.max_row+1):
    soc = ws_tasks.cell(r, 1).value
    job = ws_tasks.cell(r, 2).value
    task_id = ws_tasks.cell(r, 3).value
    task_desc = ws_tasks.cell(r, 4).value
    task_type = ws_tasks.cell(r, 5).value
    time_share = ws_tasks.cell(r, 6).value
    importance = ws_tasks.cell(r, 7).value
    frequency = ws_tasks.cell(r, 8).value
    gwa = ws_tasks.cell(r, 9).value
    dedup_emp = ws_tasks.cell(r, 10).value
    econ_weight = ws_tasks.cell(r, 11).value
    aut_mod = ws_tasks.cell(r, 12).value
    aut_sig = ws_tasks.cell(r, 13).value

    if soc is None and job is None:
        continue
    task_count += 1
    task_jobs.add((soc, job))
    task_soc_codes.add(soc)

    # Duplicate task IDs
    if task_id in task_ids_seen:
        issue("HIGH", "3 Tasks", f"Duplicate Task_ID: {task_id} at row {r}")
    task_ids_seen.add(task_id)

    # Null checks
    for col_name, val in [("SOC_Code", soc), ("Job_Title", job), ("Task_ID", task_id),
                           ("Task_Description", task_desc), ("Task_Type", task_type),
                           ("Time_Share_Pct", time_share), ("GWA", gwa)]:
        if val is None:
            issue("HIGH", "3 Tasks", f"Null {col_name} at row {r}")

    # Aut score validation
    if aut_mod is not None and aut_mod not in VALID_AUT_SCORES:
        issue("HIGH", "3 Tasks", f"Invalid Aut_Score_Mod={aut_mod} at row {r} (task {task_id})")
        aut_issues += 1
    if aut_sig is not None and aut_sig not in VALID_AUT_SCORES:
        issue("HIGH", "3 Tasks", f"Invalid Aut_Score_Sig={aut_sig} at row {r} (task {task_id})")
        aut_issues += 1
    if aut_mod is None:
        issue("HIGH", "3 Tasks", f"Missing Aut_Score_Mod at row {r} (task {task_id})")
    if aut_sig is None:
        issue("HIGH", "3 Tasks", f"Missing Aut_Score_Sig at row {r} (task {task_id})")

    # sig >= mod constraint
    if aut_mod is not None and aut_sig is not None and aut_sig < aut_mod:
        issue("HIGH", "3 Tasks", f"sig ({aut_sig}) < mod ({aut_mod}) at row {r} (task {task_id}, job {job})")

    # Task_Type valid
    if task_type not in ("Core", "Supplemental", None):
        issue("MEDIUM", "3 Tasks", f"Invalid Task_Type='{task_type}' at row {r}")

    # Time share accumulation
    if time_share is not None:
        time_shares_by_job[(soc, job)] += time_share

    # Employment checks
    if dedup_emp is not None and dedup_emp < 0:
        issue("HIGH", "3 Tasks", f"Negative Dedup_Employment_K={dedup_emp} at row {r}")

# Time share sums
time_share_issues = 0
for (soc, job), total in time_shares_by_job.items():
    if abs(total - 100.0) > 2.0:  # Allow 2% tolerance
        issue("MEDIUM", "3 Tasks", f"Time_Share_Pct sums to {total:.1f}% for {job} ({soc}), expected ~100%")
        time_share_issues += 1

print(f"  {task_count} tasks, {len(task_jobs)} unique jobs, {time_share_issues} time-share issues")

# ========================================
# 2. 1A SECTOR SUMMARY VALIDATION
# ========================================
print("Checking 1A Sector Summary...")
ws_summary = wb['1A Sector Summary']
summary_sectors = {}
for r in range(2, ws_summary.max_row+1):
    sid = ws_summary.cell(r, 1).value
    name = ws_summary.cell(r, 2).value
    num_naics = ws_summary.cell(r, 3).value
    emp = ws_summary.cell(r, 4).value
    wc_emp = ws_summary.cell(r, 6).value
    wc_pct = ws_summary.cell(r, 7).value
    if sid is None:
        continue
    summary_sectors[sid] = {"name": name, "emp": emp, "num_naics": num_naics, "wc_emp": wc_emp, "wc_pct": wc_pct}
    if emp is None or emp <= 0:
        issue("HIGH", "1A Sector Summary", f"Invalid employment for sector {sid} ({name})")
    if wc_emp is not None and emp is not None and wc_emp > emp:
        issue("HIGH", "1A Sector Summary", f"WC employment ({wc_emp}) > total ({emp}) for sector {sid} ({name})")

print(f"  {len(summary_sectors)} sectors")

# ========================================
# 3. 1 NAICS MAPPING VALIDATION
# ========================================
print("Checking 1 NAICS Mapping...")
ws_ind = wb['1 NAICS Mapping']
industries_by_sector = defaultdict(list)
industries_emp_by_sector = defaultdict(float)

for r in range(2, ws_ind.max_row+1):
    naics = ws_ind.cell(r, 1).value
    sid = ws_ind.cell(r, 3).value
    emp = ws_ind.cell(r, 8).value

    if naics is None and sid is None:
        continue
    industries_by_sector[sid].append(naics)
    if emp is not None:
        industries_emp_by_sector[sid] += emp

# Check NAICS counts match summary
for sid, data in summary_sectors.items():
    actual = len(industries_by_sector.get(sid, []))
    expected = data["num_naics"]
    if expected and actual != expected:
        issue("MEDIUM", "1 NAICS Mapping", f"Sector {sid} ({data['name']}): {actual} NAICS codes but Summary says {expected}")

# Check employment rollups
for sid, data in summary_sectors.items():
    ind_total = industries_emp_by_sector.get(sid, 0)
    sum_emp = data["emp"]
    if sum_emp and abs(ind_total - sum_emp) > 1.0:
        issue("MEDIUM", "1 NAICS Mapping→Summary", f"Sector {sid} ({data['name']}): Industries sum={ind_total:.1f}K vs Summary={sum_emp:.1f}K")

# ========================================
# 4. FRICTIONS TABS VALIDATION
# ========================================
print("Checking Frictions tabs...")
VALID_E = {0.15, 0.25, 0.30, 0.50, 0.75, 1.00}
friction_sectors = {}

for sheet_name in ['5L Frictions Low', '5H Frictions High']:
    ws = wb[sheet_name]
    for r in range(5, ws.max_row+1):
        sid = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        emp = ws.cell(r, 3).value
        t1 = ws.cell(r, 5).value
        t2 = ws.cell(r, 6).value
        t3 = ws.cell(r, 7).value
        t4 = ws.cell(r, 8).value
        d_idx = ws.cell(r, 9).value
        t0 = ws.cell(r, 10).value
        t18 = ws.cell(r, 11).value
        f1 = ws.cell(r, 12).value
        f2 = ws.cell(r, 13).value
        f3 = ws.cell(r, 14).value
        f_sum = ws.cell(r, 15).value
        r_val = ws.cell(r, 16).value
        e_val = ws.cell(r, 17).value

        if sid is None:
            continue

        if 'Low' in sheet_name:
            friction_sectors[sid] = name

        # T drivers 1-4
        for t_name, t_val in [("T1", t1), ("T2", t2), ("T3", t3), ("T4", t4)]:
            if t_val is not None and (t_val < 1 or t_val > 4):
                issue("HIGH", sheet_name, f"Sector {sid} ({name}): {t_name}={t_val} outside 1-4")
            if t_val is not None and t_val != int(t_val):
                issue("MEDIUM", sheet_name, f"Sector {sid} ({name}): {t_name}={t_val} not integer")

        # R sub-components 1-4
        for f_name, f_val in [("f1", f1), ("f2", f2), ("f3", f3)]:
            if f_val is not None and (f_val < 1 or f_val > 4):
                issue("HIGH", sheet_name, f"Sector {sid} ({name}): {f_name}={f_val} outside 1-4")

        # E valid values
        if e_val is not None and e_val not in VALID_E:
            issue("MEDIUM", sheet_name, f"Sector {sid} ({name}): E={e_val} not in valid set")

        # Verify D formula: avg(T1,T2,T3,T4)
        if all(v is not None for v in [t1, t2, t3, t4, d_idx]):
            expected_d = (t1+t2+t3+t4)/4
            if abs(d_idx - expected_d) > 0.01:
                issue("HIGH", sheet_name, f"Sector {sid} ({name}): D={d_idx:.4f} but expected {expected_d:.4f}")

        # Verify t0 formula: 1.5*D - 0.5*T4 - 0.5
        if all(v is not None for v in [d_idx, t4, t0]):
            expected_t0 = 1.5*d_idx - 0.5*t4 - 0.5
            if abs(t0 - expected_t0) > 0.01:
                issue("HIGH", sheet_name, f"Sector {sid} ({name}): t0={t0:.4f} but expected {expected_t0:.4f}")

        # Verify T(18mo) formula: 1/(1+exp(-alpha*(1.5-t0))) with alpha=1.2
        if all(v is not None for v in [t0, t18]):
            alpha = 1.2
            expected_t18 = 1 / (1 + math.exp(-alpha * (1.5 - t0)))
            if abs(t18 - expected_t18) > 0.01:
                issue("HIGH", sheet_name, f"Sector {sid} ({name}): T(18mo)={t18:.4f} but expected {expected_t18:.4f}")

        # Verify F = f1+f2+f3
        if all(v is not None for v in [f1, f2, f3, f_sum]):
            expected_f = f1+f2+f3
            if abs(f_sum - expected_f) > 0.01:
                issue("HIGH", sheet_name, f"Sector {sid} ({name}): F={f_sum} but expected {expected_f}")

        # Verify R = 1 - 0.7*(F-3)/9
        if all(v is not None for v in [f_sum, r_val]):
            expected_r = 1 - 0.7*(f_sum - 3)/9
            if abs(r_val - expected_r) > 0.01:
                issue("HIGH", sheet_name, f"Sector {sid} ({name}): R={r_val:.4f} but expected {expected_r:.4f}")

        # Employment match with summary
        if sid in summary_sectors and emp is not None:
            sum_emp = summary_sectors[sid]["emp"]
            if sum_emp and abs(emp - sum_emp) > 1.0:
                issue("MEDIUM", sheet_name, f"Sector {sid} ({name}): Emp={emp} vs Summary={sum_emp}")

# ========================================
# 5. STAFFING PATTERNS VALIDATION
# ========================================
print("Checking 2 Staffing Patterns...")
ws_sp = wb['2 Staffing Patterns']
staffing_soc = set()
staffing_by_sector = defaultdict(float)
occ_share_by_soc = defaultdict(float)

for r in range(2, ws_sp.max_row+1):
    sid = ws_sp.cell(r, 1).value
    soc = ws_sp.cell(r, 4).value
    staffing_share = ws_sp.cell(r, 7).value
    occ_share = ws_sp.cell(r, 8).value
    wage = ws_sp.cell(r, 9).value
    if soc:
        staffing_soc.add(soc)
    if sid and staffing_share:
        staffing_by_sector[str(sid)] += staffing_share
    if soc and occ_share:
        occ_share_by_soc[soc] += occ_share

# Check staffing share sums per sector
for sid, total in staffing_by_sector.items():
    if abs(total - 100.0) > 5.0:
        issue("MEDIUM", "2 Staffing Patterns", f"Sector {sid}: Staffing shares sum to {total:.1f}%, expected ~100%")

# Check occupation share sums to ~100% per SOC
occ_share_issues = 0
for soc, total in occ_share_by_soc.items():
    if abs(total - 100.0) > 1.0:
        issue("MEDIUM", "2 Staffing Patterns", f"SOC {soc}: Occupation_Industry_Share sums to {total:.1f}%, expected ~100%")
        occ_share_issues += 1

print(f"  {len(staffing_soc)} unique SOCs, {occ_share_issues} occupation-share issues")

# ========================================
# 6. CROSS-SHEET REFERENTIAL INTEGRITY
# ========================================
print("Checking cross-sheet references...")

# Staffing SOCs vs Tasks SOCs
sp_not_in_tasks = staffing_soc - task_soc_codes
tasks_not_in_sp = task_soc_codes - staffing_soc

if sp_not_in_tasks:
    for soc in sorted(sp_not_in_tasks):
        issue("MEDIUM", "2 Staffing→3 Tasks", f"SOC {soc} in Staffing Patterns but has no tasks")

if tasks_not_in_sp:
    for soc in sorted(tasks_not_in_sp):
        issue("HIGH", "3 Tasks→2 Staffing", f"SOC {soc} in Tasks but not in Staffing Patterns")

# Sectors in Summary vs Frictions
summary_sector_names = {d["name"] for d in summary_sectors.values()}
friction_sector_names = set(friction_sectors.values())

summary_not_in_frictions = summary_sector_names - friction_sector_names
frictions_not_in_summary = friction_sector_names - summary_sector_names

if summary_not_in_frictions:
    for s in sorted(summary_not_in_frictions):
        issue("HIGH", "1A Summary→Frictions", f"Sector '{s}' in Summary but missing from Frictions tabs")
if frictions_not_in_summary:
    for s in sorted(frictions_not_in_summary):
        issue("HIGH", "Frictions→1A Summary", f"Sector '{s}' in Frictions but missing from Summary")

# WC employment in summary vs staffing patterns totals
sp_emp_by_sector = defaultdict(float)
for r in range(2, ws_sp.max_row+1):
    sid = ws_sp.cell(r, 1).value
    emp = ws_sp.cell(r, 6).value
    if sid and emp:
        sp_emp_by_sector[int(sid)] += emp

for sid, data in summary_sectors.items():
    sp_total = sp_emp_by_sector.get(sid, 0)
    wc_emp = data.get("wc_emp", 0) or 0
    if abs(sp_total - wc_emp) > 1.0:
        issue("MEDIUM", "Summary↔Staffing", f"Sector {sid} ({data['name']}): Staffing sum={sp_total:.1f}K vs Summary WC_Emp={wc_emp:.1f}K")


# ========================================
# REPORT
# ========================================
print("\n" + "="*80)
print("VALIDATION REPORT")
print("="*80)

by_severity = defaultdict(list)
for sev, sheet, desc in issues:
    by_severity[sev].append((sheet, desc))

for sev in ["HIGH", "MEDIUM", "LOW"]:
    items = by_severity.get(sev, [])
    print(f"\n{'='*40}")
    print(f"  {sev}: {len(items)} issues")
    print(f"{'='*40}")
    for sheet, desc in items:
        print(f"  [{sheet}] {desc}")

print(f"\n\nTOTAL: {len(issues)} issues ({len(by_severity.get('HIGH',[]))} HIGH, {len(by_severity.get('MEDIUM',[]))} MEDIUM, {len(by_severity.get('LOW',[]))} LOW)")
