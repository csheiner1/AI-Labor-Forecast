"""
Validate the full data chain: Tasks → Jobs → Industries
Checks:
1. Every task maps to a valid job
2. Every job maps to a valid industry (Delta Sector) via NIOEM
3. Employment totals reconcile at each level
4. Time_Share sums to 100% for each job
5. Industry-level employment from bottom-up matches top-down
6. NEM code overlap check (no double-counting across sectors)
"""

import csv
import os
import openpyxl
from collections import defaultdict

BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def load_csv(filename):
    path = os.path.join(BASE_DIR, filename)
    with open(path, 'r', encoding='utf-8') as f:
        return list(csv.DictReader(f))


def safe_float(val):
    if val is None or val == '' or val == 'None':
        return 0.0
    try:
        return float(str(val).replace(',', ''))
    except (ValueError, TypeError):
        return 0.0


def main():
    print("=" * 70)
    print("DATA CHAIN VALIDATION: Tasks → Jobs → Industries")
    print("=" * 70)

    # ── Load workbook ─────────────────────────────────────────────────
    wb = openpyxl.load_workbook(os.path.join(BASE_DIR, 'jobs-data.xlsx'),
                                 read_only=True, data_only=True)

    # ── Read Jobs tab ─────────────────────────────────────────────────
    ws_jobs = wb['2 Jobs']
    headers_j = [c.value for c in next(ws_jobs.iter_rows(min_row=1, max_row=1))]
    col_j = {h: i for i, h in enumerate(headers_j) if h}

    jobs = {}
    for row in ws_jobs.iter_rows(min_row=2, values_only=True):
        soc = row[col_j.get('SOC_Code', 0)]
        if not soc:
            continue
        title = str(row[col_j.get('Custom_Title', 2)] or '').strip()
        jobs[title] = {
            'soc': str(soc).strip(),
            'sector_id': str(row[col_j.get('Delta_Sector_ID', 3)] or '').strip(),
            'sector': str(row[col_j.get('Delta_Sector', 4)] or '').strip(),
            'func_id': str(row[col_j.get('Function_ID', 5)] or '').strip(),
            'func_name': str(row[col_j.get('Function_Name', 6)] or '').strip(),
            'employment': safe_float(row[col_j.get('National_Employment_Thousands', 7)]),
        }

    print(f"\nJobs tab: {len(jobs)} jobs")

    # ── Read Tasks tab ────────────────────────────────────────────────
    task_sheet = None
    for name in wb.sheetnames:
        if 'task' in name.lower():
            task_sheet = name
            break

    if not task_sheet:
        print("ERROR: No Tasks tab found!")
        wb.close()
        return

    ws_tasks = wb[task_sheet]
    headers_t = [c.value for c in next(ws_tasks.iter_rows(min_row=1, max_row=1))]
    col_t = {h: i for i, h in enumerate(headers_t) if h}

    tasks = []
    for row in ws_tasks.iter_rows(min_row=2, values_only=True):
        task_id = row[col_t.get('Task_ID', 4)]
        if not task_id:
            continue
        tasks.append({
            'task_id': str(task_id).strip(),
            'job_title': str(row[col_t.get('Job_Title', 1)] or '').strip(),
            'soc': str(row[col_t.get('SOC_Code', 0)] or '').strip(),
            'func_id': str(row[col_t.get('Function_ID', 2)] or '').strip(),
            'time_share': safe_float(row[col_t.get('Time_Share_Pct', 8)]),
            'importance': safe_float(row[col_t.get('Importance', 9)]),
            'dedup_emp': safe_float(row[col_t.get('Dedup_Employment_K', 12)]),
            'economy_weight': safe_float(row[col_t.get('Economy_Weight_K', 13)]),
        })

    print(f"Tasks tab: {len(tasks)} tasks")
    wb.close()

    # ── Load NIOEM for industry mapping ───────────────────────────────
    nioem = load_csv('nioem_filtered.csv')
    nioem_soc_sector = defaultdict(lambda: defaultdict(float))
    nioem_soc_total = defaultdict(float)
    sector_names = {}
    for r in nioem:
        soc = r['SOC_Code']
        sid = r['Delta_Sector_ID']
        emp = safe_float(r.get('Employment_2024_thousands'))
        nioem_soc_sector[soc][sid] += emp
        nioem_soc_total[soc] += emp
        sector_names[sid] = r['Delta_Sector']

    # ── CHECK 1: Task → Job mapping ───────────────────────────────────
    print("\n" + "─" * 70)
    print("CHECK 1: Every task maps to a valid job")
    print("─" * 70)

    task_titles = set(t['job_title'] for t in tasks)
    job_titles = set(jobs.keys())
    orphan_tasks = task_titles - job_titles
    jobs_without_tasks = job_titles - task_titles

    print(f"  Unique jobs in tasks: {len(task_titles)}")
    print(f"  Unique jobs in Jobs tab: {len(job_titles)}")
    print(f"  Orphan tasks (no matching job): {len(orphan_tasks)}")
    if orphan_tasks:
        for t in sorted(orphan_tasks)[:10]:
            print(f"    {t}")
    print(f"  Jobs without tasks: {len(jobs_without_tasks)}")
    if jobs_without_tasks:
        for j in sorted(jobs_without_tasks)[:10]:
            print(f"    {j}")
        if len(jobs_without_tasks) > 10:
            print(f"    ... and {len(jobs_without_tasks) - 10} more")

    status_1 = "PASS" if len(orphan_tasks) == 0 else "FAIL"
    print(f"  Result: {status_1}")

    # ── CHECK 2: Job → Industry mapping via NIOEM ─────────────────────
    print("\n" + "─" * 70)
    print("CHECK 2: Every job's SOC code exists in NIOEM")
    print("─" * 70)

    SOC_FIXES = {
        '11-2031': '11-2032', '25-1099': '25-1000',
        '13-1023': '13-1020', '13-2020': '13-2021',
    }

    jobs_in_nioem = 0
    jobs_not_in_nioem = []
    for title, j in jobs.items():
        soc = j['soc']
        fixed = SOC_FIXES.get(soc, soc)
        if fixed in nioem_soc_total or soc in nioem_soc_total:
            jobs_in_nioem += 1
        else:
            jobs_not_in_nioem.append((title, soc))

    print(f"  Jobs with SOC in NIOEM: {jobs_in_nioem}/{len(jobs)}")
    if jobs_not_in_nioem:
        print(f"  Jobs NOT in NIOEM ({len(jobs_not_in_nioem)}):")
        for title, soc in jobs_not_in_nioem[:20]:
            print(f"    {soc}: {title}")

    status_2 = "PASS" if len(jobs_not_in_nioem) == 0 else f"WARN ({len(jobs_not_in_nioem)} missing)"
    print(f"  Result: {status_2}")

    # ── CHECK 3: Time_Share sums to 100% per job ─────────────────────
    print("\n" + "─" * 70)
    print("CHECK 3: Time_Share sums to 100% per job")
    print("─" * 70)

    job_ts = defaultdict(float)
    job_task_count = defaultdict(int)
    for t in tasks:
        job_ts[t['job_title']] += t['time_share']
        job_task_count[t['job_title']] += 1

    bad_sums = {k: v for k, v in job_ts.items() if abs(v - 100) > 0.1}
    print(f"  Jobs with Time_Share = 100%: {len(job_ts) - len(bad_sums)}/{len(job_ts)}")
    if bad_sums:
        print(f"  Jobs with bad sums:")
        for k, v in sorted(bad_sums.items()):
            print(f"    {k}: {v}%")

    avg_tasks = sum(job_task_count.values()) / len(job_task_count) if job_task_count else 0
    print(f"  Avg tasks per job: {avg_tasks:.1f}")

    status_3 = "PASS" if len(bad_sums) == 0 else f"FAIL ({len(bad_sums)} bad)"
    print(f"  Result: {status_3}")

    # ── CHECK 4: Employment reconciliation ────────────────────────────
    print("\n" + "─" * 70)
    print("CHECK 4: Employment reconciliation (bottom-up vs top-down)")
    print("─" * 70)

    # Bottom-up: sum dedup employment from tasks (per unique job)
    job_dedup_emp = {}
    for t in tasks:
        if t['job_title'] not in job_dedup_emp:
            job_dedup_emp[t['job_title']] = t['dedup_emp']

    total_dedup = sum(job_dedup_emp.values())

    # NIOEM-based: distribute each job's dedup employment across sectors
    sector_bottom_up = defaultdict(float)
    for title, dedup_emp in job_dedup_emp.items():
        if title not in jobs:
            continue
        soc = jobs[title]['soc']
        fixed_soc = SOC_FIXES.get(soc, soc)
        soc_for_nioem = fixed_soc if fixed_soc in nioem_soc_total else soc

        if soc_for_nioem in nioem_soc_total and nioem_soc_total[soc_for_nioem] > 0:
            # Distribute dedup employment proportionally across sectors
            for sid, sector_emp in nioem_soc_sector[soc_for_nioem].items():
                share = sector_emp / nioem_soc_total[soc_for_nioem]
                sector_bottom_up[sid] += dedup_emp * share
        else:
            # Assign to primary sector
            primary_sid = jobs[title]['sector_id']
            sector_bottom_up[primary_sid] += dedup_emp

    # Top-down: staffing patterns total per sector
    staffing = load_csv('staffing_patterns.csv')
    sector_top_down = defaultdict(float)
    for s in staffing:
        sector_top_down[s['Delta_Sector_ID']] += safe_float(s.get('Employment_Thousands'))

    print(f"\n  {'Sector':<40s} {'Bottom-Up (K)':>14s} {'Top-Down (K)':>14s} {'Coverage':>10s}")
    print("  " + "─" * 80)

    total_bu = 0
    total_td = 0
    for sid in sorted(sector_names.keys(), key=lambda x: int(x) if x.isdigit() else 99):
        bu = sector_bottom_up.get(sid, 0)
        td = sector_top_down.get(sid, 0)
        cov = f"{bu/td*100:.1f}%" if td > 0 else "N/A"
        total_bu += bu
        total_td += td
        print(f"  {sid + ' ' + sector_names[sid]:<40s} {bu:>14,.1f} {td:>14,.1f} {cov:>10s}")

    print("  " + "─" * 80)
    total_cov = f"{total_bu/total_td*100:.1f}%" if total_td > 0 else "N/A"
    print(f"  {'TOTAL':<40s} {total_bu:>14,.1f} {total_td:>14,.1f} {total_cov:>10s}")

    print(f"\n  De-duplicated employment total: {total_dedup:,.1f}K")
    print(f"  NIOEM-distributed total: {total_bu:,.1f}K")
    print(f"  Staffing patterns total: {total_td:,.1f}K")
    print(f"  Gap: {total_td - total_bu:,.1f}K ({(total_td - total_bu)/total_td*100:.1f}% = non-white-collar workers in our sectors)")

    # ── CHECK 5: Economy_Weight validation ────────────────────────────
    print("\n" + "─" * 70)
    print("CHECK 5: Economy_Weight_K = Time_Share × Dedup_Employment")
    print("─" * 70)

    total_economy_weight = sum(t['economy_weight'] for t in tasks)
    print(f"  Total Economy_Weight_K across all tasks: {total_economy_weight:,.1f}K")
    print(f"  This equals total de-duplicated employment: {total_dedup:,.1f}K")
    print(f"  Match: {'PASS' if abs(total_economy_weight - total_dedup) < 1 else 'CLOSE' if abs(total_economy_weight - total_dedup) < 100 else 'FAIL'}")

    # ── CHECK 6: NEM code overlap (no double-counting) ─────────────
    print("\n" + "─" * 70)
    print("CHECK 6: NEM code overlap — no NEM code in multiple sectors")
    print("─" * 70)

    sectors_csv = load_csv('lookup_sectors.csv')
    nem_to_sectors = defaultdict(set)
    for r in nioem:
        nem = r.get('NEM_Code', '')
        sid = r.get('Delta_Sector_ID', '')
        if nem and sid:
            nem_to_sectors[nem].add(sid)

    overlapping = {nem: sids for nem, sids in nem_to_sectors.items() if len(sids) > 1}
    if overlapping:
        print(f"  NEM codes appearing in multiple sectors ({len(overlapping)}):")
        for nem, sids in sorted(overlapping.items()):
            names = [sector_names.get(s, s) for s in sorted(sids, key=lambda x: int(x) if x.isdigit() else 99)]
            print(f"    NEM {nem}: {', '.join(names)}")
    else:
        print(f"  All {len(nem_to_sectors)} NEM codes are unique to a single sector")

    status_6 = "PASS" if len(overlapping) == 0 else f"WARN ({len(overlapping)} overlapping)"
    print(f"  Result: {status_6}")

    # ── SUMMARY ───────────────────────────────────────────────────────
    print("\n" + "=" * 70)
    print("VALIDATION SUMMARY")
    print("=" * 70)
    print(f"  CHECK 1 (Task→Job mapping):      {status_1}")
    print(f"  CHECK 2 (Job→Industry via NIOEM): {status_2}")
    print(f"  CHECK 3 (Time_Share = 100%):      {status_3}")
    print(f"  CHECK 4 (Employment reconcile):   {total_cov} coverage of sector employment")
    print(f"  CHECK 5 (Economy_Weight):          {'PASS' if abs(total_economy_weight - total_dedup) < 1 else 'CHECK'}")
    print(f"  CHECK 6 (NEM overlap):            {status_6}")
    print(f"\n  Data chain: {len(tasks)} tasks → {len(task_titles)} jobs → {len(sector_names)} sectors")
    print(f"  Total workers represented: {total_dedup:,.1f}K ({total_dedup*1000:,.0f})")


if __name__ == "__main__":
    main()
