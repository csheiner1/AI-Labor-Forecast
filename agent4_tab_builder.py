"""
Agent 4: Tab Builder
Assembles the final Excel workbook with all tabs.

Inputs: All CSVs from Agents 1–3
Output: jobs-data.xlsx
"""

import csv
import os
from collections import defaultdict

import openpyxl

try:
    import xlsxwriter
except ImportError:
    print("Installing xlsxwriter...")
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "xlsxwriter"])
    import xlsxwriter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(BASE_DIR, "jobs-data.xlsx")


def load_csv(filename):
    path = os.path.join(BASE_DIR, filename)
    with open(path, 'r', encoding='utf-8') as f:
        return list(csv.DictReader(f))


def safe_float(val):
    if val is None or val == '' or val == 'None':
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def main():
    print("=" * 60)
    print("AGENT 4: Tab Builder")
    print("=" * 60)

    # ── Load all data ───────────────────────────────────────────────────
    print("\n--- Loading all intermediate files ---")
    sectors = load_csv("lookup_sectors.csv")
    functions_lookup = load_csv("lookup_functions.csv")
    jobs_lookup = load_csv("lookup_jobs.csv")
    occupations = load_csv("occupations_master.csv")
    nioem_filtered = load_csv("nioem_filtered.csv")
    staffing = load_csv("staffing_patterns.csv")
    func_dist = load_csv("function_distribution.csv")
    matrix_raw = load_csv("matrix_industry_function.csv")
    matrix_norm = load_csv("matrix_industry_function_normalized.csv")

    # Build lookups
    occ_by_soc = {}
    for o in occupations:
        occ_by_soc[o['SOC_Code']] = o

    # Group nioem by NEM_Code for industry-level aggregation
    industry_employment = defaultdict(float)
    industry_employment_2034 = defaultdict(float)
    for r in nioem_filtered:
        key = r['NEM_Code']
        industry_employment[key] += safe_float(r.get('Employment_2024_thousands')) or 0
        industry_employment_2034[key] += safe_float(r.get('Employment_2034_thousands')) or 0

    # Get sector names/IDs
    sector_names = {}
    for s in sectors:
        sector_names[s['Delta_Sector_ID']] = s['Delta_Sector']

    # SOC → Function mapping
    soc_to_func = {}
    for f in functions_lookup:
        soc_to_func[f['SOC_Code']] = f

    print(f"  Loaded {len(sectors)} sectors, {len(functions_lookup)} functions, {len(jobs_lookup)} jobs")
    print(f"  Loaded {len(occupations)} occupations, {len(nioem_filtered)} NIOEM filtered rows")

    # ── Build SOC → industry employment distribution (for subsegment columns) ─
    soc_industry_emp = defaultdict(lambda: defaultdict(float))
    for r in nioem_filtered:
        soc = r['SOC_Code']
        sid = r['Delta_Sector_ID']
        emp = safe_float(r.get('Employment_2024_thousands')) or 0
        soc_industry_emp[soc][sid] += emp

    def get_top_industries(soc_code, n=3):
        """Return top N industries for a SOC code as list of (sector_name, pct) strings."""
        dist = soc_industry_emp.get(soc_code, {})
        if not dist:
            return [''] * n
        total = sum(dist.values())
        if total <= 0:
            return [''] * n
        ranked = sorted(dist.items(), key=lambda x: x[1], reverse=True)
        results = []
        for sid, emp in ranked[:n]:
            pct = emp / total * 100
            name = sector_names.get(sid, f'Sector {sid}')
            results.append(f"{name} ({pct:.0f}%)")
        while len(results) < n:
            results.append('')
        return results

    # ── Preserve existing Tasks tab before overwriting ──────────────────
    existing_tasks_headers = None
    existing_tasks_rows = []
    if os.path.exists(OUTPUT_FILE):
        try:
            wb_old = openpyxl.load_workbook(OUTPUT_FILE, read_only=True, data_only=True)
            task_sheet = None
            for name in wb_old.sheetnames:
                if 'task' in name.lower():
                    task_sheet = name
                    break
            if task_sheet:
                ws_t = wb_old[task_sheet]
                all_rows = list(ws_t.iter_rows(values_only=True))
                if len(all_rows) > 1:  # has header + data
                    existing_tasks_headers = [str(h) if h else '' for h in all_rows[0]]
                    existing_tasks_rows = all_rows[1:]
                    print(f"\n  Preserved '{task_sheet}' tab: {len(existing_tasks_rows)} tasks")
                else:
                    print(f"\n  Found '{task_sheet}' tab but it's empty — skipping preservation")
            else:
                print("\n  No existing Tasks tab to preserve")
            wb_old.close()
        except Exception as e:
            print(f"\n  Could not read existing workbook: {e}")

    # ── Create workbook ─────────────────────────────────────────────────
    print("\n--- Creating Excel workbook ---")
    wb = xlsxwriter.Workbook(OUTPUT_FILE)

    # ── Formats ─────────────────────────────────────────────────────────
    fmt_header = wb.add_format({
        'bold': True, 'bg_color': '#4472C4', 'font_color': 'white',
        'border': 1, 'text_wrap': True, 'valign': 'top',
    })
    fmt_number = wb.add_format({'num_format': '#,##0.0', 'border': 1})
    fmt_int = wb.add_format({'num_format': '#,##0', 'border': 1})
    fmt_dollar = wb.add_format({'num_format': '$#,##0', 'border': 1})
    fmt_pct = wb.add_format({'num_format': '0.0%', 'border': 1})
    fmt_pct_display = wb.add_format({'num_format': '0.0', 'border': 1})
    fmt_text = wb.add_format({'border': 1, 'text_wrap': True, 'valign': 'top'})
    fmt_blank_score = wb.add_format({
        'bg_color': '#FFF2CC', 'border': 1, 'num_format': '0.0',
    })
    fmt_section = wb.add_format({
        'bold': True, 'bg_color': '#D9E2F3', 'border': 1,
    })
    fmt_readme = wb.add_format({
        'text_wrap': True, 'valign': 'top', 'font_size': 11,
    })
    fmt_readme_header = wb.add_format({
        'bold': True, 'font_size': 14, 'bottom': 2,
    })
    fmt_readme_subheader = wb.add_format({
        'bold': True, 'font_size': 12,
    })

    # ── Tab 1A: Industries ──────────────────────────────────────────────
    print("  Building Tab 1A: Industries...")
    ws = wb.add_worksheet("1A Industries")
    ws.freeze_panes(1, 0)

    headers_1a = [
        'NAICS_Code', 'NAICS_Title', 'Sector_ID', 'Sector',
        'Sub_Industry', 'Mapping_Type', 'NEM_Code',
        'Employment_2024 (Thousands)', 'Employment_2034 (Thousands)',
        'Projected_Change_Pct', 'Automatability_Score', 'Weight',
    ]
    for c, h in enumerate(headers_1a):
        ws.write(0, c, h, fmt_header)

    row_num = 1
    for s in sectors:
        # Find the matching NEM code and employment
        naics = s['NAICS_Code']
        nem_code = ''
        emp = 0
        emp_2034 = 0
        for r in nioem_filtered:
            if r.get('NAICS_Code') == naics:
                nem_code = r.get('NEM_Code', '')
                break
        emp = industry_employment.get(nem_code, 0)
        emp_2034 = industry_employment_2034.get(nem_code, 0)
        change_pct = ((emp_2034 - emp) / emp * 100) if emp > 0 else None

        ws.write(row_num, 0, naics, fmt_text)
        ws.write(row_num, 1, s['NAICS_Title'], fmt_text)
        ws.write(row_num, 2, int(s['Delta_Sector_ID']), fmt_int)
        ws.write(row_num, 3, s['Delta_Sector'], fmt_text)
        ws.write(row_num, 4, s['Delta_Sub_Industry'], fmt_text)
        ws.write(row_num, 5, s['Mapping_Type'], fmt_text)
        ws.write(row_num, 6, nem_code, fmt_text)
        ws.write(row_num, 7, emp, fmt_number)
        ws.write(row_num, 8, emp_2034, fmt_number)
        if change_pct is not None:
            ws.write(row_num, 9, change_pct, fmt_pct_display)
        else:
            ws.write(row_num, 9, '', fmt_text)
        ws.write(row_num, 10, '', fmt_blank_score)
        ws.write(row_num, 11, '', fmt_blank_score)
        row_num += 1

    # Auto-fit columns
    col_widths_1a = [12, 45, 8, 35, 35, 12, 10, 15, 15, 12, 15, 8]
    for c, w in enumerate(col_widths_1a):
        ws.set_column(c, c, w)
    tab_1a_rows = row_num - 1

    # ── Tab 1A Summary: Sector Rollup ───────────────────────────────────
    print("  Building Tab 1A Summary: Sector Rollup...")
    ws_sum = wb.add_worksheet("1A Summary")
    ws_sum.freeze_panes(1, 0)

    headers_sum = [
        'Sector_ID', 'Sector', 'Num_NAICS_Codes',
        'Employment_2024 (Thousands)', 'Avg_Median_Wage',
        'Automatability_Score', 'Weight',
    ]
    for c, h in enumerate(headers_sum):
        ws_sum.write(0, c, h, fmt_header)

    # Aggregate by sector
    sector_agg = defaultdict(lambda: {'naics_count': 0, 'total_emp': 0, 'wages': [], 'name': ''})
    for s in sectors:
        sid = s['Delta_Sector_ID']
        sector_agg[sid]['naics_count'] += 1
        sector_agg[sid]['name'] = s['Delta_Sector']
        naics = s['NAICS_Code']
        for r in nioem_filtered:
            if r.get('NAICS_Code') == naics:
                nem = r.get('NEM_Code', '')
                break
        else:
            nem = ''
        sector_agg[sid]['total_emp'] += industry_employment.get(nem, 0)

    # Compute average wage per sector from staffing patterns
    for sid_str, agg in sector_agg.items():
        wages = []
        for sp in staffing:
            if sp['Delta_Sector_ID'] == sid_str:
                occ = occ_by_soc.get(sp['SOC_Code'])
                if occ:
                    w = safe_float(occ.get('Median_Annual_Wage'))
                    e = safe_float(sp.get('Employment_Thousands'))
                    if w and e and e > 0:
                        wages.append((w, e))
        if wages:
            total_w = sum(w * e for w, e in wages)
            total_e = sum(e for _, e in wages)
            agg['avg_wage'] = total_w / total_e if total_e > 0 else None
        else:
            agg['avg_wage'] = None

    row_num = 1
    for sid in sorted(sector_agg.keys(), key=lambda x: int(x)):
        agg = sector_agg[sid]
        ws_sum.write(row_num, 0, int(sid), fmt_int)
        ws_sum.write(row_num, 1, agg['name'], fmt_text)
        ws_sum.write(row_num, 2, agg['naics_count'], fmt_int)
        ws_sum.write(row_num, 3, agg['total_emp'], fmt_number)
        if agg.get('avg_wage'):
            ws_sum.write(row_num, 4, agg['avg_wage'], fmt_dollar)
        else:
            ws_sum.write(row_num, 4, '', fmt_text)
        ws_sum.write(row_num, 5, '', fmt_blank_score)
        ws_sum.write(row_num, 6, '', fmt_blank_score)
        row_num += 1

    col_widths_sum = [8, 40, 12, 18, 15, 15, 8]
    for c, w in enumerate(col_widths_sum):
        ws_sum.set_column(c, c, w)
    tab_sum_rows = row_num - 1

    # ── Tab 1B: Functions ───────────────────────────────────────────────
    print("  Building Tab 1B: Functions...")
    ws_fn = wb.add_worksheet("1B Functions")
    ws_fn.freeze_panes(1, 0)

    headers_fn = [
        'Function_ID', 'Function_Name', 'SOC_Code', 'SOC_Title',
        'Employment_2024 (Thousands)', 'Median_Wage', 'Education',
        'Shared', 'Automatability_Score', 'Weight',
    ]
    for c, h in enumerate(headers_fn):
        ws_fn.write(0, c, h, fmt_header)

    row_num = 1
    for f in functions_lookup:
        occ = occ_by_soc.get(f['SOC_Code'], {})
        ws_fn.write(row_num, 0, int(f['Function_ID']), fmt_int)
        ws_fn.write(row_num, 1, f['Function_Name'], fmt_text)
        ws_fn.write(row_num, 2, f['SOC_Code'], fmt_text)
        ws_fn.write(row_num, 3, f['SOC_Title'], fmt_text)
        emp = safe_float(occ.get('Employment_2024'))
        ws_fn.write(row_num, 4, emp if emp else '', fmt_number)
        wage = safe_float(occ.get('Median_Annual_Wage'))
        ws_fn.write(row_num, 5, wage if wage else '', fmt_dollar)
        ws_fn.write(row_num, 6, occ.get('Typical_Education', ''), fmt_text)
        ws_fn.write(row_num, 7, f.get('Shared', 'False'), fmt_text)
        ws_fn.write(row_num, 8, '', fmt_blank_score)
        ws_fn.write(row_num, 9, '', fmt_blank_score)
        row_num += 1

    col_widths_fn = [8, 40, 10, 55, 18, 12, 30, 8, 15, 8]
    for c, w in enumerate(col_widths_fn):
        ws_fn.set_column(c, c, w)
    tab_fn_rows = row_num - 1

    # ── Tab 2: Jobs ─────────────────────────────────────────────────────
    print("  Building Tab 2: Jobs...")
    ws_jobs = wb.add_worksheet("2 Jobs")
    ws_jobs.freeze_panes(1, 0)

    headers_jobs = [
        'SOC_Code', 'SOC_Title', 'Custom_Title', 'Sector_ID', 'Sector',
        'Function_ID', 'Function_Name', 'Employment_2024 (Thousands)',
        'Median_Wage', 'Education', 'Work_Experience', 'OJT',
        'Projected_Change_Pct', 'Mapping_Notes',
        'Primary_Industry', 'Industry_2', 'Industry_3',
        'Automatability_Score', 'Weight',
    ]
    for c, h in enumerate(headers_jobs):
        ws_jobs.write(0, c, h, fmt_header)

    row_num = 1
    socs_covered = set()
    for j in jobs_lookup:
        soc = j['SOC_Code']
        occ = occ_by_soc.get(soc, {})
        func = soc_to_func.get(soc, {})

        sector_id = j['Delta_Sector_ID']
        sector_name = sector_names.get(sector_id, '')

        # Compute top 3 industries from NIOEM distribution
        top3 = get_top_industries(soc)

        ws_jobs.write(row_num, 0, soc, fmt_text)
        ws_jobs.write(row_num, 1, j['SOC_Title'], fmt_text)
        ws_jobs.write(row_num, 2, j['Custom_Title'], fmt_text)
        ws_jobs.write(row_num, 3, str(sector_id), fmt_text)
        ws_jobs.write(row_num, 4, sector_name, fmt_text)
        ws_jobs.write(row_num, 5, func.get('Function_ID', ''), fmt_text)
        ws_jobs.write(row_num, 6, func.get('Function_Name', ''), fmt_text)
        emp = safe_float(occ.get('Employment_2024'))
        ws_jobs.write(row_num, 7, emp if emp else '', fmt_number)
        wage = safe_float(occ.get('Median_Annual_Wage'))
        ws_jobs.write(row_num, 8, wage if wage else '', fmt_dollar)
        ws_jobs.write(row_num, 9, occ.get('Typical_Education', ''), fmt_text)
        ws_jobs.write(row_num, 10, occ.get('Work_Experience', ''), fmt_text)
        ws_jobs.write(row_num, 11, occ.get('OJT_Training', ''), fmt_text)
        change = safe_float(occ.get('Change_Percent'))
        ws_jobs.write(row_num, 12, change if change else '', fmt_pct_display)
        ws_jobs.write(row_num, 13, j.get('Mapping_Notes', ''), fmt_text)
        ws_jobs.write(row_num, 14, top3[0], fmt_text)
        ws_jobs.write(row_num, 15, top3[1], fmt_text)
        ws_jobs.write(row_num, 16, top3[2], fmt_text)
        ws_jobs.write(row_num, 17, '', fmt_blank_score)
        ws_jobs.write(row_num, 18, '', fmt_blank_score)
        row_num += 1
        socs_covered.add(soc)

    col_widths_jobs = [10, 55, 30, 10, 35, 8, 35, 18, 12, 25, 18, 20, 12, 20, 28, 28, 28, 15, 8]
    for c, w in enumerate(col_widths_jobs):
        ws_jobs.set_column(c, c, w)
    tab_jobs_rows = row_num - 1

    # ── Matrix Tab ──────────────────────────────────────────────────────
    print("  Building Matrix tab...")
    ws_matrix = wb.add_worksheet("Matrix")
    ws_matrix.freeze_panes(2, 2)

    # Write headers from the matrix CSV
    if matrix_raw:
        headers_m = list(matrix_raw[0].keys())
        # Label row above headers with units
        ws_matrix.write(0, 0, "EMPLOYMENT (Thousands)", fmt_section)
        for c, h in enumerate(headers_m):
            # Clean up column names for display
            display_h = h.replace('Sector_', '').replace('_', ' ')
            ws_matrix.write(1, c, display_h, fmt_header)

        for r_idx, row in enumerate(matrix_raw):
            for c_idx, h in enumerate(headers_m):
                val = row.get(h, '')
                num = safe_float(val)
                if c_idx < 2:
                    ws_matrix.write(r_idx + 2, c_idx, val, fmt_text)
                elif num is not None:
                    ws_matrix.write(r_idx + 2, c_idx, num, fmt_number)
                else:
                    ws_matrix.write(r_idx + 2, c_idx, val, fmt_text)

        # Write normalized matrix below
        norm_start = len(matrix_raw) + 4
        ws_matrix.write(norm_start - 1, 0, "NORMALIZED (% of row total)", fmt_section)
        for c, h in enumerate(headers_m):
            display_h = h.replace('Sector_', '').replace('_', ' ')
            ws_matrix.write(norm_start, c, display_h, fmt_header)

        for r_idx, row in enumerate(matrix_norm):
            for c_idx, h in enumerate(headers_m):
                val = row.get(h, '')
                num = safe_float(val)
                if c_idx < 2:
                    ws_matrix.write(norm_start + r_idx + 1, c_idx, val, fmt_text)
                elif num is not None:
                    ws_matrix.write(norm_start + r_idx + 1, c_idx, num, fmt_pct_display)
                else:
                    ws_matrix.write(norm_start + r_idx + 1, c_idx, val, fmt_text)

        ws_matrix.set_column(0, 0, 8)
        ws_matrix.set_column(1, 1, 35)
        for c in range(2, len(headers_m)):
            ws_matrix.set_column(c, c, 14)
    tab_matrix_rows = len(matrix_raw) + len(matrix_norm) + 3

    # ── Staffing Patterns Tab ───────────────────────────────────────────
    print("  Building Staffing Patterns tab...")
    ws_staff = wb.add_worksheet("Staffing Patterns")
    ws_staff.freeze_panes(1, 0)

    headers_sp = ['Sector_ID', 'Sector', 'SOC_Code', 'SOC_Title',
                   'Employment (Thousands)', 'Staffing_Share_Pct']
    for c, h in enumerate(headers_sp):
        ws_staff.write(0, c, h, fmt_header)

    for r_idx, r in enumerate(staffing):
        ws_staff.write(r_idx + 1, 0, r['Delta_Sector_ID'], fmt_text)
        ws_staff.write(r_idx + 1, 1, r['Delta_Sector'], fmt_text)
        ws_staff.write(r_idx + 1, 2, r['SOC_Code'], fmt_text)
        ws_staff.write(r_idx + 1, 3, r['SOC_Title'], fmt_text)
        emp = safe_float(r.get('Employment_Thousands'))
        ws_staff.write(r_idx + 1, 4, emp if emp else 0, fmt_number)
        pct = safe_float(r.get('Staffing_Share_Pct'))
        ws_staff.write(r_idx + 1, 5, pct if pct else 0, fmt_pct_display)

    col_widths_sp = [8, 35, 10, 55, 15, 12]
    for c, w in enumerate(col_widths_sp):
        ws_staff.set_column(c, c, w)

    # ── Tab 2B: Job_Industry pivot table ─────────────────────────────
    print("  Building Tab 2B: Job_Industry pivot...")
    ws_ji = wb.add_worksheet("2B Job_Industry")
    ws_ji.freeze_panes(2, 2)

    # Get sorted unique sector IDs and all unique SOC codes
    unique_sids = sorted(set(s['Delta_Sector_ID'] for s in sectors), key=lambda x: int(x))
    all_socs_in_nioem = sorted(soc_industry_emp.keys())

    # Headers: SOC_Code, SOC_Title, then one column per sector
    ji_headers = ['SOC_Code', 'SOC_Title'] + [sector_names.get(sid, f'Sector {sid}') for sid in unique_sids]
    ws_ji.write(0, 0, "Employment by Industry (Thousands)", fmt_section)
    for c, h in enumerate(ji_headers):
        ws_ji.write(1, c, h, fmt_header)

    row_num_ji = 2
    for soc in all_socs_in_nioem:
        occ = occ_by_soc.get(soc, {})
        title = occ.get('SOC_Title', soc)
        ws_ji.write(row_num_ji, 0, soc, fmt_text)
        ws_ji.write(row_num_ji, 1, title, fmt_text)
        for c_idx, sid in enumerate(unique_sids):
            emp = soc_industry_emp.get(soc, {}).get(sid, 0)
            if emp > 0:
                ws_ji.write(row_num_ji, c_idx + 2, emp, fmt_number)
            else:
                ws_ji.write(row_num_ji, c_idx + 2, '', fmt_text)
        row_num_ji += 1

    ws_ji.set_column(0, 0, 10)
    ws_ji.set_column(1, 1, 50)
    for c in range(2, len(ji_headers)):
        ws_ji.set_column(c, c, 14)
    tab_ji_rows = row_num_ji - 1

    # ── Lookup tabs ─────────────────────────────────────────────────────
    print("  Building Lookup tabs...")

    # Lookup_Sectors
    ws_ls = wb.add_worksheet("Lookup_Sectors")
    ws_ls.freeze_panes(1, 0)
    ls_headers = ['Sector_ID', 'Sector', 'Sub_Industry', 'NAICS_Code', 'NAICS_Title', 'Mapping_Type']
    for c, h in enumerate(ls_headers):
        ws_ls.write(0, c, h, fmt_header)
    ls_csv_keys = ['Delta_Sector_ID', 'Delta_Sector', 'Delta_Sub_Industry', 'NAICS_Code', 'NAICS_Title', 'Mapping_Type']
    for r_idx, s in enumerate(sectors):
        for c, csv_key in enumerate(ls_csv_keys):
            ws_ls.write(r_idx + 1, c, s.get(csv_key, ''), fmt_text)
    for c, w in enumerate([8, 35, 35, 12, 50, 12]):
        ws_ls.set_column(c, c, w)

    # Lookup_Functions
    ws_lf = wb.add_worksheet("Lookup_Functions")
    ws_lf.freeze_panes(1, 0)
    lf_headers = ['Function_ID', 'Function_Name', 'SOC_Code', 'SOC_Title', 'Shared']
    for c, h in enumerate(lf_headers):
        ws_lf.write(0, c, h, fmt_header)
    for r_idx, f in enumerate(functions_lookup):
        for c, h in enumerate(lf_headers):
            ws_lf.write(r_idx + 1, c, f.get(h, ''), fmt_text)
    for c, w in enumerate([8, 40, 10, 55, 8]):
        ws_lf.set_column(c, c, w)

    # Lookup_Jobs
    ws_lj = wb.add_worksheet("Lookup_Jobs")
    ws_lj.freeze_panes(1, 0)
    lj_headers = ['Custom_Title', 'Sector_ID', 'SOC_Code', 'SOC_Title', 'Mapping_Notes']
    for c, h in enumerate(lj_headers):
        ws_lj.write(0, c, h, fmt_header)
    lj_csv_keys = ['Custom_Title', 'Delta_Sector_ID', 'SOC_Code', 'SOC_Title', 'Mapping_Notes']
    for r_idx, j in enumerate(jobs_lookup):
        for c, csv_key in enumerate(lj_csv_keys):
            ws_lj.write(r_idx + 1, c, j.get(csv_key, ''), fmt_text)
    for c, w in enumerate([30, 10, 10, 55, 20]):
        ws_lj.set_column(c, c, w)

    # ── Restore Tasks tab (if preserved) ─────────────────────────────
    tab_tasks_rows = 0
    if existing_tasks_headers and existing_tasks_rows:
        print("  Restoring preserved Tasks tab...")
        ws_tasks = wb.add_worksheet("3 Tasks")
        ws_tasks.freeze_panes(1, 0)
        for c, h in enumerate(existing_tasks_headers):
            ws_tasks.write(0, c, h, fmt_header)
        for r_idx, row in enumerate(existing_tasks_rows):
            for c_idx, val in enumerate(row):
                if val is None:
                    ws_tasks.write(r_idx + 1, c_idx, '', fmt_text)
                elif isinstance(val, (int, float)):
                    ws_tasks.write(r_idx + 1, c_idx, val, fmt_number)
                else:
                    ws_tasks.write(r_idx + 1, c_idx, str(val), fmt_text)
        tab_tasks_rows = len(existing_tasks_rows)
        # Auto-size columns
        for c in range(len(existing_tasks_headers)):
            max_len = len(existing_tasks_headers[c])
            ws_tasks.set_column(c, c, min(max_len + 4, 40))

    # ── ReadMe Tab ──────────────────────────────────────────────────────
    print("  Building ReadMe tab...")
    ws_rm = wb.add_worksheet("ReadMe")
    ws_rm.set_column(0, 0, 100)

    readme_lines = [
        ("AI Labor Analysis Dataset", fmt_readme_header),
        ("", fmt_readme),
        ("DATA SOURCES", fmt_readme_subheader),
        ("- BLS Employment Projections 2024-2034 (occupation.xlsx, Tables 1.2/1.8/1.9)", fmt_readme),
        ("- National Industry-Occupation Employment Matrix (NIOEM) via BLS API", fmt_readme),
        ("- Employment figures are in THOUSANDS (152.3 = 152,300 workers)", fmt_readme),
        ("- Wage data: BLS 2024 median annual wages. Downloaded March 2026.", fmt_readme),
        ("", fmt_readme),
        ("ARCHITECTURE", fmt_readme_subheader),
        ("Two independent dimensions, bridged by the BLS NIOEM:", fmt_readme),
        ("- 20 Industries (NAICS-based) = where a company operates", fmt_readme),
        ("- 17 Functions (SOC-based) = what a worker does", fmt_readme),
        ("Cross-industry roles (CEO, HR Manager, etc.) are assigned to their", fmt_readme),
        ("highest-employment industry. The subsegment columns (Primary_Industry,", fmt_readme),
        ("Industry_2, Industry_3) and the 2B tab show the full distribution.", fmt_readme),
        ("", fmt_readme),
        ("TABS", fmt_readme_subheader),
        ("1A Industries    — 76 NAICS codes with employment, rolled up to 20 sectors", fmt_readme),
        ("1A Summary       — Sector-level rollup with avg wages", fmt_readme),
        ("1B Functions     — 393 SOC codes across 17 business functions", fmt_readme),
        ("2 Jobs           — 502 job titles with BLS data + top-3 industry subsegments", fmt_readme),
        ("2B Job_Industry  — Full SOC x 20-Industry employment pivot table", fmt_readme),
        ("3 Tasks          — ~3,800 tasks with Time_Share and Economy_Weight", fmt_readme),
        ("Matrix           — 20 Industry x 17 Function cross-tabulation (raw + normalized)", fmt_readme),
        ("Staffing Patterns— Top occupations by share within each sector", fmt_readme),
        ("Lookup_*         — Reference tables for the taxonomy", fmt_readme),
        ("", fmt_readme),
        ("KEY COLUMNS", fmt_readme_subheader),
        ("Employment_Thousands  — BLS employment in thousands", fmt_readme),
        ("Median_Wage           — BLS median annual wage ($)", fmt_readme),
        ("Projected_Change_Pct  — Projected 2024-2034 change (%)", fmt_readme),
        ("Primary_Industry      — Highest-employment industry for this SOC, with %", fmt_readme),
        ("Industry_2 / 3        — Second and third highest industries, with %", fmt_readme),
        ("Automatability_Score  — [BLANK] Yellow columns are placeholders for", fmt_readme),
        ("Weight                — [BLANK] AI displacement/augmentation scores", fmt_readme),
        ("", fmt_readme),
        ("KNOWN LIMITATIONS", fmt_readme_subheader),
        ("- Some NAICS codes share a single BLS aggregate ('Composite' mapping type)", fmt_readme),
        ("- Education (Sector 10): NIOEM covers private sector only. ~85-90% of", fmt_readme),
        ("  teachers work in public schools, so national employment figures overstate", fmt_readme),
        ("  this sector in reconciliation. Sector-level NIOEM data is accurate.", fmt_readme),
        ("- Sectors 15-20 (Manufacturing, Retail, Construction, Transportation,", fmt_readme),
        ("  Wholesale, Accommodation) are heavily blue-collar. White-collar SOC", fmt_readme),
        ("  coverage is 30-50% of total sector employment — this is expected.", fmt_readme),
        ("- Government uses BLS aggregate NEM codes (999100/999200), not NAICS", fmt_readme),
        ("- SOC 13-1111 (Management Analysts) shared between Functions 8 and 10", fmt_readme),
    ]

    for r_idx, (text, fmt) in enumerate(readme_lines):
        ws_rm.write(r_idx, 0, text, fmt)

    # ── Close workbook ──────────────────────────────────────────────────
    wb.close()

    # ── Final Summary ───────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("FINAL SUMMARY")
    print("=" * 60)
    print(f"\nOutput: {OUTPUT_FILE}")
    print(f"\nTabs and row counts:")
    print(f"  1A Industries:      {tab_1a_rows} rows")
    print(f"  1A Summary:         {tab_sum_rows} rows")
    print(f"  1B Functions:       {tab_fn_rows} rows")
    print(f"  2 Jobs:             {tab_jobs_rows} rows")
    print(f"  2B Job_Industry:   {tab_ji_rows} rows")
    print(f"  3 Tasks:           {tab_tasks_rows} rows {'(preserved)' if tab_tasks_rows > 0 else '(empty — run task_generator_v2.py)'}")
    print(f"  Matrix:             {tab_matrix_rows} rows")
    print(f"  Staffing Patterns:  {len(staffing)} rows")
    print(f"  Lookup_Sectors:     {len(sectors)} rows")
    print(f"  Lookup_Functions:   {len(functions_lookup)} rows")
    print(f"  Lookup_Jobs:        {len(jobs_lookup)} rows")
    print(f"  ReadMe:             Documentation")
    print(f"\nTotal unique SOC codes in Jobs tab: {len(socs_covered)}")

    # Data completeness
    total_jobs = len(jobs_lookup)
    with_wage = sum(1 for j in jobs_lookup if safe_float(occ_by_soc.get(j['SOC_Code'], {}).get('Median_Annual_Wage')))
    with_edu = sum(1 for j in jobs_lookup if occ_by_soc.get(j['SOC_Code'], {}).get('Typical_Education') not in (None, '', 'None'))
    with_proj = sum(1 for j in jobs_lookup if safe_float(occ_by_soc.get(j['SOC_Code'], {}).get('Change_Percent')))
    print(f"\nData completeness (Jobs tab):")
    print(f"  With wages:       {with_wage}/{total_jobs} ({with_wage/total_jobs*100:.0f}%)")
    print(f"  With education:   {with_edu}/{total_jobs} ({with_edu/total_jobs*100:.0f}%)")
    print(f"  With projections: {with_proj}/{total_jobs} ({with_proj/total_jobs*100:.0f}%)")

    # SOC codes from lookup_jobs not found in BLS
    missing_socs = []
    for j in jobs_lookup:
        if j['SOC_Code'] not in occ_by_soc:
            missing_socs.append((j['Custom_Title'], j['SOC_Code']))
    if missing_socs:
        print(f"\nSOC codes from lookup_jobs NOT found in BLS data ({len(missing_socs)}):")
        for title, soc in missing_socs:
            print(f"  {soc}: {title}")
    else:
        print(f"\nAll SOC codes from lookup_jobs found in BLS data.")

    print("\nAgent 4 complete.")


if __name__ == "__main__":
    main()
