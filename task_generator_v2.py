"""
Task Generator v2: Time-Share Weighted
Generates O*NET-style tasks with Time_Share (% of workday) as the primary weight.
Also de-duplicates employment across jobs sharing the same SOC code.

Usage:
  python task_generator_v2.py --input jobs-data.xlsx --sheet "2 Jobs"
  python task_generator_v2.py --input jobs-data.xlsx --sheet "2 Jobs" --limit 5
"""

import argparse
import csv
import json
import os
import sys
import time
import threading
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from anthropic import Anthropic

client = Anthropic()
MODEL = "claude-sonnet-4-20250514"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------------------
# SOC CODE FIXES: Map incorrect/missing SOC codes to valid BLS codes
# ---------------------------------------------------------------------------
SOC_FIXES = {
    '11-2031': '11-2032',  # PR Managers → BLS uses 11-2032
    '25-1099': '25-1000',  # Postsecondary Teachers All Other → use summary
    '13-1023': '13-1020',  # Purchasing Agents → BLS uses summary 13-1020
    '13-2020': '13-2021',  # Property Appraisers → BLS uses 13-2021
}

# ---------------------------------------------------------------------------
# STEP 0: DE-DUPLICATE EMPLOYMENT
# ---------------------------------------------------------------------------
def build_employment_table(jobs: list[dict], occ_lookup: dict) -> dict:
    """
    For each job, compute de-duplicated employment:
      De-dup Employment = National SOC Employment / (# of custom titles using that SOC)

    Returns dict keyed by Custom_Title → employment in thousands.
    """
    # Count how many custom titles share each SOC code
    soc_counts = defaultdict(int)
    soc_titles = defaultdict(list)
    for j in jobs:
        soc = j['soc_code']
        soc_counts[soc] += 1
        soc_titles[soc].append(j['title'])

    employment = {}
    for j in jobs:
        soc = j['soc_code']
        fixed_soc = SOC_FIXES.get(soc, soc)

        occ = occ_lookup.get(fixed_soc) or occ_lookup.get(soc)
        if occ:
            nat_emp = safe_float(occ.get('Employment_2024'))
            if nat_emp and nat_emp > 0:
                employment[j['title']] = {
                    'national_employment': nat_emp,
                    'dedup_employment': round(nat_emp / soc_counts[soc], 2),
                    'soc_code': soc,
                    'fixed_soc': fixed_soc,
                    'shared_with': soc_counts[soc],
                    'median_wage': safe_float(occ.get('Median_Annual_Wage')),
                    'education': occ.get('Typical_Education', ''),
                    'change_pct': safe_float(occ.get('Change_Percent')),
                }
            else:
                employment[j['title']] = _empty_emp(soc, fixed_soc, soc_counts[soc])
        else:
            employment[j['title']] = _empty_emp(soc, fixed_soc, soc_counts[soc])

    return employment


def _empty_emp(soc, fixed_soc, shared):
    return {
        'national_employment': None,
        'dedup_employment': None,
        'soc_code': soc,
        'fixed_soc': fixed_soc,
        'shared_with': shared,
        'median_wage': None,
        'education': '',
        'change_pct': None,
    }


def safe_float(val):
    if val is None or val == '' or val == 'None':
        return None
    try:
        return float(str(val).replace(',', ''))
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# AGENT 1: TASK GENERATOR (v2 — with Time_Share)
# ---------------------------------------------------------------------------
GENERATOR_SYSTEM = """You are an occupational analyst specializing in job task decomposition.
You produce task statements in O*NET format: verb-object-context, one sentence each.

Rules:
- Generate exactly 6-8 tasks per occupation
- Tasks should collectively cover ~95% of the role's working time
- Use active verbs (analyze, develop, prepare, coordinate, evaluate, implement, manage)
- Each task must be distinct — no overlapping responsibilities
- Include a mix of strategic/analytical and operational/communication tasks
- time_share values MUST sum to exactly 100 across all tasks
- time_share represents the percentage of a typical workday/workweek spent on this task
- A task done frequently but briefly (e.g., checking email) may have lower time_share
  than a task done less frequently but for extended periods (e.g., building financial models)

Respond ONLY with valid JSON, no markdown fences, no preamble."""

def generate_tasks(job_title: str, soc_code: str, function_name: str, sector: str) -> list[dict]:
    """Agent 1: Generate tasks with time allocation."""
    prompt = f"""Decompose this role into its 6-8 most important tasks with time allocation:

Job Title: {job_title}
SOC Code: {soc_code}
Function: {function_name}
Primary Sector: {sector}

Return a JSON array where each element has:
{{
  "task": "O*NET-style task statement (verb-object-context, one sentence)",
  "time_share": <int 5-40, percentage of workday spent on this task — ALL must sum to 100>,
  "importance": <float 1.0-5.0, how critical this task is to the role>,
  "frequency": "<daily|weekly|monthly|quarterly|annually>"
}}

CRITICAL: time_share values must sum to exactly 100. Think carefully about how a typical
professional in this role actually spends their day/week.

Order by time_share descending (most time-consuming task first)."""

    resp = client.messages.create(
        model=MODEL,
        max_tokens=1500,
        system=GENERATOR_SYSTEM,
        messages=[{"role": "user", "content": prompt}],
    )
    text = resp.content[0].text.strip()
    if text.startswith("```"):
        text = text.split("\n", 1)[1].rsplit("```", 1)[0]
    tasks = json.loads(text)

    # Validate time_share sums to 100
    total = sum(t.get("time_share", 0) for t in tasks)
    if total != 100:
        # Normalize to 100
        for t in tasks:
            t["time_share"] = round(t["time_share"] * 100 / total)
        # Fix rounding to hit exactly 100
        diff = 100 - sum(t["time_share"] for t in tasks)
        if diff != 0:
            tasks[0]["time_share"] += diff

    return tasks


# ---------------------------------------------------------------------------
# AGENT 2: QUALITY REVIEWER (v2 — validates Time_Share)
# ---------------------------------------------------------------------------
REVIEWER_SYSTEM = """You are a senior occupational data quality reviewer.
You review AI-generated task statements against O*NET standards.

Your job:
1. Check each task follows verb-object-context format
2. Flag and fix any vague or overlapping tasks
3. Validate time_share allocations are realistic:
   - Does the total equal 100%?
   - Are the proportions plausible for this role? (e.g., a software engineer should NOT
     spend 40% on meetings — that's more like 10-15%)
   - Adjust if needed based on your knowledge of how these roles actually spend time
4. Verify importance ratings are calibrated (not all 4.5+)
5. Assign each task a Task Type: "Core" if time_share >= 10 AND importance >= 3.0, else "Supplemental"
6. Assign the closest Generalized Work Activity (GWA) category from this list:
   - Analyzing Data/Information
   - Making Decisions and Solving Problems
   - Processing Information
   - Communicating with Supervisors, Peers, or Subordinates
   - Communicating with People Outside the Organization
   - Organizing, Planning, and Prioritizing Work
   - Updating and Using Relevant Knowledge
   - Establishing and Maintaining Interpersonal Relationships
   - Documenting/Recording Information
   - Evaluating Information to Determine Compliance
   - Developing Objectives and Strategies
   - Interacting With Computers
   - Getting Information
   - Monitoring Processes, Materials, or Surroundings
   - Coordinating the Work and Activities of Others
   - Training and Teaching Others
   - Thinking Creatively
   - Selling or Influencing Others
   - Performing Administrative Activities

Respond ONLY with valid JSON, no markdown fences, no preamble."""

def review_tasks(job_title: str, soc_code: str, raw_tasks: list[dict]) -> list[dict]:
    """Agent 2: Review, validate time_share, and enrich tasks."""
    prompt = f"""Review these AI-generated tasks for: {job_title} ({soc_code})

Raw tasks:
{json.dumps(raw_tasks, indent=2)}

Return a JSON array with the reviewed tasks. Each element:
{{
  "task": "corrected task statement if needed",
  "time_share": <int, adjusted so total = 100>,
  "importance": <float, adjusted if miscalibrated>,
  "frequency": "<daily|weekly|monthly|quarterly|annually>",
  "task_type": "Core or Supplemental",
  "gwa": "closest GWA category from the list",
  "review_note": "what you changed, or 'approved' if no changes"
}}

CRITICAL: time_share must sum to exactly 100 across all tasks."""

    resp = client.messages.create(
        model=MODEL,
        max_tokens=2000,
        system=REVIEWER_SYSTEM,
        messages=[{"role": "user", "content": prompt}],
    )
    text = resp.content[0].text.strip()
    if text.startswith("```"):
        text = text.split("\n", 1)[1].rsplit("```", 1)[0]
    tasks = json.loads(text)

    # Final validation: ensure sum = 100
    total = sum(t.get("time_share", 0) for t in tasks)
    if total != 100 and total > 0:
        for t in tasks:
            t["time_share"] = round(t["time_share"] * 100 / total)
        diff = 100 - sum(t["time_share"] for t in tasks)
        if diff != 0:
            tasks[0]["time_share"] += diff

    return tasks


# ---------------------------------------------------------------------------
# AGENT 3: FORMATTER (v2 — Time_Share weight + de-duplicated employment)
# ---------------------------------------------------------------------------
def format_tasks(soc_code: str, job_title: str, function_id: str,
                 function_name: str, reviewed_tasks: list[dict],
                 emp_info: dict) -> list[dict]:
    """Agent 3: Calculate economy-level weights, assign IDs, produce final rows."""
    dedup_emp = emp_info.get('dedup_employment') or 0

    rows = []
    for i, t in enumerate(reviewed_tasks, start=1):
        task_id = f"CL-{soc_code.replace('-', '')}-{i:03d}"
        time_share = t.get("time_share", 0) / 100.0  # Convert to decimal

        # Economy-level weight: time_share × de-duplicated employment
        # This gives "thousands of FTE-equivalents doing this task"
        economy_weight = round(time_share * dedup_emp, 2) if dedup_emp else None

        rows.append({
            "SOC_Code": soc_code,
            "Job_Title": job_title,
            "Function_ID": function_id,
            "Function_Name": function_name,
            "Task_ID": task_id,
            "Task_Description": t["task"],
            "Task_Source": "Claude-generated",
            "Task_Type": t.get("task_type", ""),
            "Time_Share_Pct": t.get("time_share", 0),
            "Importance": t.get("importance", 0),
            "Frequency": t.get("frequency", ""),
            "GWA": t.get("gwa", ""),
            "Dedup_Employment_K": dedup_emp if dedup_emp else "",
            "Economy_Weight_K": economy_weight if economy_weight else "",
            "Automatability_Score": "",  # blank for your scoring
            "Review_Note": t.get("review_note", ""),
        })
    return rows


# ---------------------------------------------------------------------------
# PIPELINE
# ---------------------------------------------------------------------------
def process_job(job: dict, emp_info: dict) -> list[dict]:
    """Run the full 3-agent pipeline for one job."""
    title = job["title"]
    soc = job["soc_code"]
    func_id = job.get("function_id", "")
    func_name = job.get("function_name", "")
    sector = job.get("sector", "")

    print(f"  [Agent 1] Generating tasks for: {title} ({soc})")
    raw = generate_tasks(title, soc, func_name, sector)

    print(f"  [Agent 2] Reviewing {len(raw)} tasks (time_share sum={sum(t.get('time_share',0) for t in raw)})...")
    reviewed = review_tasks(title, soc, raw)
    ts_sum = sum(t.get('time_share', 0) for t in reviewed)
    print(f"            Reviewed time_share sum={ts_sum}")

    print(f"  [Agent 3] Formatting with dedup_emp={emp_info.get('dedup_employment', 'N/A')}K...")
    final = format_tasks(soc, title, func_id, func_name, reviewed, emp_info)

    return final


def read_jobs_from_workbook(filepath: str, sheet_name: str) -> list[dict]:
    """Read job list from the workbook."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    col_map = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        hl = str(h).strip()
        if hl == "SOC_Code":
            col_map["soc_code"] = i
        elif hl == "Custom_Title":
            col_map["title"] = i
        elif hl == "SOC_Title":
            col_map["soc_title"] = i
        elif hl == "Function_ID":
            col_map["function_id"] = i
        elif hl == "Function_Name":
            col_map["function_name"] = i
        elif hl == "Delta_Sector":
            col_map["sector"] = i

    if "title" not in col_map:
        for i, h in enumerate(headers):
            if h and "title" in str(h).lower():
                col_map["title"] = i
                break

    jobs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        soc = row[col_map.get("soc_code", 0)]
        if soc is None:
            continue
        title = str(row[col_map.get("title", 2)] or "").strip()
        if not title and "soc_title" in col_map:
            title = str(row[col_map["soc_title"]] or "").strip()
        jobs.append({
            "soc_code": str(soc).strip(),
            "title": title,
            "function_id": str(row[col_map.get("function_id", 5)] or "").strip(),
            "function_name": str(row[col_map.get("function_name", 6)] or "").strip(),
            "sector": str(row[col_map.get("sector", 4)] or "").strip(),
        })

    wb.close()
    print(f"Loaded {len(jobs)} jobs from '{sheet_name}'")
    return jobs


def load_occupations():
    """Load occupations_master.csv into a lookup dict."""
    occ = {}
    path = os.path.join(BASE_DIR, "occupations_master.csv")
    with open(path, 'r', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            occ[row['SOC_Code']] = row
    return occ


def read_existing_tasks(filepath: str) -> list[dict]:
    """Read existing tasks from the '3 Tasks' tab if it exists."""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception:
        return []

    sheet_name = None
    for name in wb.sheetnames:
        if 'task' in name.lower():
            sheet_name = name
            break

    if not sheet_name:
        wb.close()
        return []

    ws = wb[sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    tasks = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        task = {}
        for i, h in enumerate(headers):
            if h:
                task[h] = row[i] if i < len(row) else ''
        if task.get('Task_ID'):
            tasks.append(task)
    wb.close()
    print(f"Read {len(tasks)} existing tasks from '{sheet_name}'")
    return tasks


def write_tasks_to_workbook(filepath: str, all_tasks: list[dict]):
    """Write tasks to a '3 Tasks' tab in the workbook."""
    wb = openpyxl.load_workbook(filepath)

    # Remove any existing tasks tab
    for name in list(wb.sheetnames):
        if 'task' in name.lower():
            del wb[name]

    ws = wb.create_sheet("3 Tasks")

    columns = [
        "SOC_Code", "Job_Title", "Function_ID", "Function_Name",
        "Task_ID", "Task_Description", "Task_Source", "Task_Type",
        "Time_Share_Pct", "Importance", "Frequency", "GWA",
        "Dedup_Employment_K", "Economy_Weight_K",
        "Automatability_Score", "Review_Note",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for col_i, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A2"

    for row_i, task in enumerate(all_tasks, start=2):
        for col_i, key in enumerate(columns, start=1):
            val = task.get(key, "")
            cell = ws.cell(row=row_i, column=col_i, value=val)
            if key == "Automatability_Score":
                cell.fill = yellow_fill

    # Column widths
    widths = {
        "SOC_Code": 10, "Job_Title": 28, "Function_ID": 8, "Function_Name": 30,
        "Task_ID": 16, "Task_Description": 65, "Task_Source": 14, "Task_Type": 12,
        "Time_Share_Pct": 12, "Importance": 10, "Frequency": 10, "GWA": 35,
        "Dedup_Employment_K": 16, "Economy_Weight_K": 16,
        "Automatability_Score": 16, "Review_Note": 30,
    }
    for col_i, key in enumerate(columns, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_i)].width = widths.get(key, 12)

    wb.save(filepath)
    print(f"\nWrote {len(all_tasks)} tasks to 'Tasks' tab in {filepath}")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Generate tasks v2 with Time_Share weighting")
    parser.add_argument("--input", default="jobs-data.xlsx", help="Path to workbook")
    parser.add_argument("--sheet", default="2 Jobs", help="Sheet with job list")
    parser.add_argument("--output", default=None, help="Output path (defaults to input)")
    parser.add_argument("--limit", type=int, default=None, help="Process first N jobs only")
    parser.add_argument("--delay", type=float, default=0.5, help="Delay between API calls")
    parser.add_argument("--workers", type=int, default=1, help="Number of parallel workers")
    parser.add_argument("--new-only", action="store_true",
                        help="Only process jobs that don't have tasks yet; merge with existing")
    args = parser.parse_args()
    output = args.output or args.input

    # Load jobs and occupation data
    jobs = read_jobs_from_workbook(args.input, args.sheet)
    occ_lookup = load_occupations()

    # If --new-only, read existing tasks and filter to jobs without tasks
    existing_tasks = []
    if args.new_only:
        existing_tasks = read_existing_tasks(args.input)
        existing_titles = set(t.get('Job_Title', '') for t in existing_tasks)
        all_job_titles = [j['title'] for j in jobs]
        jobs = [j for j in jobs if j['title'] not in existing_titles]
        print(f"--new-only: {len(existing_titles)} jobs already have tasks, {len(jobs)} new jobs to process")

    if args.limit:
        jobs = jobs[:args.limit]
        print(f"Limited to first {args.limit} jobs")

    # Build de-duplicated employment table
    print("\n--- Building de-duplicated employment table ---")
    emp_table = build_employment_table(jobs, occ_lookup)

    # Report de-duplication
    shared_count = sum(1 for v in emp_table.values() if v['shared_with'] > 1)
    missing_count = sum(1 for v in emp_table.values() if v['dedup_employment'] is None)
    print(f"  Jobs with shared SOC codes: {shared_count}")
    print(f"  Jobs with missing employment: {missing_count}")
    if missing_count:
        for title, info in emp_table.items():
            if info['dedup_employment'] is None:
                print(f"    {title} ({info['soc_code']} → {info['fixed_soc']})")

    # Show de-dup examples
    print("\n  De-duplication examples:")
    examples = ['Software Engineer', 'Machine Learning Engineer', 'Management Consultant', 'Strategy Consultant']
    for ex in examples:
        if ex in emp_table:
            e = emp_table[ex]
            print(f"    {ex}: {e['national_employment']}K national / {e['shared_with']} titles = {e['dedup_employment']}K")

    # Process jobs
    all_tasks = []
    errors = []
    print_lock = threading.Lock()
    counter = {'done': 0}

    def process_one(idx_job):
        i, job = idx_job
        title = job['title']
        emp_info = emp_table.get(title, {})
        try:
            tasks = process_job(job, emp_info)
            ts_total = sum(t.get('Time_Share_Pct', 0) for t in tasks)
            with print_lock:
                counter['done'] += 1
                print(f"[{counter['done']}/{len(jobs)}] {title} ({job['soc_code']}) — {len(tasks)} tasks, Time_Share={ts_total}%", flush=True)
            return ('ok', title, tasks)
        except Exception as e:
            with print_lock:
                counter['done'] += 1
                print(f"[{counter['done']}/{len(jobs)}] {title} — ERROR: {e}", flush=True)
            return ('error', title, str(e))

    num_workers = args.workers
    if num_workers > 1:
        print(f"\n--- Processing {len(jobs)} jobs with {num_workers} parallel workers ---", flush=True)
        with ThreadPoolExecutor(max_workers=num_workers) as executor:
            futures = {executor.submit(process_one, (i, job)): job
                       for i, job in enumerate(jobs, start=1)}
            for future in as_completed(futures):
                result = future.result()
                if result[0] == 'ok':
                    all_tasks.extend(result[2])
                else:
                    errors.append((result[1], result[2]))
    else:
        print(f"\n--- Processing {len(jobs)} jobs sequentially ---", flush=True)
        for i, job in enumerate(jobs, start=1):
            result = process_one((i, job))
            if result[0] == 'ok':
                all_tasks.extend(result[2])
            else:
                errors.append((result[1], result[2]))
            if args.delay and i < len(jobs):
                time.sleep(args.delay)

    # Merge with existing tasks if --new-only
    if args.new_only and existing_tasks:
        print(f"\nMerging {len(existing_tasks)} existing + {len(all_tasks)} new tasks")
        all_tasks = existing_tasks + all_tasks

    # Write output
    write_tasks_to_workbook(output, all_tasks)

    # Summary
    print(f"\n{'='*60}")
    print(f"SUMMARY")
    print(f"{'='*60}")
    print(f"Jobs processed: {len(jobs) - len(errors)}/{len(jobs)}")
    print(f"Errors: {len(errors)}")
    print(f"Total tasks generated: {len(all_tasks)}")
    processed = len(jobs) - len(errors)
    if processed > 0:
        print(f"Avg tasks per job: {len(all_tasks)/processed:.1f}")

    # Verify time_share sums
    job_ts = defaultdict(int)
    for t in all_tasks:
        job_ts[t['Job_Title']] += t.get('Time_Share_Pct', 0)
    bad_sums = {k: v for k, v in job_ts.items() if v != 100}
    print(f"\nTime_Share validation:")
    print(f"  Jobs summing to 100%: {len(job_ts) - len(bad_sums)}/{len(job_ts)}")
    if bad_sums:
        print(f"  Jobs with incorrect sums:")
        for k, v in bad_sums.items():
            print(f"    {k}: {v}%")

    # Employment coverage
    total_dedup = sum(
        emp_table[t]['dedup_employment']
        for t in set(t_row['Job_Title'] for t_row in all_tasks)
        if t in emp_table and emp_table[t]['dedup_employment']
    )
    print(f"\nTotal de-duplicated employment: {total_dedup:,.1f}K ({total_dedup*1000:,.0f} workers)")

    if errors:
        print(f"\nFailed jobs:")
        for title, err in errors:
            print(f"  {title}: {err}")

    print(f"\nOutput: {output} → 'Tasks' tab")


if __name__ == "__main__":
    main()
