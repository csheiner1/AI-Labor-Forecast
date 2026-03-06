"""Load workbook data for the Flask dashboard.

Reads from 4 Results and 6 Social Impact tabs on startup.
Caches in memory for fast page rendering.
"""
import logging
import openpyxl
import os
import threading

logger = logging.getLogger(__name__)

from social_impact.config import WORKBOOK


class DataStore:
    """In-memory cache of workbook data for the dashboard."""

    def __init__(self):
        self.results = []       # 4 Results tab rows
        self.social = []        # 6 Social Impact tab rows
        self.soc_lookup = {}    # SOC -> merged dict of results + social
        self._displacement_data = None  # cached for transition API
        self._loaded = False
        self._lock = threading.Lock()

    def load(self):
        """Load data from workbook into memory (thread-safe)."""
        if self._loaded:
            return
        with self._lock:
            # Double-check after acquiring lock
            if self._loaded:
                return
            self._load_impl()

    def _load_impl(self):
        """Internal load, called under lock."""

        wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)

        # Load 4 Results -- derive column count from header row
        ws = wb["4 Results"]
        headers = None
        for row_vals in ws.iter_rows(values_only=True):
            if headers is None:
                headers = list(row_vals)
                continue
            if len(row_vals) < len(headers):
                continue
            soc = row_vals[0]
            if not soc:
                continue
            row = {}
            for c, h in enumerate(headers):
                if h:
                    row[h] = row_vals[c]
            self.results.append(row)

        # Load 6 Social Impact -- derive column count from header row
        try:
            ws2 = wb["6 Social Impact"]
            headers2 = None
            for row_vals in ws2.iter_rows(values_only=True):
                if headers2 is None:
                    headers2 = list(row_vals)
                    continue
                if len(row_vals) < len(headers2):
                    continue
                soc = row_vals[0]
                if not soc:
                    continue
                row = {}
                for c, h in enumerate(headers2):
                    if h:
                        row[h] = row_vals[c]
                self.social.append(row)
        except KeyError:
            print("WARNING: '6 Social Impact' tab not found. Run social_impact/run.py first.")

        wb.close()

        # Build merged lookup -- for SOCs with multiple rows in Results,
        # keep the first row (which has the primary sector assignment)
        social_by_soc = {}
        for r in self.social:
            soc_val = r.get("SOC_Code")
            if soc_val:
                social_by_soc[soc_val] = r
        results_socs = set()
        for r in self.results:
            soc = r.get("SOC_Code")
            if not soc or soc in self.soc_lookup:
                continue
            results_socs.add(soc)
            merged = dict(r)
            if soc in social_by_soc:
                merged.update(social_by_soc[soc])
            self.soc_lookup[soc] = merged

        # Warn about SOCs in Results that are missing from Social tab
        if social_by_soc:
            missing = results_socs - set(social_by_soc.keys())
            if missing:
                logger.warning(
                    "%d SOC(s) in Results tab missing from Social Impact tab: %s",
                    len(missing), sorted(missing)[:10]
                )

        self._loaded = True
        print(f"DataStore loaded: {len(self.results)} results, {len(self.social)} social impact rows, "
              f"{len(self.soc_lookup)} unique SOCs")

    def get_all(self):
        """Return all SOC records as merged dicts."""
        self.load()
        return list(self.soc_lookup.values())

    def get_soc(self, soc_code):
        """Return one SOC record."""
        self.load()
        return self.soc_lookup.get(soc_code)

    def get_sectors(self):
        """Return list of unique sectors."""
        self.load()
        return sorted(set(r.get("Sector", "") for r in self.soc_lookup.values() if r.get("Sector")))

    def get_displacement_data(self):
        """Return displacement_data dict for transition API, cached after first call."""
        self.load()
        if self._displacement_data is None:
            with self._lock:
                # Double-check after acquiring lock
                if self._displacement_data is None:
                    data = {}
                    for soc, rec in self.soc_lookup.items():
                        data[soc] = {
                            "title": rec.get("Job_Title") or rec.get("Custom_Title", ""),
                            "d_mod_low": rec.get("d_mod_low", 0),
                            "employment_K": rec.get("Employment_2024_K", 0),
                        }
                    self._displacement_data = data
        return self._displacement_data

    def get_wage_quintiles(self):
        """Return SOC codes grouped by wage quintile."""
        self.load()
        wages = [(r["SOC_Code"], r.get("Median_Wage", 0) or 0) for r in self.soc_lookup.values()]
        wages.sort(key=lambda x: x[1])
        n = len(wages)
        quintile_size = n // 5
        quintiles = {}
        labels = ["Q1 (lowest)", "Q2", "Q3", "Q4", "Q5 (highest)"]
        for i, label in enumerate(labels):
            start = i * quintile_size
            end = start + quintile_size if i < 4 else n
            quintiles[label] = [w[0] for w in wages[start:end]]
        return quintiles


# Singleton instance
store = DataStore()
