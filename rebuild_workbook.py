#!/usr/bin/env python3
"""
HISTORICAL — This script was run once to expand the workbook from 17→21 sectors.
Tab names in this script are outdated (pre-rename). Do not re-run without updating.
Current tab names: 1 NAICS Mapping, 1A Sector Summary, 2 Staffing Patterns,
3 Tasks, 4 Results, 5L/5H Frictions Low/High. Deleted tabs: 2 Jobs, 2B Job_Industry,
Jobs_All_Industry.

Original description:
Rebuild jobs-data-v3.xlsx to 21 sectors + 10 new SOC codes.

Changes:
- 17 modeled sectors → 21 (restore Staffing, Construction, Transport, Wholesale, Accommodation&Food)
- Split "Logistics & Distribution" → "Transportation & Logistics" + "Wholesale Trade"
- Insurance remains carved out from Finance (existing)
- Add 10 missing white-collar SOC codes
- Rebuild: Lookup_Sectors, 1A Industries, 1A Summary, Staffing Patterns,
           Jobs_All_Industry, 2B Job_Industry, 2 Jobs, Lookup_Jobs, 4 Results,
           Low/High Industry Frictions, ReadMe
"""

import csv
import copy
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict
from datetime import datetime

# ============================================================
# CONFIGURATION
# ============================================================

WORKBOOK = 'jobs-data-v3.xlsx'
NIOEM_CSV = 'nioem_filtered.csv'
OCC_MASTER_CSV = 'archive/csv_intermediates/occupations_master.csv'
LOOKUP_SECTORS_CSV = 'archive/csv_intermediates/lookup_sectors.csv'

# New 21-sector layout
# NIOEM sectors map to new workbook sector IDs
# NIOEM 1 is split: Finance NAICS → WB 1, Insurance NAICS → WB 2
NEW_SECTORS = {
    1: 'Finance & Banking',
    2: 'Insurance',
    3: 'Technology & Software',
    4: 'Healthcare & Life Sciences',
    5: 'Law Firms & Legal Services',
    6: 'Management Consulting Firms',
    7: 'Accounting & Tax Firms',
    8: 'Advertising & PR Agencies',
    9: 'Staffing & Recruitment Agencies',
    10: 'Real Estate & Property',
    11: 'Education & Academia',
    12: 'Government & Public Administration',
    13: 'Media Publishing & Entertainment',
    14: 'Energy & Utilities',
    15: 'Architecture & Engineering Firms',
    16: 'Manufacturing',
    17: 'Retail Trade',
    18: 'Construction',
    19: 'Transportation & Logistics',
    20: 'Wholesale Trade',
    21: 'Accommodation & Food Services',
}

# Insurance NAICS codes (carved from NIOEM 1 → WB 2)
INSURANCE_NAICS = {'5241', '5242', '5251', '524114'}

# NIOEM sector ID → new workbook sector ID
# NIOEM 1 is special: split by NAICS into Finance (1) and Insurance (2)
NIOEM_TO_WB = {
    # '1': split by NAICS
    '2': 3,   # Technology
    '3': 4,   # Healthcare
    '4': 5,   # Law
    '5': 6,   # Consulting
    '6': 7,   # Accounting
    '7': 8,   # Advertising
    '8': 9,   # Staffing (restored)
    '9': 10,  # Real Estate
    '10': 11, # Education
    '11': 12, # Government
    '12': 13, # Media
    '13': 14, # Energy
    '14': 15, # Architecture & Engineering
    '15': 16, # Manufacturing
    '16': 17, # Retail
    '17': 18, # Construction (restored)
    '18': 19, # Transportation (restored, was part of "Logistics")
    '19': 20, # Wholesale (restored, was part of "Logistics")
    '20': 21, # Accommodation & Food Services (restored)
}

# For NIOEM 1, determine WB sector by NAICS code
def nioem_to_wb_sector(nioem_sid, naics_code):
    """Convert NIOEM sector + NAICS to new workbook sector ID."""
    if nioem_sid == '1':
        # Check if insurance NAICS
        naics_str = str(naics_code)
        if any(naics_str.startswith(ins) for ins in INSURANCE_NAICS):
            return 2  # Insurance
        return 1  # Finance & Banking
    return NIOEM_TO_WB.get(nioem_sid)

# NAICS → sub-industry mapping from lookup_sectors.csv
# We'll load this dynamically

# New SOC codes to add
NEW_SOCS = [
    {'soc': '13-1071', 'title': 'HR Specialist'},
    {'soc': '15-1244', 'title': 'Systems Administrator'},  # Already has tasks (orphan)
    {'soc': '29-1051', 'title': 'Pharmacist'},
    {'soc': '29-2052', 'title': 'Pharmacy Technician'},
    {'soc': '13-1121', 'title': 'Event Planner'},
    {'soc': '43-5032', 'title': 'Dispatcher'},
    {'soc': '43-4161', 'title': 'HR Assistant'},
    {'soc': '41-2021', 'title': 'Counter Clerk'},
    {'soc': '43-4151', 'title': 'Order Clerk'},
    {'soc': '43-4181', 'title': 'Reservation Agent'},
]

# Old sector name → new sector name mapping (for updating existing data)
OLD_TO_NEW_SECTOR_NAME = {
    'Finance & Financial Services': 'Finance & Banking',
    'Logistics & Distribution': None,  # Split: check NAICS
}

# ============================================================
# LOAD SOURCE DATA
# ============================================================

print("Loading source data...")

# Load NIOEM
nioem_rows = []
with open(NIOEM_CSV) as f:
    reader = csv.DictReader(f)
    for row in reader:
        nioem_rows.append(row)
print(f"  NIOEM: {len(nioem_rows)} rows")

# Load occupations master
occ_master = {}
with open(OCC_MASTER_CSV) as f:
    reader = csv.DictReader(f)
    for row in reader:
        if row['Occupation_Type'] == 'Line item':
            occ_master[row['SOC_Code']] = row
print(f"  Occupations master: {len(occ_master)} line items")

# Load lookup sectors (old 20-sector NAICS mapping)
naics_lookup = []
with open(LOOKUP_SECTORS_CSV) as f:
    reader = csv.DictReader(f)
    for row in reader:
        naics_lookup.append(row)
print(f"  Lookup sectors: {len(naics_lookup)} NAICS codes")

# Build NAICS → sub-industry mapping
naics_to_subindustry = {}
naics_to_mapping_type = {}
for row in naics_lookup:
    naics_to_subindustry[row['NAICS_Code']] = row['Delta_Sub_Industry']
    naics_to_mapping_type[row['NAICS_Code']] = row['Mapping_Type']

# ============================================================
# BUILD NEW NIOEM AGGREGATION
# ============================================================

print("\nBuilding 21-sector aggregation from NIOEM...")

# Aggregate NIOEM employment by (new_sector_id, SOC_Code)
emp_by_sector_soc = defaultdict(float)  # (wb_sid, soc) → emp_K
emp_by_sector_naics = defaultdict(float)  # (wb_sid, naics) → emp_K
emp_by_soc_all = defaultdict(float)  # soc → total emp across all 21 sectors

# Also track per-sector totals
sector_total_emp = defaultdict(float)

# Track all unique SOC codes in NIOEM
all_nioem_socs = set()

for row in nioem_rows:
    nioem_sid = row['Delta_Sector_ID']
    naics = row['NAICS_Code']
    soc = row['SOC_Code']
    emp = float(row['Employment_2024_thousands'])

    wb_sid = nioem_to_wb_sector(nioem_sid, naics)
    if wb_sid is None:
        continue

    emp_by_sector_soc[(wb_sid, soc)] += emp
    emp_by_sector_naics[(wb_sid, naics)] += emp
    emp_by_soc_all[soc] += emp
    sector_total_emp[wb_sid] += emp
    all_nioem_socs.add(soc)

print(f"  Unique SOC codes in NIOEM: {len(all_nioem_socs)}")
for sid in sorted(NEW_SECTORS.keys()):
    print(f"  Sector {sid} ({NEW_SECTORS[sid]}): {sector_total_emp[sid]:.1f}K")

# ============================================================
# LOAD EXISTING WORKBOOK
# ============================================================

print(f"\nLoading workbook {WORKBOOK}...")
# Backup first
backup_name = f'{WORKBOOK}.bak_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
shutil.copy2(WORKBOOK, backup_name)
print(f"  Backed up to {backup_name}")

wb = openpyxl.load_workbook(WORKBOOK)

# ============================================================
# PRESERVE EXISTING SCORED DATA
# ============================================================

print("\nPreserving existing scored data...")

# Preserve 4 Results data (keyed by SOC_Code + Custom_Title)
ws_results = wb['4 Results']
results_headers = [ws_results.cell(1, c).value for c in range(1, ws_results.max_column + 1)]
existing_results = {}
for r in range(2, ws_results.max_row + 1):
    soc = ws_results.cell(r, 1).value
    title = ws_results.cell(r, 2).value
    if soc is None:
        continue
    row_data = {}
    for c, h in enumerate(results_headers, 1):
        row_data[h] = ws_results.cell(r, c).value
    existing_results[(soc, title)] = row_data
print(f"  Preserved {len(existing_results)} results rows")

# Preserve 3 Tasks data (all existing tasks)
ws_tasks = wb['3 Tasks']
task_headers = [ws_tasks.cell(1, c).value for c in range(1, ws_tasks.max_column + 1)]
existing_tasks = []
for r in range(2, ws_tasks.max_row + 1):
    soc = ws_tasks.cell(r, 1).value
    if soc is None:
        continue
    row_data = {}
    for c, h in enumerate(task_headers, 1):
        row_data[h] = ws_tasks.cell(r, c).value
    existing_tasks.append(row_data)
print(f"  Preserved {len(existing_tasks)} task rows")

# Preserve 2 Jobs data (keyed by SOC_Code + Custom_Title)
ws_jobs = wb['2 Jobs']
jobs_headers = [ws_jobs.cell(1, c).value for c in range(1, ws_jobs.max_column + 1)]
existing_jobs = {}
for r in range(2, ws_jobs.max_row + 1):
    soc = ws_jobs.cell(r, 1).value
    title = ws_jobs.cell(r, 2).value
    if soc is None:
        continue
    row_data = {}
    for c, h in enumerate(jobs_headers, 1):
        row_data[h] = ws_jobs.cell(r, c).value
    existing_jobs[(soc, title)] = row_data
print(f"  Preserved {len(existing_jobs)} job rows")

# Preserve existing Jobs_All_Industry data (keyed by SOC_Code)
ws_jai = wb['Jobs_All_Industry']
jai_headers = [ws_jai.cell(1, c).value for c in range(1, ws_jai.max_column + 1)]
existing_jai = {}
for r in range(2, ws_jai.max_row + 1):
    soc = ws_jai.cell(r, 1).value
    if soc is None:
        continue
    row_data = {}
    for c, h in enumerate(jai_headers, 1):
        row_data[h] = ws_jai.cell(r, c).value
    existing_jai[soc] = row_data
print(f"  Preserved {len(existing_jai)} Jobs_All_Industry rows")

# Preserve existing frictions data (keyed by old sector name)
existing_frictions = {'Low': {}, 'High': {}}
for scenario in ['Low', 'High']:
    ws_fr = wb[f'{scenario} Industry Frictions']
    for r in range(5, ws_fr.max_row + 1):
        sid = ws_fr.cell(r, 1).value
        name = ws_fr.cell(r, 2).value
        if sid is None:
            continue
        row_data = []
        for c in range(1, ws_fr.max_column + 1):
            row_data.append(ws_fr.cell(r, c).value)
        existing_frictions[scenario][name] = row_data
print(f"  Preserved {len(existing_frictions['Low'])} Low friction rows, {len(existing_frictions['High'])} High friction rows")

# Get existing SOC codes in the model
existing_soc_set = set()
existing_job_titles = {}  # soc -> list of custom titles
for (soc, title) in existing_jobs:
    existing_soc_set.add(soc)
    if soc not in existing_job_titles:
        existing_job_titles[soc] = []
    existing_job_titles[soc].append(title)

# ============================================================
# DETERMINE PRIMARY SECTOR FOR EACH SOC
# ============================================================

def get_primary_sector(soc_code):
    """Determine primary sector (highest employment) for a SOC code across 21 sectors."""
    sector_emp = {}
    for sid in NEW_SECTORS:
        emp = emp_by_sector_soc.get((sid, soc_code), 0)
        if emp > 0:
            sector_emp[sid] = emp
    if not sector_emp:
        return None, {}
    primary = max(sector_emp, key=sector_emp.get)
    return primary, sector_emp

def get_top3_industries(soc_code):
    """Get top 3 industries with percentages for a SOC code."""
    _, sector_emp = get_primary_sector(soc_code)
    total = sum(sector_emp.values())
    if total == 0:
        return None, None, None
    sorted_sectors = sorted(sector_emp.items(), key=lambda x: -x[1])
    result = []
    for sid, emp in sorted_sectors[:3]:
        pct = round(100 * emp / total)
        result.append(f"{NEW_SECTORS[sid]} ({pct}%)")
    while len(result) < 3:
        result.append(None)
    return result[0], result[1], result[2]


# ============================================================
# REBUILD LOOKUP_SECTORS
# ============================================================

print("\nRebuilding Lookup_Sectors...")
ws = wb['Lookup_Sectors']

# Clear existing data
for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        ws.cell(r, c).value = None

# Write header
ls_headers = ['Sector_ID', 'Sector', 'Sub_Industry', 'NAICS_Code', 'NAICS_Title', 'Mapping_Type']
for c, h in enumerate(ls_headers, 1):
    ws.cell(1, c).value = h
    ws.cell(1, c).font = Font(bold=True)

# Build new lookup from naics_lookup, remapping sector IDs
row_num = 2
for entry in naics_lookup:
    old_sid = entry['Delta_Sector_ID']
    naics = entry['NAICS_Code']

    # Map to new sector ID
    new_sid = nioem_to_wb_sector(old_sid, naics)
    if new_sid is None:
        continue

    ws.cell(row_num, 1).value = str(new_sid)
    ws.cell(row_num, 2).value = NEW_SECTORS[new_sid]
    ws.cell(row_num, 3).value = entry['Delta_Sub_Industry']
    ws.cell(row_num, 4).value = naics
    ws.cell(row_num, 5).value = entry['NAICS_Title']
    ws.cell(row_num, 6).value = entry['Mapping_Type']
    row_num += 1

print(f"  Wrote {row_num - 2} rows")


# ============================================================
# REBUILD 1A INDUSTRIES
# ============================================================

print("\nRebuilding 1A Industries...")
ws = wb['1A Industries']

# Clear existing data
for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        ws.cell(r, c).value = None

# Write header
ind_headers = ['NAICS_Code', 'NAICS_Title', 'Sector_ID', 'Sector', 'Sub_Industry',
               'Mapping_Type', 'NEM_Code', 'Employment_2024 (Thousands)']
for c, h in enumerate(ind_headers, 1):
    ws.cell(1, c).value = h
    ws.cell(1, c).font = Font(bold=True)

# Aggregate NIOEM employment by NAICS code within each new sector
naics_emp = defaultdict(float)  # (new_sid, naics) → emp
naics_nem = {}  # naics → NEM code (from NIOEM)

for row in nioem_rows:
    nioem_sid = row['Delta_Sector_ID']
    naics = row['NAICS_Code']
    nem = row['NEM_Code']
    emp = float(row['Employment_2024_thousands'])

    new_sid = nioem_to_wb_sector(nioem_sid, naics)
    if new_sid is None:
        continue

    naics_emp[(new_sid, naics)] += emp
    naics_nem[naics] = nem

# Get NAICS titles from lookup
naics_titles = {}
for entry in naics_lookup:
    naics_titles[entry['NAICS_Code']] = entry['NAICS_Title']

# Write rows, grouped by sector
row_num = 2
for sid in sorted(NEW_SECTORS.keys()):
    # Get all NAICS codes for this sector
    sector_naics = set()
    for (s, n) in naics_emp:
        if s == sid:
            sector_naics.add(n)

    for naics in sorted(sector_naics):
        ws.cell(row_num, 1).value = naics
        ws.cell(row_num, 2).value = naics_titles.get(naics, naics)
        ws.cell(row_num, 3).value = sid
        ws.cell(row_num, 4).value = NEW_SECTORS[sid]
        ws.cell(row_num, 5).value = naics_to_subindustry.get(naics, '')
        ws.cell(row_num, 6).value = naics_to_mapping_type.get(naics, 'Direct')
        ws.cell(row_num, 7).value = naics_nem.get(naics, '')
        ws.cell(row_num, 8).value = round(naics_emp[(sid, naics)], 1)
        row_num += 1

print(f"  Wrote {row_num - 2} rows")


# ============================================================
# REBUILD 1A SUMMARY
# ============================================================

print("\nRebuilding 1A Summary...")
ws = wb['1A Summary']

# Clear existing data
for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        ws.cell(r, c).value = None

# Write header
sum_headers = ['Sector_ID', 'Sector', 'Num_NAICS_Codes', 'Employment_2024 (Thousands)',
               'Avg_Median_Wage']
for c, h in enumerate(sum_headers, 1):
    ws.cell(1, c).value = h
    ws.cell(1, c).font = Font(bold=True)

# Compute sector summaries
# Get NAICS count and employment per sector from 1A Industries
ws_ind = wb['1A Industries']
sector_naics_count = defaultdict(int)
sector_emp_sum = defaultdict(float)

for r in range(2, row_num):  # row_num from industries rebuild
    sid = ws_ind.cell(r, 3).value
    emp = ws_ind.cell(r, 8).value
    if sid is not None:
        sector_naics_count[sid] += 1
        if emp:
            sector_emp_sum[sid] += emp

# Compute average median wage per sector from NIOEM SOC codes
# Weight by employment within sector
sector_wage_sum = defaultdict(float)
sector_wage_weight = defaultdict(float)

for (wb_sid, soc), emp in emp_by_sector_soc.items():
    if soc in occ_master and emp > 0:
        wage_str = occ_master[soc].get('Median_Annual_Wage', '')
        try:
            wage = float(wage_str.replace(',', ''))
            sector_wage_sum[wb_sid] += wage * emp
            sector_wage_weight[wb_sid] += emp
        except (ValueError, AttributeError):
            pass

row_num = 2
for sid in sorted(NEW_SECTORS.keys()):
    ws.cell(row_num, 1).value = sid
    ws.cell(row_num, 2).value = NEW_SECTORS[sid]
    ws.cell(row_num, 3).value = sector_naics_count.get(sid, 0)
    ws.cell(row_num, 4).value = round(sector_emp_sum.get(sid, 0), 1)

    # Avg median wage
    if sector_wage_weight[sid] > 0:
        avg_wage = sector_wage_sum[sid] / sector_wage_weight[sid]
        ws.cell(row_num, 5).value = round(avg_wage, 2)

    row_num += 1

print(f"  Wrote {row_num - 2} sectors")


# ============================================================
# REBUILD 2 JOBS (add new SOCs + update sector names)
# ============================================================

print("\nRebuilding 2 Jobs...")
ws = wb['2 Jobs']

# Clear existing data (keep header)
for r in range(2, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        ws.cell(r, c).value = None

# Column layout: SOC_Code, Custom_Title, Sector, Employment_2024 (Thousands),
# Median_Wage, Primary_Industry, Industry_2, Industry_3,
# Projected_Change_Pct, Mapping_Notes, Automatability_Score, Weight

row_num = 2

# Write existing jobs with updated sector names
for (soc, title), data in sorted(existing_jobs.items()):
    old_sector = data.get('Sector', '')

    # Map old sector name to new
    if old_sector == 'Finance & Financial Services':
        new_sector = 'Finance & Banking'
    elif old_sector == 'Logistics & Distribution':
        # Determine from NIOEM: is this primarily Transport or Wholesale?
        primary_sid, sector_emp = get_primary_sector(soc)
        # Check if more in Transport (19) or Wholesale (20)
        trans_emp = sector_emp.get(19, 0)
        whole_emp = sector_emp.get(20, 0)
        if trans_emp >= whole_emp:
            new_sector = 'Transportation & Logistics'
        else:
            new_sector = 'Wholesale Trade'
    else:
        new_sector = old_sector  # Unchanged

    # Recompute employment scoped to all 21 modeled sectors
    primary_sid, sector_emp = get_primary_sector(soc)
    total_modeled_emp = sum(sector_emp.values())

    # For split-SOC jobs, divide evenly among custom titles
    num_titles = len(existing_job_titles.get(soc, [title]))
    job_emp = round(total_modeled_emp / num_titles, 1) if total_modeled_emp > 0 else data.get('Employment_2024 (Thousands)', 0)

    # Get top 3 industries
    ind1, ind2, ind3 = get_top3_industries(soc)

    ws.cell(row_num, 1).value = soc
    ws.cell(row_num, 2).value = title
    ws.cell(row_num, 3).value = new_sector
    ws.cell(row_num, 4).value = job_emp
    ws.cell(row_num, 5).value = data.get('Median_Wage')
    ws.cell(row_num, 6).value = ind1
    ws.cell(row_num, 7).value = ind2
    ws.cell(row_num, 8).value = ind3
    ws.cell(row_num, 9).value = data.get('Projected_Change_Pct')
    ws.cell(row_num, 10).value = data.get('Mapping_Notes')
    row_num += 1

# Add new SOC codes
for new_soc in NEW_SOCS:
    soc = new_soc['soc']
    title = new_soc['title']

    if (soc, title) in existing_jobs:
        continue  # Already exists

    primary_sid, sector_emp = get_primary_sector(soc)
    total_modeled_emp = sum(sector_emp.values())

    # Get BLS data
    bls = occ_master.get(soc, {})
    wage_str = bls.get('Median_Annual_Wage', '')
    try:
        wage = float(wage_str.replace(',', ''))
    except (ValueError, AttributeError):
        wage = None

    chg_str = bls.get('Change_Percent', '')
    try:
        chg = float(chg_str)
    except (ValueError, AttributeError):
        chg = None

    ind1, ind2, ind3 = get_top3_industries(soc)

    ws.cell(row_num, 1).value = soc
    ws.cell(row_num, 2).value = title
    ws.cell(row_num, 3).value = NEW_SECTORS[primary_sid] if primary_sid else ''
    ws.cell(row_num, 4).value = round(total_modeled_emp, 1)
    ws.cell(row_num, 5).value = wage
    ws.cell(row_num, 6).value = ind1
    ws.cell(row_num, 7).value = ind2
    ws.cell(row_num, 8).value = ind3
    ws.cell(row_num, 9).value = chg
    ws.cell(row_num, 10).value = 'Direct'
    row_num += 1

print(f"  Wrote {row_num - 2} job rows")


# ============================================================
# REBUILD LOOKUP_JOBS
# ============================================================

print("\nRebuilding Lookup_Jobs...")
ws = wb['Lookup_Jobs']

# Clear existing data
for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        ws.cell(r, c).value = None

# Write header
lj_headers = ['Custom_Title', 'Sector_ID', 'SOC_Code', 'SOC_Title', 'Mapping_Notes']
for c, h in enumerate(lj_headers, 1):
    ws.cell(1, c).value = h
    ws.cell(1, c).font = Font(bold=True)

# Read back 2 Jobs to build lookup
ws_jobs = wb['2 Jobs']
row_num = 2
for r in range(2, ws_jobs.max_row + 1):
    soc = ws_jobs.cell(r, 1).value
    title = ws_jobs.cell(r, 2).value
    sector_name = ws_jobs.cell(r, 3).value
    notes = ws_jobs.cell(r, 10).value

    if soc is None:
        break

    # Find sector ID from name
    sector_id = None
    for sid, sname in NEW_SECTORS.items():
        if sname == sector_name:
            sector_id = str(sid)
            break

    # Get BLS title
    bls_title = occ_master.get(soc, {}).get('SOC_Title', '')

    ws.cell(row_num, 1).value = title
    ws.cell(row_num, 2).value = sector_id
    ws.cell(row_num, 3).value = soc
    ws.cell(row_num, 4).value = bls_title
    ws.cell(row_num, 5).value = notes
    row_num += 1

print(f"  Wrote {row_num - 2} rows")


# ============================================================
# REBUILD STAFFING PATTERNS
# ============================================================

print("\nRebuilding Staffing Patterns...")
ws = wb['Staffing Patterns']

# Clear existing data
for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        ws.cell(r, c).value = None

# Write header
sp_headers = ['Sector_ID', 'Sector', 'SOC_Code', 'SOC_Title', 'Employment (Thousands)', 'Staffing_Share_Pct']
for c, h in enumerate(sp_headers, 1):
    ws.cell(1, c).value = h
    ws.cell(1, c).font = Font(bold=True)

# Get all SOC codes in the model (existing + new)
model_socs = set(existing_soc_set)
for ns in NEW_SOCS:
    model_socs.add(ns['soc'])

# Build staffing patterns from NIOEM, filtered to model SOCs
row_num = 2
for sid in sorted(NEW_SECTORS.keys()):
    sector_name = NEW_SECTORS[sid]

    # Get all (soc, emp) pairs for this sector
    sector_entries = []
    total_sector_emp = 0

    for (s, soc), emp in emp_by_sector_soc.items():
        if s == sid and soc in model_socs and emp > 0:
            soc_title = occ_master.get(soc, {}).get('SOC_Title', soc)
            sector_entries.append((soc, soc_title, emp))
            total_sector_emp += emp

    # Sort by employment descending
    sector_entries.sort(key=lambda x: -x[2])

    for soc, soc_title, emp in sector_entries:
        share = round(100 * emp / total_sector_emp, 2) if total_sector_emp > 0 else 0
        ws.cell(row_num, 1).value = str(sid)
        ws.cell(row_num, 2).value = sector_name
        ws.cell(row_num, 3).value = soc
        ws.cell(row_num, 4).value = soc_title
        ws.cell(row_num, 5).value = round(emp, 1)
        ws.cell(row_num, 6).value = share
        row_num += 1

print(f"  Wrote {row_num - 2} staffing pattern rows")


# ============================================================
# REBUILD JOBS_ALL_INDUSTRY
# ============================================================

print("\nRebuilding Jobs_All_Industry...")
ws = wb['Jobs_All_Industry']

# Clear existing data
for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        ws.cell(r, c).value = None

# New column layout
# Cols 1-9: metadata, then one column per sector
sector_order = sorted(NEW_SECTORS.keys())
jai_headers = ['SOC_Code', 'BLS_Title', 'Custom_Titles', 'Primary_Sector',
               'All_21_NIOEM_K', 'K12_Residual_K', 'Model_Employment_K']
for sid in sector_order:
    jai_headers.append(NEW_SECTORS[sid])

for c, h in enumerate(jai_headers, 1):
    ws.cell(1, c).value = h
    ws.cell(1, c).font = Font(bold=True)

# Get all unique SOC codes in the model, with custom title mappings
# Build from 2 Jobs
ws_jobs_rebuilt = wb['2 Jobs']
soc_to_titles = defaultdict(list)
for r in range(2, ws_jobs_rebuilt.max_row + 1):
    soc = ws_jobs_rebuilt.cell(r, 1).value
    title = ws_jobs_rebuilt.cell(r, 2).value
    if soc is None:
        break
    soc_to_titles[soc].append(title)

row_num = 2
for soc in sorted(soc_to_titles.keys()):
    titles = soc_to_titles[soc]
    bls_title = occ_master.get(soc, {}).get('SOC_Title', soc)
    custom_str = ' | '.join(titles)

    primary_sid, sector_emp = get_primary_sector(soc)
    total_emp = sum(sector_emp.values())

    ws.cell(row_num, 1).value = soc
    ws.cell(row_num, 2).value = bls_title
    ws.cell(row_num, 3).value = custom_str
    ws.cell(row_num, 4).value = NEW_SECTORS[primary_sid] if primary_sid else ''
    ws.cell(row_num, 5).value = round(total_emp, 1)  # All 21 NIOEM
    ws.cell(row_num, 6).value = None  # K12 residual - keep blank
    ws.cell(row_num, 7).value = round(total_emp, 1)  # Model employment = all 21

    # Per-sector employment columns
    for i, sid in enumerate(sector_order):
        col = 8 + i  # starts at col 8
        emp = emp_by_sector_soc.get((sid, soc), 0)
        ws.cell(row_num, col).value = round(emp, 1) if emp > 0 else None

    row_num += 1

print(f"  Wrote {row_num - 2} rows")


# ============================================================
# REBUILD 2B JOB_INDUSTRY
# ============================================================

print("\nRebuilding 2B Job_Industry...")
ws = wb['2B Job_Industry']

# Clear existing data
for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        ws.cell(r, c).value = None

# Header row 1: title
ws.cell(1, 1).value = 'Employment by Industry (Thousands)'
ws.cell(1, 1).font = Font(bold=True)

# Header row 2: column names
ji_headers = ['SOC_Code', 'SOC_Title']
for sid in sector_order:
    ji_headers.append(NEW_SECTORS[sid])

for c, h in enumerate(ji_headers, 1):
    ws.cell(2, c).value = h
    ws.cell(2, c).font = Font(bold=True)

# Data rows starting at row 3
row_num = 3
for soc in sorted(soc_to_titles.keys()):
    bls_title = occ_master.get(soc, {}).get('SOC_Title', soc)

    ws.cell(row_num, 1).value = soc
    ws.cell(row_num, 2).value = bls_title

    for i, sid in enumerate(sector_order):
        col = 3 + i
        emp = emp_by_sector_soc.get((sid, soc), 0)
        ws.cell(row_num, col).value = round(emp, 1) if emp > 0 else None

    row_num += 1

print(f"  Wrote {row_num - 3} rows")


# ============================================================
# UPDATE 4 RESULTS
# ============================================================

print("\nRebuilding 4 Results...")
ws = wb['4 Results']

# Clear data rows (keep header)
for r in range(2, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        ws.cell(r, c).value = None

# Columns stay the same:
# SOC_Code, Custom_Title, Sector, Employment_2024_K, Median_Wage,
# task_coverage_mod, task_coverage_sig, workflow_simplicity,
# x_depth_mod, x_depth_sig, x_scale, x_sub,
# a_mod, a_sig, phi_mod, phi_sig,
# d_max, E, T_18mo_low, T_18mo_high, R_value_low, R_value_high,
# d_mod_low, d_mod_high, d_sig_low, d_sig_high,
# displaced_K_mod_low, displaced_K_mod_high, displaced_K_sig_low, displaced_K_sig_high

row_num = 2

# Read back rebuilt 2 Jobs to get the authoritative job list + sectors
ws_jobs_new = wb['2 Jobs']
for r in range(2, ws_jobs_new.max_row + 1):
    soc = ws_jobs_new.cell(r, 1).value
    title = ws_jobs_new.cell(r, 2).value
    sector = ws_jobs_new.cell(r, 3).value
    emp = ws_jobs_new.cell(r, 4).value
    wage = ws_jobs_new.cell(r, 5).value

    if soc is None:
        break

    ws.cell(row_num, 1).value = soc
    ws.cell(row_num, 2).value = title
    ws.cell(row_num, 3).value = sector
    ws.cell(row_num, 4).value = emp
    ws.cell(row_num, 5).value = wage

    # Preserve existing scoring data if available
    # Match by SOC + old title (titles haven't changed, only sectors)
    old_result = existing_results.get((soc, title))
    if old_result:
        # Copy scoring columns (6-30)
        scoring_cols = ['task_coverage_mod', 'task_coverage_sig', 'workflow_simplicity',
                       'x_depth_mod', 'x_depth_sig', 'x_scale', 'x_sub',
                       'a_mod', 'a_sig', 'phi_mod', 'phi_sig']
        for i, col_name in enumerate(scoring_cols):
            ws.cell(row_num, 6 + i).value = old_result.get(col_name)

        # d_max, E, T, R values will need recalculation for new sectors
        # but preserve them as-is for existing sectors
        old_sector = old_result.get('Sector', '')
        if old_sector == sector or (old_sector == 'Finance & Financial Services' and sector == 'Finance & Banking'):
            # Same sector, preserve frictions
            friction_cols = ['d_max', 'E', 'T_18mo_low', 'T_18mo_high',
                           'R_value_low', 'R_value_high',
                           'd_mod_low', 'd_mod_high', 'd_sig_low', 'd_sig_high',
                           'displaced_K_mod_low', 'displaced_K_mod_high',
                           'displaced_K_sig_low', 'displaced_K_sig_high']
            for i, col_name in enumerate(friction_cols):
                ws.cell(row_num, 17 + i).value = old_result.get(col_name)

    row_num += 1

print(f"  Wrote {row_num - 2} results rows")


# ============================================================
# UPDATE FRICTIONS TABS
# ============================================================

print("\nRebuilding Frictions tabs...")

for scenario in ['Low', 'High']:
    ws = wb[f'{scenario} Industry Frictions']

    # Preserve header rows (1-4) - just update them
    # Row 1: title
    # Row 2: legend
    # Row 3: sub-headers
    # Row 4: column labels

    # Clear data rows (5+)
    for r in range(5, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None

    # Write new sector rows
    row_num = 5
    for sid in sorted(NEW_SECTORS.keys()):
        sector_name = NEW_SECTORS[sid]

        # Try to find existing friction data by matching sector name
        old_data = None
        if sector_name in existing_frictions[scenario]:
            old_data = existing_frictions[scenario][sector_name]
        elif sector_name == 'Finance & Banking':
            old_data = existing_frictions[scenario].get('Finance & Financial Services')
        elif sector_name == 'Transportation & Logistics':
            old_data = existing_frictions[scenario].get('Logistics & Distribution')
        elif sector_name == 'Wholesale Trade':
            # Use Logistics data as starting point for Wholesale too
            old_data = existing_frictions[scenario].get('Logistics & Distribution')

        ws.cell(row_num, 1).value = sid
        ws.cell(row_num, 2).value = sector_name
        ws.cell(row_num, 3).value = round(sector_emp_sum.get(sid, 0), 1)

        # Compute avg wage for this sector
        if sector_wage_weight[sid] > 0:
            ws.cell(row_num, 4).value = round(sector_wage_sum[sid] / sector_wage_weight[sid])

        if old_data:
            # Copy friction values (columns 5-20)
            for c in range(5, min(len(old_data) + 1, 21)):
                if c <= len(old_data):
                    ws.cell(row_num, c).value = old_data[c - 1]
        else:
            # New sector - leave friction values blank (need manual scoring)
            pass

        row_num += 1

print(f"  Wrote 21 rows per friction tab")


# ============================================================
# UPDATE README
# ============================================================

print("\nUpdating ReadMe...")
ws = wb['ReadMe']

# Clear and rewrite
for r in range(1, ws.max_row + 1):
    ws.cell(r, 1).value = None

readme_lines = [
    'AI Labor Analysis Dataset',
    '',
    'DATA SOURCES',
    '- BLS Employment Projections 2024-2034 (occupation.xlsx, Tables 1.2/1.8/1.9)',
    '- National Industry-Occupation Employment Matrix (NIOEM) via BLS API',
    '- Employment figures are in THOUSANDS (152.3 = 152,300 workers)',
    '- Wage data: BLS 2024 median annual wages. Downloaded March 2026.',
    '',
    'ARCHITECTURE',
    '21 industries (NAICS-based) × ~420 occupations (SOC-based).',
    'Cross-industry roles (CEO, HR Manager, etc.) are assigned to their',
    'highest-employment industry. The subsegment columns (Primary_Industry,',
    'Industry_2, Industry_3) and the 2B/Jobs_All_Industry tabs show the full distribution.',
    '',
    'SECTORS (21)',
    ' 1  Finance & Banking           (NAICS 5221-5239)',
    ' 2  Insurance                   (NAICS 5241-5251)',
    ' 3  Technology & Software       (NAICS 5112,5182,5415,3344,517)',
    ' 4  Healthcare & Life Sciences  (NAICS 3254,3391,54171,524114,621,622)',
    ' 5  Law Firms & Legal Services  (NAICS 5411)',
    ' 6  Management Consulting       (NAICS 5416)',
    ' 7  Accounting & Tax Firms      (NAICS 5412)',
    ' 8  Advertising & PR Agencies   (NAICS 5418,54191)',
    ' 9  Staffing & Recruitment      (NAICS 5613)',
    '10  Real Estate & Property      (NAICS 531)',
    '11  Education & Academia        (NAICS 6111,6113,6117)',
    '12  Government & Public Admin   (NAICS 921,922,926,928)',
    '13  Media, Publishing & Entertainment (NAICS 5111,5121,5122,516)',
    '14  Energy & Utilities          (NAICS 2111,2211,2212)',
    '15  Architecture & Engineering  (NAICS 54131,54133,54138)',
    '16  Manufacturing               (NAICS 311,312,325,326,331,332,333,335,336)',
    '17  Retail Trade                (NAICS 441,445,452,455,456)',
    '18  Construction                (NAICS 236,237,238)',
    '19  Transportation & Logistics  (NAICS 481,482,484,485,488,492,493)',
    '20  Wholesale Trade             (NAICS 423,424,4251)',
    '21  Accommodation & Food Services (NAICS 721,722)',
    '',
    'TABS',
    '1A Industries    — NAICS codes with employment, rolled up to 21 sectors',
    '1A Summary       — Sector-level rollup with avg wages',
    '2 Jobs           — Job titles with BLS data + top-3 industry subsegments',
    '2B Job_Industry  — Full SOC × 21-Industry employment pivot table',
    '3 Tasks          — ~3,200 tasks with Time_Share and autonomy scores',
    '4 Results        — Displacement model results',
    'Jobs_All_Industry— Full NIOEM employment matrix (SOC × 21 sectors)',
    'Staffing Patterns— Top occupations by share within each sector',
    'Industry Frictions— T, R, E scoring for Low/High scenarios',
    'Lookup_*         — Reference tables for the taxonomy',
    '',
    'KEY COLUMNS',
    'Employment_Thousands  — BLS employment in thousands',
    'Median_Wage           — BLS median annual wage ($)',
    'Projected_Change_Pct  — Projected 2024-2034 change (%)',
    'Primary_Industry      — Highest-employment industry for this SOC, with %',
    '',
    'KNOWN LIMITATIONS',
    '- Some NAICS codes share a single BLS aggregate (Composite mapping type)',
    '- Education (Sector 11): NIOEM covers private sector only.',
    '- Blue-collar-heavy sectors (Manufacturing, Retail, Construction, Transport,',
    '  Wholesale, Accommodation) have white-collar SOC coverage of 30-50%.',
    '- Government uses BLS aggregate NEM codes (999100/999200), not NAICS',
    '',
    f'Last rebuilt: {datetime.now().strftime("%Y-%m-%d")}',
]

for i, line in enumerate(readme_lines, 1):
    ws.cell(i, 1).value = line

print("  ReadMe updated")


# ============================================================
# SAVE
# ============================================================

print(f"\nSaving workbook...")
wb.save(WORKBOOK)
print(f"  Saved to {WORKBOOK}")
print(f"  Backup at {backup_name}")


# ============================================================
# VERIFICATION
# ============================================================

print("\n" + "=" * 60)
print("VERIFICATION")
print("=" * 60)

# Reload to verify
wb2 = openpyxl.load_workbook(WORKBOOK, data_only=True)

# Count rows
sheets_to_check = {
    '1A Industries': 2,
    '1A Summary': 2,
    '2 Jobs': 2,
    'Lookup_Jobs': 2,
    'Lookup_Sectors': 2,
    'Staffing Patterns': 2,
    'Jobs_All_Industry': 2,
    '2B Job_Industry': 3,
    '3 Tasks': 2,
    '4 Results': 2,
}

for sheet_name, start_row in sheets_to_check.items():
    ws = wb2[sheet_name]
    data_rows = 0
    for r in range(start_row, ws.max_row + 1):
        if ws.cell(r, 1).value is not None:
            data_rows += 1
    print(f"  {sheet_name}: {data_rows} data rows")

# Check SOC coverage
ws_jobs = wb2['2 Jobs']
ws_tasks = wb2['3 Tasks']
job_socs = set()
task_socs = set()
for r in range(2, ws_jobs.max_row + 1):
    soc = ws_jobs.cell(r, 1).value
    if soc:
        job_socs.add(soc)
for r in range(2, ws_tasks.max_row + 1):
    soc = ws_tasks.cell(r, 1).value
    if soc:
        task_socs.add(soc)

orphan_tasks = task_socs - job_socs
jobs_no_tasks = job_socs - task_socs

print(f"\n  Unique SOC codes in Jobs: {len(job_socs)}")
print(f"  Unique SOC codes in Tasks: {len(task_socs)}")
if orphan_tasks:
    print(f"  WARNING: {len(orphan_tasks)} SOC codes in Tasks but not Jobs: {sorted(orphan_tasks)}")
if jobs_no_tasks:
    print(f"  NOTE: {len(jobs_no_tasks)} SOC codes in Jobs but not Tasks (need task generation): {sorted(jobs_no_tasks)}")

# Check sector coverage
ws_sum = wb2['1A Summary']
print(f"\n  Sectors in 1A Summary:")
for r in range(2, ws_sum.max_row + 1):
    sid = ws_sum.cell(r, 1).value
    name = ws_sum.cell(r, 2).value
    emp = ws_sum.cell(r, 4).value
    if sid:
        print(f"    {sid}: {name} ({emp}K)")

print("\nDone!")
