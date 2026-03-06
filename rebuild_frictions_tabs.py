"""
Rebuild 5L/5H Frictions tabs with rows = sector x occupation group.
Pre-fills friction scores from existing sector-level values.
Derived columns use Excel formulas.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

WORKBOOK = 'jobs-data-v3.xlsx'

GROUP_ORDER = [
    'Core', 'G1_Exec_Management', 'G2_HR_People', 'G3_Finance_Accounting',
    'G4_IT_Digital', 'G5_Marketing_Creative', 'G6_Sales_BizDev',
    'G7_Legal_Compliance', 'G8_Procurement_Supply', 'G9_Admin_Office',
]

# Display names (no underscores)
GROUP_DISPLAY = {
    'Core': 'Core',
    'G1_Exec_Management': 'G1 Executive & Management',
    'G2_HR_People': 'G2 HR & People Ops',
    'G3_Finance_Accounting': 'G3 Finance & Accounting',
    'G4_IT_Digital': 'G4 IT & Digital',
    'G5_Marketing_Creative': 'G5 Marketing & Creative',
    'G6_Sales_BizDev': 'G6 Sales & Business Dev',
    'G7_Legal_Compliance': 'G7 Legal & Compliance',
    'G8_Procurement_Supply': 'G8 Procurement & Supply Chain',
    'G9_Admin_Office': 'G9 Admin & Office Support',
}

# Column layout (1-indexed)
# A=1  Sector_ID
# B=2  Sector
# C=3  Occupation_Group
# D=4  Emp (K)
# E=5  Avg Median Wage ($)
# F=6  T1
# G=7  T2
# H=8  T3
# I=9  T4
# J=10 D (Delay Index)
# K=11 t0 (Inflection)
# L=12 T(18mo)
# M=13 f1
# N=14 f2
# O=15 f3
# P=16 F (sum)
# Q=17 R Value
# R=18 E Elasticity
# S=19 Notes
# T=20 mu Velocity Tier
# U=21 d_max Effective


def read_staffing_data(wb):
    """Read staffing patterns and aggregate by (sector, group)."""
    ws = wb['2 Staffing Patterns']

    agg = defaultdict(lambda: {'emp': 0.0, 'wage_emp': 0.0})
    sector_names = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[2]:
            continue
        sid = int(row[0])
        sector = row[1]
        group = row[2]
        emp = float(row[5]) if row[5] else 0
        wage = float(row[8]) if row[8] else 0

        sector_names[sid] = sector
        key = (sid, group)
        agg[key]['emp'] += emp
        agg[key]['wage_emp'] += emp * wage

    # Build sorted list of (sector_id, sector_name, group, emp, avg_wage)
    rows = []
    for (sid, group), data in agg.items():
        if data['emp'] <= 0:
            continue
        avg_wage = data['wage_emp'] / data['emp'] if data['emp'] > 0 else 0
        rows.append({
            'sector_id': sid,
            'sector': sector_names[sid],
            'group': group,
            'emp': round(data['emp'], 1),
            'avg_wage': round(avg_wage, 0),
        })

    # Sort by sector_id, then group order
    group_rank = {g: i for i, g in enumerate(GROUP_ORDER)}
    rows.sort(key=lambda r: (r['sector_id'], group_rank.get(r['group'], 99)))
    return rows


def read_existing_frictions(ws):
    """Read existing friction scores from a 5L or 5H tab (current layout has Occ Group in col C).
    Only reads Core rows to get the sector-level scores. Returns dict: sector_id -> scores."""
    frictions = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row[0] or not isinstance(row[0], (int, float)):
            continue
        # Only read Core rows (they have the sector-level scores)
        if row[2] != 'Core':
            continue
        sid = int(row[0])
        # Current layout: T1=col F(5), T2=G(6), T3=H(7), T4=I(8),
        #                 f1=M(12), f2=N(13), f3=O(14), E=R(17)
        frictions[sid] = {
            'T1': row[5],
            'T2': row[6],
            'T3': row[7],
            'T4': row[8],
            'f1': row[12],
            'f2': row[13],
            'f3': row[14],
            'E': row[17],
        }
    return frictions


def build_frictions_tab(wb, tab_name, staffing_rows, existing_frictions):
    """Create or overwrite a frictions tab with sector x group rows."""

    # Delete existing tab if present
    if tab_name in wb.sheetnames:
        del wb[tab_name]

    ws = wb.create_sheet(tab_name)

    # --- Styles ---
    title_font = Font(name='Calibri', size=12, bold=True)
    legend_font = Font(name='Calibri', size=9, italic=True)
    header_font = Font(name='Calibri', size=10, bold=True)
    data_font = Font(name='Calibri', size=10)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    core_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    wrap = Alignment(wrap_text=True, vertical='top')

    scenario = 'LOW' if 'Low' in tab_name else 'HIGH'

    # --- Row 1: Title ---
    ws.cell(row=1, column=1, value=f'AI DISPLACEMENT -- INDUSTRY x OCCUPATION GROUP FRICTIONS  |  T / R / E SCORING  [{scenario} SCENARIO]')
    ws.cell(row=1, column=1).font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=21)

    # --- Row 2: Legend ---
    legend = ('T drivers (1-4): 1=low friction/fast  2=low-medium  3=medium-high  4=high friction/slow  --  '
              'R (1-4): 1=permissive  2=minor  3=significant  4=hard constraint  --  '
              'E: 1.0=headcount drops  0.5=partial  0.25=demand absorbs')
    ws.cell(row=2, column=1, value=legend)
    ws.cell(row=2, column=1).font = legend_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=21)

    # --- Row 3: Column group headers ---
    group_headers = {
        1: 'Reference', 6: 'T -- Restructuring Timing Drivers  (rate 1-4 each)',
        10: 'T Derived', 13: 'R -- Resistance Sub-components  (rate 1-4 each)',
        16: 'R Derived', 18: 'E', 19: 'Notes', 20: 'd_max',
    }
    for col, label in group_headers.items():
        c = ws.cell(row=3, column=col, value=label)
        c.font = header_font
        c.fill = header_fill

    # --- Row 4: Column detail headers ---
    col_headers = [
        'Sector_ID', 'Sector', 'Occupation\nGroup',
        'Emp 2024\n(K workers)', 'Avg Median\nWage ($)',
        'T1\nInstitutional\nInertia', 'T2\nSystems\nIntegration',
        'T3\nCustomer\nAcceptance', 'T4\nCompetitive\nPressure',
        'Delay\nIndex D', 't0\n(Inflection)', 'T(18mo)',
        'f1\nLiability &\nInsurability', 'f2\nRegulatory',
        'f3\nUnions', 'F\n(sum)', 'R\nValue',
        'E\nElasticity', 'Notes / Rationale',
        'mu(i)\nVelocity Tier', 'd_max\nEffective',
    ]
    for i, h in enumerate(col_headers):
        c = ws.cell(row=4, column=i + 1, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = wrap
        c.border = thin_border

    # --- Data rows (row 5+) ---
    for idx, r in enumerate(staffing_rows):
        row_num = idx + 5
        sid = r['sector_id']
        is_core = r['group'] == 'Core'

        # Reference columns
        display_name = GROUP_DISPLAY.get(r['group'], r['group'])
        ws.cell(row=row_num, column=1, value=sid).font = data_font
        ws.cell(row=row_num, column=2, value=r['sector']).font = data_font
        ws.cell(row=row_num, column=3, value=display_name).font = data_font
        ws.cell(row=row_num, column=4, value=r['emp']).font = data_font
        ws.cell(row=row_num, column=5, value=r['avg_wage']).font = data_font

        # Only pre-fill friction inputs for Core rows (G1-G9 left empty for manual entry)
        if is_core:
            scores = existing_frictions.get(sid, {})
            for col, key in [(6, 'T1'), (7, 'T2'), (8, 'T3'), (9, 'T4'),
                             (13, 'f1'), (14, 'f2'), (15, 'f3'), (18, 'E')]:
                val = scores.get(key)
                if val is not None:
                    ws.cell(row=row_num, column=col, value=val).font = data_font

        # Derived formulas (only if T inputs exist)
        # D = AVERAGE(T1:T4)  -> col 10
        ws.cell(row=row_num, column=10).font = data_font
        ws.cell(row=row_num, column=10, value=f'=IF(COUNTA(F{row_num}:I{row_num})=4,AVERAGE(F{row_num}:I{row_num}),"")')

        # t0 = 1.5*D - 0.5*T4 - 0.5  -> col 11
        ws.cell(row=row_num, column=11).font = data_font
        ws.cell(row=row_num, column=11, value=f'=IF(J{row_num}<>"",1.5*J{row_num}-0.5*I{row_num}-0.5,"")')

        # T(18mo) = 1/(1+EXP(-1.2*(1.5-t0)))  -> col 12
        ws.cell(row=row_num, column=12).font = data_font
        ws.cell(row=row_num, column=12, value=f'=IF(K{row_num}<>"",1/(1+EXP(-1.2*(1.5-K{row_num}))),"")')

        # F = f1 + f2 + f3  -> col 16
        ws.cell(row=row_num, column=16).font = data_font
        ws.cell(row=row_num, column=16, value=f'=IF(COUNTA(M{row_num}:O{row_num})=3,SUM(M{row_num}:O{row_num}),"")')

        # R = 1 - 0.7*(F-3)/9  -> col 17
        ws.cell(row=row_num, column=17).font = data_font
        ws.cell(row=row_num, column=17, value=f'=IF(P{row_num}<>"",1-0.7*(P{row_num}-3)/9,"")')

        # Apply borders and Core highlight
        for col in range(1, 22):
            cell = ws.cell(row=row_num, column=col)
            cell.border = thin_border

        if r['group'] == 'Core':
            for col in range(1, 22):
                ws.cell(row=row_num, column=col).fill = core_fill

    # --- Column widths ---
    widths = {1: 10, 2: 35, 3: 30, 4: 12, 5: 14,
              6: 12, 7: 12, 8: 12, 9: 12,
              10: 10, 11: 10, 12: 10,
              13: 12, 14: 10, 15: 10, 16: 8, 17: 8,
              18: 10, 19: 30, 20: 12, 21: 12}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Number formats
    last_row = len(staffing_rows) + 4
    for row_num in range(5, last_row + 1):
        ws.cell(row=row_num, column=4).number_format = '#,##0.0'
        ws.cell(row=row_num, column=5).number_format = '$#,##0'
        ws.cell(row=row_num, column=10).number_format = '0.00'
        ws.cell(row=row_num, column=11).number_format = '0.000'
        ws.cell(row=row_num, column=12).number_format = '0.0000'
        ws.cell(row=row_num, column=16).number_format = '0'
        ws.cell(row=row_num, column=17).number_format = '0.0000'

    print(f"  Built {tab_name}: {len(staffing_rows)} data rows (row 5 to {last_row})")
    return ws


def main():
    wb = openpyxl.load_workbook(WORKBOOK)

    # Read staffing data
    staffing_rows = read_staffing_data(wb)
    print(f"Staffing: {len(staffing_rows)} (sector, group) combos")

    # Read existing friction scores
    frictions_low = read_existing_frictions(wb['5L Frictions Low'])
    frictions_high = read_existing_frictions(wb['5H Frictions High'])
    print(f"Existing Low scores: {len([s for s in frictions_low.values() if s.get('T1') is not None])} sectors scored")
    print(f"Existing High scores: {len([s for s in frictions_high.values() if s.get('T1') is not None])} sectors scored")

    # Get sheet positions — place new tabs where old ones were
    sheet_names = wb.sheetnames
    low_idx = sheet_names.index('5L Frictions Low')
    high_idx = sheet_names.index('5H Frictions High')

    # Build new tabs
    ws_low = build_frictions_tab(wb, '5L Frictions Low', staffing_rows, frictions_low)
    ws_high = build_frictions_tab(wb, '5H Frictions High', staffing_rows, frictions_high)

    # Move tabs to correct positions
    new_names = wb.sheetnames
    wb.move_sheet('5L Frictions Low', offset=low_idx - new_names.index('5L Frictions Low'))
    new_names = wb.sheetnames
    wb.move_sheet('5H Frictions High', offset=high_idx - new_names.index('5H Frictions High'))

    wb.save(WORKBOOK)
    print(f"\nSaved {WORKBOOK}")

    # Verification
    print("\n--- Verification ---")
    wb2 = openpyxl.load_workbook(WORKBOOK, data_only=True)
    for tab in ['5L Frictions Low', '5H Frictions High']:
        ws = wb2[tab]
        data_rows = 0
        scored = 0
        for row in ws.iter_rows(min_row=5, values_only=True):
            if row[0] and isinstance(row[0], (int, float)):
                data_rows += 1
                if row[5] is not None:  # T1 filled
                    scored += 1
        print(f"  {tab}: {data_rows} data rows, {scored} pre-scored from sector defaults")

    print("\n--- Tab order ---")
    for i, name in enumerate(wb2.sheetnames):
        print(f"  {i}: {name}")


if __name__ == '__main__':
    main()
