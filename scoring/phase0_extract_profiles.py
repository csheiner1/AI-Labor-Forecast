"""Phase 0: Extract structured occupation profiles for job-level scoring."""
import json
import openpyxl
import statistics

# GWA categories tagged for construct-specific routing
INTERPERSONAL_GWAS = {
    'Communicating with People Outside the Organization',
    'Communicating with Supervisors, Peers, or Subordinates',
    'Establishing and Maintaining Interpersonal Relationships',
    'Selling or Influencing Others',
    'Training and Teaching Others',
    'Coordinating the Work and Activities of Others',
}
DIGITAL_GWAS = {
    'Interacting With Computers',
    'Processing Information',
    'Analyzing Data/Information',
    'Documenting/Recording Information',
}
JUDGMENT_GWAS = {
    'Making Decisions and Solving Problems',
    'Evaluating Information to Determine Compliance',
    'Thinking Creatively',
    'Developing Objectives and Strategies',
}

wb = openpyxl.load_workbook('jobs-data-v3.xlsx', data_only=True)

# -- Load from 4 Results (authoritative for SOC, title, employment, task_coverage) --
ws_res = wb['4 Results']
jobs = {}
for r in range(2, ws_res.max_row + 1):
    soc = ws_res.cell(r, 1).value
    title = ws_res.cell(r, 2).value
    jobs[title] = {
        'soc_code': soc,
        'custom_title': title,
        'sector': ws_res.cell(r, 3).value,
        'employment_K': ws_res.cell(r, 4).value,
        'median_wage': ws_res.cell(r, 5).value,
        'task_coverage_mod': ws_res.cell(r, 6).value,
        'task_coverage_sig': ws_res.cell(r, 7).value,
    }

# -- Enrich from 2 Staffing Patterns (industry concentration, projections) --
ws_sp = wb['2 Staffing Patterns']
# Build per-SOC: top 3 industries by Occupation_Industry_Share_Pct, median_wage, projected_change
soc_industries = {}  # soc -> list of (sector, share)
soc_meta = {}  # soc -> {wage, proj_change}
for r in range(2, ws_sp.max_row + 1):
    soc = ws_sp.cell(r, 4).value
    sector = ws_sp.cell(r, 2).value
    occ_share = ws_sp.cell(r, 8).value
    wage = ws_sp.cell(r, 9).value
    proj = ws_sp.cell(r, 10).value
    if soc:
        soc_industries.setdefault(soc, []).append((sector, occ_share or 0))
        if soc not in soc_meta:
            soc_meta[soc] = {'wage': wage, 'proj_change': proj}

# Derive top 3 industries per SOC
soc_top3 = {}
for soc, industries in soc_industries.items():
    ranked = sorted(industries, key=lambda x: x[1], reverse=True)[:3]
    soc_top3[soc] = [f"{name} ({share:.0f}%)" for name, share in ranked]

for title, job in jobs.items():
    soc = job['soc_code']
    top3 = soc_top3.get(soc, [])
    job['primary_industry'] = top3[0] if len(top3) > 0 else None
    job['industry_2'] = top3[1] if len(top3) > 1 else None
    job['industry_3'] = top3[2] if len(top3) > 2 else None
    meta = soc_meta.get(soc, {})
    job['projected_change_pct'] = meta.get('proj_change')

# -- Load Tasks (join on title, fallback from SOC|title) --
ws_tasks = wb['3 Tasks']
task_data_by_soc_title = {}  # SOC|title -> list
task_data_by_title = {}      # title -> list
for r in range(2, ws_tasks.max_row + 1):
    soc = ws_tasks.cell(r, 1).value
    job_title = ws_tasks.cell(r, 2).value
    task = {
        'task_id': ws_tasks.cell(r, 3).value,
        'description': ws_tasks.cell(r, 4).value,
        'task_type': ws_tasks.cell(r, 5).value,
        'time_share_pct': ws_tasks.cell(r, 6).value,
        'importance': ws_tasks.cell(r, 7).value,
        'frequency': ws_tasks.cell(r, 8).value,
        'gwa': ws_tasks.cell(r, 9).value,
        'aut_score_mod': ws_tasks.cell(r, 12).value,
        'aut_score_sig': ws_tasks.cell(r, 13).value,
    }
    key = f"{soc}|{job_title}"
    task_data_by_soc_title.setdefault(key, []).append(task)
    task_data_by_title.setdefault(job_title, []).append(task)

# -- Build profiles --
profiles = []
fallback_count = 0
for title, job in jobs.items():
    # Try exact SOC|title match first, fall back to title-only
    key = f"{job['soc_code']}|{title}"
    tasks = task_data_by_soc_title.get(key, [])
    if not tasks:
        tasks = task_data_by_title.get(title, [])
        if tasks:
            fallback_count += 1
    if not tasks:
        print(f"WARNING: No tasks for {title} (SOC {job['soc_code']})")
        continue

    mod_scores = [t['aut_score_mod'] for t in tasks if t['aut_score_mod'] is not None]
    sig_scores = [t['aut_score_sig'] for t in tasks if t['aut_score_sig'] is not None]

    # Task-level summary stats
    def stats(scores):
        if len(scores) < 2:
            return {'mean': scores[0] if scores else None, 'std': 0, 'min': scores[0] if scores else None, 'max': scores[0] if scores else None, 'iqr': 0}
        s = sorted(scores)
        q1_idx = len(s) // 4
        q3_idx = 3 * len(s) // 4
        return {
            'mean': round(statistics.mean(s), 4),
            'std': round(statistics.stdev(s), 4),
            'min': s[0],
            'max': s[-1],
            'iqr': round(s[q3_idx] - s[q1_idx], 4),
        }

    # Top/bottom 3 tasks by sig score
    sorted_tasks_sig = sorted(tasks, key=lambda t: t['aut_score_sig'] or 0, reverse=True)
    top3 = [{'desc': t['description'][:120], 'gwa': t['gwa'], 'score_mod': t['aut_score_mod'], 'score_sig': t['aut_score_sig'], 'time_share': t['time_share_pct']} for t in sorted_tasks_sig[:3]]
    bottom3 = [{'desc': t['description'][:120], 'gwa': t['gwa'], 'score_mod': t['aut_score_mod'], 'score_sig': t['aut_score_sig'], 'time_share': t['time_share_pct']} for t in sorted_tasks_sig[-3:]]

    # GWA distribution
    gwa_counts = {}
    for t in tasks:
        g = t['gwa']
        if g:
            gwa_counts[g] = gwa_counts.get(g, 0) + 1
    total_tasks = len(tasks)

    # Construct-specific features
    interpersonal_share = sum(gwa_counts.get(g, 0) for g in INTERPERSONAL_GWAS) / total_tasks if total_tasks else 0
    digital_share = sum(gwa_counts.get(g, 0) for g in DIGITAL_GWAS) / total_tasks if total_tasks else 0
    judgment_share = sum(gwa_counts.get(g, 0) for g in JUDGMENT_GWAS) / total_tasks if total_tasks else 0

    # Task heterogeneity (std of autonomy scores / mean)
    het_mod = round(stats(mod_scores)['std'] / max(stats(mod_scores)['mean'], 0.01), 4) if mod_scores else 0
    het_sig = round(stats(sig_scores)['std'] / max(stats(sig_scores)['mean'], 0.01), 4) if sig_scores else 0

    profile = {
        'soc_code': job['soc_code'],
        'custom_title': job['custom_title'],
        'sector': job['sector'],
        'employment_K': job['employment_K'],
        'median_wage': job['median_wage'],
        'primary_industry': job['primary_industry'],
        'industry_2': job.get('industry_2'),
        'industry_3': job.get('industry_3'),
        'projected_change_pct': job.get('projected_change_pct'),
        'task_coverage_mod': job.get('task_coverage_mod'),
        'task_coverage_sig': job.get('task_coverage_sig'),
        'num_tasks': total_tasks,
        'mod_stats': stats(mod_scores),
        'sig_stats': stats(sig_scores),
        'heterogeneity_mod': het_mod,
        'heterogeneity_sig': het_sig,
        'interpersonal_task_share': round(interpersonal_share, 3),
        'digital_task_share': round(digital_share, 3),
        'judgment_task_share': round(judgment_share, 3),
        'top3_tasks': top3,
        'bottom3_tasks': bottom3,
        'gwa_distribution': {k: v for k, v in sorted(gwa_counts.items(), key=lambda x: -x[1])},
    }
    profiles.append(profile)

# Sort by sector then title for consistent ordering
profiles.sort(key=lambda p: (p['sector'], p['custom_title']))

# Add index
for i, p in enumerate(profiles):
    p['idx'] = i

if fallback_count:
    print(f"Title-only fallback used for {fallback_count} jobs (SOC mismatch between Results and Tasks)")

print(f"Extracted {len(profiles)} profiles")
print(f"Sectors: {len(set(p['sector'] for p in profiles))}")

# Save
with open('scoring/job_profiles.json', 'w') as f:
    json.dump(profiles, f, indent=2)

# Print sector distribution
from collections import Counter
sector_counts = Counter(p['sector'] for p in profiles)
for s, c in sorted(sector_counts.items()):
    print(f"  {s}: {c}")

print("\nSample profile:")
print(json.dumps(profiles[0], indent=2))
