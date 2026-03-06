"""Phase 0: Extract O*NET tasks and map to our 357 Staffing Patterns SOC entries.

Reads O*NET bulk data files, maps tasks to GWA categories via the DWA chain,
pulls importance/frequency ratings, and organizes by our merged SOC entries.

Output: scoring/task_pipeline/onet_raw.json
"""

import csv
import json
import os
import openpyxl
from collections import Counter

ONET_DIR = "onet_data/db_29_1_text"
WORKBOOK = "jobs-data-v3.xlsx"
OUTPUT = "scoring/task_pipeline/onet_raw.json"

# GWA Element ID → short name mapping (matching our workbook conventions)
GWA_NAMES = {
    "4.A.1.a.1": "Getting Information",
    "4.A.1.a.2": "Monitoring Processes, Materials, or Surroundings",
    "4.A.1.b.1": "Identifying Objects, Actions, and Events",
    "4.A.1.b.2": "Inspecting Equipment, Structures, or Materials",
    "4.A.1.b.3": "Estimating Quantifiable Characteristics",
    "4.A.2.a.1": "Judging Qualities of Objects, Services, or People",
    "4.A.2.a.2": "Processing Information",
    "4.A.2.a.3": "Evaluating Information to Determine Compliance",
    "4.A.2.a.4": "Analyzing Data/Information",
    "4.A.2.b.1": "Making Decisions and Solving Problems",
    "4.A.2.b.2": "Thinking Creatively",
    "4.A.2.b.3": "Updating and Using Relevant Knowledge",
    "4.A.2.b.4": "Developing Objectives and Strategies",
    "4.A.2.b.5": "Scheduling Work and Activities",
    "4.A.2.b.6": "Organizing, Planning, and Prioritizing Work",
    "4.A.3.a.1": "Performing General Physical Activities",
    "4.A.3.a.2": "Handling and Moving Objects",
    "4.A.3.a.3": "Controlling Machines and Processes",
    "4.A.3.a.4": "Operating Vehicles or Equipment",
    "4.A.3.b.1": "Interacting With Computers",
    "4.A.3.b.2": "Drafting/Specifying Technical Devices",
    "4.A.3.b.4": "Repairing Mechanical Equipment",
    "4.A.3.b.5": "Repairing Electronic Equipment",
    "4.A.3.b.6": "Documenting/Recording Information",
    "4.A.4.a.1": "Interpreting Information for Others",
    "4.A.4.a.2": "Communicating with Supervisors, Peers, or Subordinates",
    "4.A.4.a.3": "Communicating with People Outside the Organization",
    "4.A.4.a.4": "Establishing and Maintaining Interpersonal Relationships",
    "4.A.4.a.5": "Assisting and Caring for Others",
    "4.A.4.a.6": "Selling or Influencing Others",
    "4.A.4.a.7": "Resolving Conflicts and Negotiating",
    "4.A.4.a.8": "Performing for or Working with the Public",
    "4.A.4.b.1": "Coordinating the Work and Activities of Others",
    "4.A.4.b.2": "Developing and Building Teams",
    "4.A.4.b.3": "Training and Teaching Others",
    "4.A.4.b.4": "Guiding, Directing, and Motivating Subordinates",
    "4.A.4.b.5": "Coaching and Developing Others",
    "4.A.4.b.6": "Providing Consultation and Advice",
    "4.A.4.c.1": "Performing Administrative Activities",
    "4.A.4.c.2": "Staffing Organizational Units",
    "4.A.4.c.3": "Monitoring and Controlling Resources",
}

# Frequency category → label
FREQ_LABELS = {
    1: "yearly",
    2: "yearly",
    3: "monthly",
    4: "weekly",
    5: "daily",
    6: "daily",
    7: "daily",
}


def load_onet_data():
    """Load all required O*NET data files."""

    # 1. Task Statements
    tasks = {}  # (soc6, task_id) -> {task, task_type}
    with open(os.path.join(ONET_DIR, "Task Statements.txt"), "r") as f:
        reader = csv.reader(f, delimiter="\t")
        next(reader)
        for r in reader:
            soc_full = r[0]
            soc6 = soc_full[:7]
            task_id = r[1]
            tasks[(soc6, task_id)] = {
                "onet_soc": soc_full,
                "task_id": task_id,
                "task_text": r[2],
                "task_type": r[3],  # Core or Supplemental
            }

    # 2. Task Ratings (IM = importance, FT = frequency, RT = relevance)
    ratings = {}  # (soc_full, task_id) -> {importance, frequency_modal, relevance}
    with open(os.path.join(ONET_DIR, "Task Ratings.txt"), "r") as f:
        reader = csv.reader(f, delimiter="\t")
        next(reader)
        ft_dist = {}  # (soc_full, task_id) -> {cat: pct}
        for r in reader:
            soc_full, task_id, scale, category, value = r[0], r[1], r[2], r[3], float(r[4])
            key = (soc_full, task_id)
            if key not in ratings:
                ratings[key] = {}

            if scale == "IM":
                ratings[key]["importance"] = round(value, 1)
            elif scale == "RT":
                ratings[key]["relevance"] = round(value, 1)
            elif scale == "FT":
                ft_dist.setdefault(key, {})[int(category)] = value

        # Convert FT distribution to modal frequency label
        for key, dist in ft_dist.items():
            if dist:
                modal_cat = max(dist, key=dist.get)
                if key in ratings:
                    ratings[key]["frequency"] = FREQ_LABELS.get(modal_cat, "weekly")

    # 3. DWA Reference: DWA ID -> GWA Element ID
    dwa_to_gwa = {}
    with open(os.path.join(ONET_DIR, "DWA Reference.txt"), "r") as f:
        reader = csv.reader(f, delimiter="\t")
        next(reader)
        for r in reader:
            gwa_element_id = r[0]  # e.g., 4.A.2.a.4
            dwa_id = r[2]          # e.g., 4.A.2.a.4.I09.D03
            dwa_to_gwa[dwa_id] = gwa_element_id

    # 4. Tasks to DWAs: (soc_full, task_id) -> list of DWA IDs
    task_dwas = {}
    with open(os.path.join(ONET_DIR, "Tasks to DWAs.txt"), "r") as f:
        reader = csv.reader(f, delimiter="\t")
        next(reader)
        for r in reader:
            soc_full, task_id, dwa_id = r[0], r[1], r[2]
            soc6 = soc_full[:7]
            task_dwas.setdefault((soc6, task_id), []).append(dwa_id)

    return tasks, ratings, dwa_to_gwa, task_dwas


def get_gwa_for_task(soc6, task_id, dwa_to_gwa, task_dwas):
    """Map a task to its primary GWA category via the DWA chain."""
    dwas = task_dwas.get((soc6, task_id), [])
    if not dwas:
        return None

    # Map each DWA to its GWA, pick the most common
    gwa_ids = [dwa_to_gwa.get(d) for d in dwas]
    gwa_ids = [g for g in gwa_ids if g]
    if not gwa_ids:
        return None

    most_common = Counter(gwa_ids).most_common(1)[0][0]
    return GWA_NAMES.get(most_common, most_common)


def load_staffing_entries():
    """Load the 357 SOC entries from Staffing Patterns tab."""
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    sp = wb["2 Staffing Patterns"]

    entries = {}  # soc_code_str -> {title, individual_socs, sectors, employment}
    for row in sp.iter_rows(min_row=2, max_row=sp.max_row, values_only=True):
        sector_id, sector, occ_group, soc_code, soc_title = row[0], row[1], row[2], row[3], row[4]
        emp = row[5]
        if not soc_code or not isinstance(soc_code, str):
            continue

        if soc_code not in entries:
            # Parse individual SOC codes from merged entries
            individual = [s.strip() for s in soc_code.split(",") if "-" in s.strip()]
            entries[soc_code] = {
                "soc_code": soc_code,
                "soc_title": soc_title,
                "individual_socs": individual,
                "is_merged": len(individual) > 1,
                "sectors": [],
                "total_employment_K": 0,
            }

        entries[soc_code]["sectors"].append({
            "sector_id": sector_id,
            "sector": sector,
            "occ_group": occ_group,
            "employment_K": emp,
        })
        entries[soc_code]["total_employment_K"] += (emp or 0)

    wb.close()
    return entries


def build_task_set(entry, tasks, ratings, dwa_to_gwa, task_dwas):
    """Build the O*NET task set for a Staffing Patterns entry."""
    individual_socs = entry["individual_socs"]

    # Collect all O*NET tasks for all constituent SOC codes
    collected = []
    for soc6 in individual_socs:
        # O*NET uses .00 suffix for base SOC codes
        for (s, tid), task_info in tasks.items():
            if s == soc6:
                onet_soc = task_info["onet_soc"]
                rating_key = (onet_soc, tid)
                r = ratings.get(rating_key, {})

                gwa = get_gwa_for_task(soc6, tid, dwa_to_gwa, task_dwas)

                collected.append({
                    "source_soc": soc6,
                    "onet_task_id": tid,
                    "task_text": task_info["task_text"],
                    "task_type": task_info["task_type"],
                    "importance": r.get("importance"),
                    "frequency": r.get("frequency"),
                    "relevance": r.get("relevance"),
                    "gwa": gwa,
                })

    # Sort by importance (descending), then relevance
    collected.sort(
        key=lambda t: (-(t["importance"] or 0), -(t["relevance"] or 0))
    )

    return collected


def main():
    print("Loading O*NET data...")
    tasks, ratings, dwa_to_gwa, task_dwas = load_onet_data()
    print(f"  Task Statements: {len(tasks)}")
    print(f"  Task Ratings: {len(ratings)}")
    print(f"  DWA→GWA mappings: {len(dwa_to_gwa)}")

    print("Loading Staffing Patterns entries...")
    entries = load_staffing_entries()
    print(f"  SOC entries: {len(entries)}")
    merged = sum(1 for e in entries.values() if e["is_merged"])
    print(f"  Single: {len(entries) - merged}, Merged: {merged}")

    print("Building task sets...")
    output = []
    onet_covered = 0
    llm_needed = 0

    for soc_code, entry in sorted(entries.items()):
        task_set = build_task_set(entry, tasks, ratings, dwa_to_gwa, task_dwas)

        record = {
            "soc_code": entry["soc_code"],
            "soc_title": entry["soc_title"],
            "is_merged": entry["is_merged"],
            "individual_socs": entry["individual_socs"],
            "total_employment_K": round(entry["total_employment_K"], 1),
            "sector_count": len(entry["sectors"]),
            "onet_task_count": len(task_set),
            "source": "onet" if task_set else "llm_generate",
            "tasks": task_set,
        }

        if task_set:
            onet_covered += 1
        else:
            llm_needed += 1

        output.append(record)

    print(f"\nResults:")
    print(f"  O*NET covered: {onet_covered}")
    print(f"  LLM generation needed: {llm_needed}")

    # Show LLM-needed entries
    if llm_needed:
        print(f"\n  SOCs needing LLM generation:")
        for rec in output:
            if rec["source"] == "llm_generate":
                print(f"    {rec['soc_code']}: {rec['soc_title']}")

    # Task count stats for O*NET-covered entries
    onet_counts = [r["onet_task_count"] for r in output if r["source"] == "onet"]
    print(f"\n  O*NET task counts: min={min(onet_counts)}, max={max(onet_counts)}, avg={sum(onet_counts)/len(onet_counts):.1f}")

    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    with open(OUTPUT, "w") as f:
        json.dump(output, f, indent=2)
    print(f"\nSaved to {OUTPUT}")


if __name__ == "__main__":
    main()
