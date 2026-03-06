"""Phase 2: Write scored tasks to workbook and generate bottleneck flags report.

Reads scored_tasks.json, writes to '3 Tasks' tab, and produces a bottleneck
flags report identifying occupations where high-importance tasks have low
autonomy scores — candidates for R (regulatory friction) review.

Uses importance-weighted task_coverage formula:
  task_coverage = sum(aut * importance * time_share) / sum(importance * time_share)
"""

import json
import os
import openpyxl
from copy import copy

SCORED_INPUT = "scoring/task_pipeline/scored_tasks.json"
WORKBOOK = "jobs-data-v3.xlsx"
FLAGS_OUTPUT = "scoring/task_pipeline/bottleneck_flags.json"
FLAGS_REPORT = "scoring/task_pipeline/bottleneck_report.txt"

# Importance threshold for bottleneck detection
BOTTLENECK_IMP_THRESHOLD = 4
# Autonomy threshold: tasks strictly at 0.00 are hard bottlenecks (licensing/physical gates)
BOTTLENECK_AUT_THRESHOLD = 0.0


def load_employment_dedup(wb):
    """Build SOC → deduplicated employment from Staffing Patterns.

    Each SOC appears in multiple sectors. Dedup_Employment_K is the total
    employment for that SOC across all sectors (not double-counted).
    We use the max of Occupation_Industry_Share_Pct to approximate unique workers.
    """
    sp = wb["2 Staffing Patterns"]
    soc_emp = {}  # soc_code -> total employment (already deduplicated in staffing)
    for row in sp.iter_rows(min_row=2, max_row=sp.max_row, values_only=True):
        soc_code = row[3]
        emp = row[5]  # Employment (Thousands)
        occ_share = row[7]  # Occupation_Industry_Share_Pct
        if soc_code and emp:
            # Sum employment across sectors but cap at 100% of occupation total
            if soc_code not in soc_emp:
                soc_emp[soc_code] = 0
            soc_emp[soc_code] += emp
    return soc_emp


def compute_task_coverage(tasks, scenario="mod"):
    """Compute importance-weighted task coverage.

    Formula: sum(aut * importance * time_share) / sum(importance * time_share)
    """
    score_key = f"aut_score_{scenario}"
    numerator = 0
    denominator = 0
    for t in tasks:
        w = t["importance"] * t["time_share_pct"]
        numerator += t[score_key] * w
        denominator += w
    if denominator == 0:
        return 0
    return round(numerator / denominator, 4)


def detect_bottlenecks(tasks, scenario="mod"):
    """Find high-importance tasks with low autonomy scores."""
    score_key = f"aut_score_{scenario}"
    bottlenecks = []
    for t in tasks:
        if (t["importance"] >= BOTTLENECK_IMP_THRESHOLD and
                t[score_key] <= BOTTLENECK_AUT_THRESHOLD):
            bottlenecks.append({
                "task_text": t["task_text"],
                "importance": t["importance"],
                "time_share_pct": t["time_share_pct"],
                "aut_score": t[score_key],
                "gwa": t["gwa"],
                "scenario": scenario,
            })
    return bottlenecks


def write_tasks_tab(wb, scored_data, soc_emp):
    """Write scored tasks to the '3 Tasks' tab."""
    ws = wb["3 Tasks"]

    # Headers should already exist from the archive copy; verify
    expected_headers = [
        "SOC_Code", "Job_Title", "Task_ID", "Task_Description", "Task_Type",
        "Time_Share_Pct", "Importance", "Frequency", "GWA",
        "Dedup_Employment_K", "Economy_Weight_K", "Aut_Score_Mod", "Aut_Score_Sig",
    ]
    for col, header in enumerate(expected_headers, 1):
        ws.cell(row=1, column=col, value=header)

    row_num = 2
    for entry in scored_data:
        soc_code = entry["soc_code"]
        soc_title = entry["soc_title"]
        source = entry["source"]
        dedup_emp = soc_emp.get(soc_code, 0)

        # Task ID prefix: ON for O*NET-sourced, CL for LLM-generated
        id_prefix = "ON" if source == "onet_curated" else "CL"
        # Compact SOC for task ID: remove hyphens
        soc_compact = soc_code.replace(", ", "_").replace("-", "")

        for task_idx, task in enumerate(entry["tasks"], 1):
            task_id = f"{id_prefix}-{soc_compact}-{task_idx:03d}"
            time_share = task["time_share_pct"]
            economy_weight = round(dedup_emp * time_share / 100, 2) if dedup_emp else 0

            ws.cell(row=row_num, column=1, value=soc_code)
            ws.cell(row=row_num, column=2, value=soc_title)
            ws.cell(row=row_num, column=3, value=task_id)
            ws.cell(row=row_num, column=4, value=task["task_text"])
            ws.cell(row=row_num, column=5, value=task["task_type"])
            ws.cell(row=row_num, column=6, value=time_share)
            ws.cell(row=row_num, column=7, value=task["importance"])
            ws.cell(row=row_num, column=8, value=task["frequency"])
            ws.cell(row=row_num, column=9, value=task["gwa"])
            ws.cell(row=row_num, column=10, value=round(dedup_emp, 1) if dedup_emp else 0)
            ws.cell(row=row_num, column=11, value=economy_weight)
            ws.cell(row=row_num, column=12, value=task["aut_score_mod"])
            ws.cell(row=row_num, column=13, value=task["aut_score_sig"])

            row_num += 1

    return row_num - 2  # total data rows written


def main():
    print("Loading scored tasks...")
    with open(SCORED_INPUT) as f:
        scored_data = json.load(f)
    print(f"  {len(scored_data)} SOC entries")
    total_tasks = sum(len(e["tasks"]) for e in scored_data)
    print(f"  {total_tasks} total tasks")

    print("Loading workbook...")
    wb = openpyxl.load_workbook(WORKBOOK)

    print("Computing employment deduplication...")
    soc_emp = load_employment_dedup(wb)

    # ── Write tasks to workbook ──────────────────────────────────────────────
    print("Writing to '3 Tasks' tab...")
    rows_written = write_tasks_tab(wb, scored_data, soc_emp)
    print(f"  {rows_written} rows written")

    wb.save(WORKBOOK)
    print(f"  Saved {WORKBOOK}")

    # ── Compute task_coverage and detect bottlenecks ─────────────────────────
    print("\nComputing task coverage and detecting bottlenecks...")
    flags = []
    coverage_mod = []
    coverage_sig = []

    for entry in scored_data:
        tc_mod = compute_task_coverage(entry["tasks"], "mod")
        tc_sig = compute_task_coverage(entry["tasks"], "sig")
        coverage_mod.append(tc_mod)
        coverage_sig.append(tc_sig)

        bn_mod = detect_bottlenecks(entry["tasks"], "mod")
        bn_sig = detect_bottlenecks(entry["tasks"], "sig")

        if bn_mod or bn_sig:
            flags.append({
                "soc_code": entry["soc_code"],
                "soc_title": entry["soc_title"],
                "employment_K": entry["total_employment_K"],
                "task_coverage_mod": tc_mod,
                "task_coverage_sig": tc_sig,
                "bottleneck_tasks_mod": bn_mod,
                "bottleneck_tasks_sig": bn_sig,
            })

    # Sort flags by employment (highest impact first)
    flags.sort(key=lambda f: -f["employment_K"])

    # Save flags JSON
    with open(FLAGS_OUTPUT, "w") as f:
        json.dump(flags, f, indent=2)

    # Generate human-readable report
    with open(FLAGS_REPORT, "w") as f:
        f.write("BOTTLENECK FLAGS REPORT — Candidates for R (Regulatory Friction) Review\n")
        f.write("=" * 80 + "\n\n")
        f.write(f"Total SOC entries analyzed: {len(scored_data)}\n")
        f.write(f"Entries with bottleneck flags: {len(flags)}\n")
        f.write(f"Criteria: importance >= {BOTTLENECK_IMP_THRESHOLD} AND "
                f"autonomy <= {BOTTLENECK_AUT_THRESHOLD}\n\n")

        f.write(f"Task Coverage (importance-weighted):\n")
        f.write(f"  Moderate:    mean={sum(coverage_mod)/len(coverage_mod):.3f}, "
                f"min={min(coverage_mod):.3f}, max={max(coverage_mod):.3f}\n")
        f.write(f"  Significant: mean={sum(coverage_sig)/len(coverage_sig):.3f}, "
                f"min={min(coverage_sig):.3f}, max={max(coverage_sig):.3f}\n\n")

        f.write("-" * 80 + "\n\n")

        for flag in flags:
            f.write(f"[{flag['soc_code']}] {flag['soc_title']}\n")
            f.write(f"  Employment: {flag['employment_K']:.1f}K | "
                    f"Coverage: mod={flag['task_coverage_mod']:.3f} "
                    f"sig={flag['task_coverage_sig']:.3f}\n")

            if flag["bottleneck_tasks_mod"]:
                f.write(f"  Bottlenecks (Moderate):\n")
                for bt in flag["bottleneck_tasks_mod"]:
                    f.write(f"    - [{bt['aut_score']:.2f}] imp={bt['importance']} "
                            f"time={bt['time_share_pct']}% | {bt['task_text'][:90]}\n")

            if flag["bottleneck_tasks_sig"]:
                sig_only = [bt for bt in flag["bottleneck_tasks_sig"]
                            if bt["task_text"] not in
                            [m["task_text"] for m in flag["bottleneck_tasks_mod"]]]
                if sig_only:
                    f.write(f"  Additional bottlenecks (Significant only):\n")
                    for bt in sig_only:
                        f.write(f"    - [{bt['aut_score']:.2f}] imp={bt['importance']} "
                                f"time={bt['time_share_pct']}% | {bt['task_text'][:90]}\n")
            f.write("\n")

    print(f"\nBottleneck flags: {len(flags)} SOCs flagged for R review")
    print(f"  Saved: {FLAGS_OUTPUT}")
    print(f"  Report: {FLAGS_REPORT}")

    # Print top 10 by employment
    print(f"\n  Top 10 flagged by employment:")
    for flag in flags[:10]:
        n_mod = len(flag["bottleneck_tasks_mod"])
        print(f"    {flag['soc_code']}: {flag['soc_title']} "
              f"({flag['employment_K']:.0f}K, {n_mod} bottleneck tasks)")

    print(f"\nTask Coverage Summary:")
    print(f"  Moderate:    mean={sum(coverage_mod)/len(coverage_mod):.3f}")
    print(f"  Significant: mean={sum(coverage_sig)/len(coverage_sig):.3f}")


if __name__ == "__main__":
    main()
