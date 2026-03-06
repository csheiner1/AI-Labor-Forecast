"""Assign Time_Share_Pct to tasks in the workbook using Claude Opus.

Reads the 3 Tasks tab, groups tasks by SOC, sends each SOC's task list to the
LLM for time-share allocation, validates/normalizes results, and writes back
Time_Share_Pct and Economy_Weight_K to the workbook.

Uses parallel batching with retry logic and checkpoint/resume support.

Output: scoring/task_pipeline/time_share_results.json
"""

import json
import os
import sys
import time
import openpyxl
import anthropic
from concurrent.futures import ThreadPoolExecutor, as_completed

WORKBOOK = "jobs-data-v3.xlsx"
ONET_RAW = "scoring/task_pipeline/onet_raw.json"
RESULTS_FILE = "scoring/task_pipeline/time_share_results.json"
MODEL = "claude-opus-4-6"
MAX_WORKERS = 4
MAX_RETRIES = 3


# ── Prompt templates ──────────────────────────────────────────────────────────

SYSTEM_PROMPT = """You are an occupational analyst specializing in AI labor displacement research.
You produce structured JSON output. Never include markdown formatting or commentary outside the JSON."""

USER_PROMPT = """Assign time-share percentages to the task inventory for the occupation below.

## Occupation
- SOC Code: {soc_code}
- Title: {job_title}
- Employment: {employment_K}K workers

## Tasks ({task_count} total)
{task_list}

## Instructions

For EACH task above, return a JSON object with exactly two fields:
- "task_id": The task_id exactly as shown above
- "time_share_pct": Integer percentage of typical work time spent on this task

Rules:
1. All time_share_pct values must sum to exactly 100.
2. Use frequency, importance, and task type to guide allocation:
   - Daily core tasks with high importance get larger shares (15-25%)
   - Weekly tasks get moderate shares (5-15%)
   - Monthly/yearly or supplemental tasks get smaller shares (2-8%)
3. No single task should exceed 35%.
4. Every task must get at least 1%.
5. Core tasks should collectively account for 70-90% of total time.

Return ONLY a JSON array of objects:
```json
[
  {{"task_id": "ON-111011-001", "time_share_pct": 15}},
  {{"task_id": "ON-111011-002", "time_share_pct": 10}}
]
```"""


# ── Input extraction ─────────────────────────────────────────────────────────

def load_frequency_lookup():
    """Build task_text -> frequency dict from onet_raw.json."""
    with open(ONET_RAW) as f:
        entries = json.load(f)
    lookup = {}
    for entry in entries:
        for t in entry["tasks"]:
            text = t["task_text"].strip()
            freq = t.get("frequency")
            if text and freq:
                lookup[text] = freq
    print(f"  Frequency lookup: {len(lookup)} tasks with frequency data")
    return lookup


def load_tasks_from_workbook():
    """Read 3 Tasks tab, group by SOC_Code.

    Returns dict keyed by soc_code, each value is a dict with:
      job_title, employment_K, tasks (list of dicts with row, task_id,
      task_description, task_type, importance, gwa)
    """
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    ws = wb["3 Tasks"]

    soc_entries = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        vals = [cell.value for cell in row]
        soc_code = vals[0]
        if not soc_code:
            continue

        if soc_code not in soc_entries:
            soc_entries[soc_code] = {
                "soc_code": soc_code,
                "job_title": vals[1],
                "employment_K": vals[8],  # Col 9: Employment_K
                "tasks": [],
            }

        soc_entries[soc_code]["tasks"].append({
            "row": row_idx,
            "task_id": vals[2],         # Col 3: Task_ID
            "task_description": vals[3], # Col 4: Task_Description
            "task_type": vals[4],        # Col 5: Task_Type
            "importance": vals[6],       # Col 7: Importance
            "gwa": vals[7],             # Col 8: GWA
        })

    wb.close()
    print(f"  Loaded {len(soc_entries)} SOC entries, {sum(len(e['tasks']) for e in soc_entries.values())} tasks")
    return soc_entries


def enrich_with_frequency(soc_entries, freq_lookup):
    """Add frequency to each task from the frequency lookup."""
    matched = 0
    total = 0
    for entry in soc_entries.values():
        for task in entry["tasks"]:
            total += 1
            text = (task["task_description"] or "").strip()
            freq = freq_lookup.get(text)
            task["frequency"] = freq if freq else "N/A"
            if freq:
                matched += 1
    print(f"  Frequency enrichment: {matched}/{total} tasks matched ({100*matched/total:.1f}%)")


# ── Prompt building ──────────────────────────────────────────────────────────

def format_task_list(tasks):
    """Format tasks as a numbered list for the prompt."""
    lines = []
    for i, t in enumerate(tasks, 1):
        tid = t["task_id"]
        ttype = t["task_type"] or "n/a"
        imp = t["importance"] if t["importance"] is not None else "N/A"
        freq = t["frequency"]
        gwa = t["gwa"] or "N/A"
        desc = t["task_description"]
        lines.append(
            f"{i}. task_id={tid}  type={ttype}  importance={imp}  frequency={freq}  gwa={gwa}\n"
            f"   {desc}"
        )
    return "\n".join(lines)


# ── LLM calling ──────────────────────────────────────────────────────────────

def call_opus(client, soc_entry):
    """Call Claude Opus API with retry logic."""
    task_list = format_task_list(soc_entry["tasks"])
    prompt = USER_PROMPT.format(
        soc_code=soc_entry["soc_code"],
        job_title=soc_entry["job_title"],
        employment_K=soc_entry["employment_K"],
        task_count=len(soc_entry["tasks"]),
        task_list=task_list,
    )
    soc_code = soc_entry["soc_code"]

    for attempt in range(MAX_RETRIES):
        try:
            response = client.messages.create(
                model=MODEL,
                max_tokens=4096,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.2,
            )
            text = response.content[0].text.strip()

            # Extract JSON from response (handle markdown wrapping)
            if text.startswith("```"):
                text = text.split("```")[1]
                if text.startswith("json"):
                    text = text[4:]
            text = text.strip()

            result = json.loads(text)
            return result

        except json.JSONDecodeError as e:
            print(f"  [{soc_code}] JSON parse error: {e}")
            if attempt < MAX_RETRIES - 1:
                time.sleep(2)
                continue
            return None

        except anthropic.RateLimitError:
            wait = 30 * (attempt + 1)
            print(f"  [{soc_code}] Rate limited, waiting {wait}s...")
            time.sleep(wait)
            continue

        except anthropic.APIError as e:
            print(f"  [{soc_code}] API error: {e}")
            if attempt < MAX_RETRIES - 1:
                time.sleep(5)
                continue
            return None

    return None


# ── Validation + normalization ───────────────────────────────────────────────

def validate_and_normalize(result, soc_entry):
    """Validate and normalize LLM response.

    Returns (id_to_share dict, warnings list, needs_retry bool).
    """
    warnings = []
    expected_ids = {t["task_id"] for t in soc_entry["tasks"]}
    expected_count = len(expected_ids)

    if not isinstance(result, list):
        return None, ["Response is not a list"], True

    # Check task count
    if len(result) != expected_count:
        warnings.append(f"Got {len(result)} tasks, expected {expected_count}")
        return None, warnings, True

    # Extract task_id -> time_share_pct mapping
    id_to_share = {}
    for item in result:
        tid = item.get("task_id")
        pct = item.get("time_share_pct")
        if tid is None or pct is None:
            warnings.append(f"Missing task_id or time_share_pct in item: {item}")
            continue
        if tid not in expected_ids:
            warnings.append(f"Unknown task_id: {tid}")
            return None, warnings, True
        if tid in id_to_share:
            warnings.append(f"Duplicate task_id: {tid}")
        id_to_share[tid] = pct

    # Check all expected IDs are present
    missing = expected_ids - set(id_to_share.keys())
    if missing:
        warnings.append(f"Missing task_ids: {missing}")
        return None, warnings, True

    # Check sum
    raw_sum = sum(id_to_share.values())
    deviation = abs(raw_sum - 100)

    if deviation > 15:
        warnings.append(f"Sum {raw_sum} deviates >15% from 100")
        return None, warnings, True

    if deviation > 2:
        warnings.append(f"Sum {raw_sum} deviates >2% from 100 — normalizing")

    # Normalize: scale proportionally
    if raw_sum != 100 and raw_sum > 0:
        factor = 100.0 / raw_sum
        for tid in id_to_share:
            id_to_share[tid] = max(1, round(id_to_share[tid] * factor))

        # Fix remainder on largest share (ensure it stays >= 1)
        current_sum = sum(id_to_share.values())
        remainder = 100 - current_sum
        if remainder != 0:
            largest_tid = max(id_to_share, key=id_to_share.get)
            id_to_share[largest_tid] = max(1, id_to_share[largest_tid] + remainder)

    # Warn on extreme values (but allow)
    for tid, pct in id_to_share.items():
        if pct < 1:
            warnings.append(f"{tid}: time_share_pct={pct} (<1%)")
        if pct > 40:
            warnings.append(f"{tid}: time_share_pct={pct} (>40%)")

    return id_to_share, warnings, False


# ── Checkpoint/resume ────────────────────────────────────────────────────────

def load_checkpoint():
    """Load completed results from checkpoint file.

    Returns dict keyed by soc_code.
    """
    if os.path.exists(RESULTS_FILE):
        with open(RESULTS_FILE) as f:
            data = json.load(f)
        # Handle both list format (sorted for output) and dict format
        if isinstance(data, list):
            return {item["soc_code"]: item for item in data}
        return data
    return {}


def save_checkpoint(completed):
    """Save results as sorted list to checkpoint file."""
    sorted_list = sorted(completed.values(), key=lambda x: x["soc_code"])
    with open(RESULTS_FILE, "w") as f:
        json.dump(sorted_list, f, indent=2)


# ── Process function ─────────────────────────────────────────────────────────

def process_soc(client, soc_entry):
    """Call Opus, validate, retry if needed. Return result dict or None."""
    soc_code = soc_entry["soc_code"]

    for attempt in range(MAX_RETRIES):
        raw = call_opus(client, soc_entry)
        if raw is None:
            if attempt < MAX_RETRIES - 1:
                print(f"  [{soc_code}] No response, retrying ({attempt+1}/{MAX_RETRIES})...")
                continue
            return None

        id_to_share, warnings, needs_retry = validate_and_normalize(raw, soc_entry)

        if warnings:
            for w in warnings[:3]:
                print(f"  [{soc_code}] {w}")

        if needs_retry:
            if attempt < MAX_RETRIES - 1:
                print(f"  [{soc_code}] Validation failed, retrying ({attempt+1}/{MAX_RETRIES})...")
                continue
            print(f"  [{soc_code}] FAILED after {MAX_RETRIES} attempts")
            return None

        return {
            "soc_code": soc_code,
            "job_title": soc_entry["job_title"],
            "task_count": len(id_to_share),
            "time_shares": id_to_share,
            "warnings": warnings,
        }

    return None


# ── Writeback ────────────────────────────────────────────────────────────────

def writeback(completed, soc_entries):
    """Write Time_Share_Pct and Economy_Weight_K back to the workbook.

    Only runs when all SOCs are complete.
    """
    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb["3 Tasks"]

    rows_written = 0
    for row_idx in range(2, ws.max_row + 1):
        soc_code = ws.cell(row=row_idx, column=1).value
        task_id = ws.cell(row=row_idx, column=3).value
        employment_k = ws.cell(row=row_idx, column=9).value

        if not soc_code or soc_code not in completed:
            continue

        shares = completed[soc_code]["time_shares"]
        if task_id not in shares:
            continue

        pct = shares[task_id]

        # Col 6: Time_Share_Pct
        ws.cell(row=row_idx, column=6).value = pct

        # Col 10: Economy_Weight_K = Employment_K * Time_Share_Pct / 100
        if employment_k is not None:
            ws.cell(row=row_idx, column=10).value = round(employment_k * pct / 100, 2)

        rows_written += 1

    wb.save(WORKBOOK)
    wb.close()

    # Summary stats
    all_pcts = []
    for entry in completed.values():
        all_pcts.extend(entry["time_shares"].values())

    print(f"\nWriteback complete:")
    print(f"  Rows written: {rows_written}")
    print(f"  SOCs: {len(completed)}")
    print(f"  Time_Share_Pct: min={min(all_pcts)}, max={max(all_pcts)}, "
          f"mean={sum(all_pcts)/len(all_pcts):.1f}")
    total_warnings = sum(len(e["warnings"]) for e in completed.values())
    print(f"  Total warnings: {total_warnings}")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("ERROR: ANTHROPIC_API_KEY environment variable not set.")
        sys.exit(1)

    client = anthropic.Anthropic()

    print("Loading workbook tasks...")
    soc_entries = load_tasks_from_workbook()

    print("Loading frequency lookup...")
    freq_lookup = load_frequency_lookup()

    print("Enriching tasks with frequency data...")
    enrich_with_frequency(soc_entries, freq_lookup)

    # Resume support
    completed = load_checkpoint()
    if completed:
        print(f"Resuming: {len(completed)} already done, {len(soc_entries) - len(completed)} remaining")

    remaining = [e for soc, e in sorted(soc_entries.items()) if soc not in completed]
    if not remaining:
        print("All SOCs already complete.")
        print("Running writeback...")
        writeback(completed, soc_entries)
        return

    print(f"\nProcessing {len(remaining)} SOCs with {MAX_WORKERS} parallel workers...\n")

    failed = []
    processed = 0
    completions_since_save = 0

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(process_soc, client, entry): entry
            for entry in remaining
        }

        for future in as_completed(futures):
            entry = futures[future]
            processed += 1

            try:
                result = future.result()
                if result:
                    completed[result["soc_code"]] = result
                    completions_since_save += 1
                    print(f"  [{processed}/{len(remaining)}] {entry['soc_code']}: "
                          f"{entry['job_title']} ({len(entry['tasks'])} tasks)")
                else:
                    failed.append(entry["soc_code"])
                    print(f"  [{processed}/{len(remaining)}] {entry['soc_code']}: FAILED")
            except Exception as e:
                failed.append(entry["soc_code"])
                print(f"  [{processed}/{len(remaining)}] {entry['soc_code']}: ERROR {e}")

            if completions_since_save >= 10:
                save_checkpoint(completed)
                completions_since_save = 0
                print(f"  --- Checkpoint saved ({len(completed)} complete, {len(failed)} failed) ---")

    # Final save
    save_checkpoint(completed)

    print(f"\nDone: {len(completed)} complete, {len(failed)} failed")
    if failed:
        print(f"Failed SOCs: {', '.join(sorted(failed))}")

    # Writeback only when all SOCs are complete
    if len(completed) == len(soc_entries):
        print("\nAll 310 SOCs complete. Running writeback...")
        writeback(completed, soc_entries)
    else:
        missing = len(soc_entries) - len(completed)
        print(f"\n{missing} SOCs still incomplete — skipping writeback.")
        print("Re-run to retry failed SOCs.")


if __name__ == "__main__":
    main()
