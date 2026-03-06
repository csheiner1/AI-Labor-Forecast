"""Re-run time_share assignment for K-12 teacher SOCs with instruction emphasis."""

import json
import os
import sys
import time
import openpyxl
import anthropic

WORKBOOK = "jobs-data-v3.xlsx"
ONET_RAW = "scoring/task_pipeline/onet_raw.json"
RESULTS_FILE = "scoring/task_pipeline/time_share_results.json"
MODEL = "claude-opus-4-6"

K12_SOCS = {
    "25-2011", "25-2012", "25-2021, 25-2022, 25-2031, 25-2032", "25-2022",
    "25-2023", "25-2051, 25-2052, 25-2059", "25-2057", "25-2058",
    "25-3011", "25-3021", "25-3041", "25-3099"
}

SYSTEM_PROMPT = (
    "You are an occupational analyst specializing in AI labor displacement research. "
    "You produce structured JSON output. Never include markdown formatting or commentary outside the JSON."
)

USER_PROMPT = """Assign time-share percentages to the task inventory for the occupation below.

## Occupation
- SOC Code: {soc_code}
- Title: {job_title}
- Employment: {employment_K}K workers

## Tasks ({task_count} total)
{task_list}

## Instructions

For EACH task above, return a JSON object with exactly two fields:
- "task_id": The task_id exactly as shown above
- "time_share_pct": Integer percentage of typical work time spent on this task

Rules:
1. All time_share_pct values must sum to exactly 100.
2. CRITICAL FOR TEACHING OCCUPATIONS: Direct instruction (teaching, lecturing, leading
   discussions, demonstrating) is THE primary activity. The single main instruction task
   should get 25-35% of time. Teachers spend the majority of their workday in front of
   students.
3. Related but distinct teaching activities (adapting methods, planning lessons, grading)
   are supporting activities — important, but secondary to actual instruction time.
4. Use frequency, importance, and task type to guide allocation for non-instruction tasks.
5. Every task must get at least 1%.
6. Core tasks should collectively account for 70-90% of total time.

Return ONLY a JSON array of objects:
```json
[
  {{"task_id": "ON-252022-001", "time_share_pct": 30}},
  {{"task_id": "ON-252022-002", "time_share_pct": 10}}
]
```"""


def main():
    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("ERROR: ANTHROPIC_API_KEY not set.")
        sys.exit(1)

    # Load frequency lookup
    with open(ONET_RAW) as f:
        raw_entries = json.load(f)
    freq_lookup = {}
    for entry in raw_entries:
        for t in entry.get("tasks", []):
            text = t["task_text"].strip()
            freq = t.get("frequency")
            if text and freq:
                freq_lookup[text] = freq

    # Load workbook tasks for K-12 SOCs
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    ws = wb["3 Tasks"]

    soc_entries = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        vals = [cell.value for cell in row]
        soc_code = vals[0]
        if not soc_code or soc_code not in K12_SOCS:
            continue
        if soc_code not in soc_entries:
            soc_entries[soc_code] = {
                "soc_code": soc_code,
                "job_title": vals[1],
                "employment_K": vals[8],
                "tasks": [],
            }
        freq = freq_lookup.get((vals[3] or "").strip(), "N/A")
        soc_entries[soc_code]["tasks"].append({
            "row": row_idx,
            "task_id": vals[2],
            "task_description": vals[3],
            "task_type": vals[4],
            "importance": vals[6],
            "gwa": vals[7],
            "frequency": freq,
        })
    wb.close()

    print(f"Loaded {len(soc_entries)} K-12 SOCs")

    def format_task_list(tasks):
        lines = []
        for i, t in enumerate(tasks, 1):
            tid = t["task_id"]
            ttype = t["task_type"] or "n/a"
            imp = t["importance"] if t["importance"] is not None else "N/A"
            freq = t["frequency"]
            gwa = t["gwa"] or "N/A"
            desc = t["task_description"]
            lines.append(
                f"{i}. task_id={tid}  type={ttype}  importance={imp}  "
                f"frequency={freq}  gwa={gwa}\n   {desc}"
            )
        return "\n".join(lines)

    # Call Opus for each SOC
    client = anthropic.Anthropic()
    results = {}
    failed = []

    for soc_code, entry in sorted(soc_entries.items()):
        task_list = format_task_list(entry["tasks"])
        prompt = USER_PROMPT.format(
            soc_code=entry["soc_code"],
            job_title=entry["job_title"],
            employment_K=entry["employment_K"],
            task_count=len(entry["tasks"]),
            task_list=task_list,
        )

        success = False
        for attempt in range(3):
            try:
                response = client.messages.create(
                    model=MODEL,
                    max_tokens=4096,
                    system=SYSTEM_PROMPT,
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0.2,
                )
                text = response.content[0].text.strip()
                if text.startswith("```"):
                    text = text.split("```")[1]
                    if text.startswith("json"):
                        text = text[4:]
                text = text.strip()
                result = json.loads(text)

                # Validate
                expected_ids = {t["task_id"] for t in entry["tasks"]}
                got_ids = {item["task_id"] for item in result}
                if got_ids != expected_ids or len(result) != len(expected_ids):
                    print(f"  [{soc_code}] ID mismatch, retrying...")
                    continue

                id_to_share = {item["task_id"]: item["time_share_pct"] for item in result}
                raw_sum = sum(id_to_share.values())

                if abs(raw_sum - 100) > 15:
                    print(f"  [{soc_code}] Sum {raw_sum} too far off, retrying...")
                    continue

                # Normalize if needed
                if raw_sum != 100 and raw_sum > 0:
                    factor = 100.0 / raw_sum
                    for tid in id_to_share:
                        id_to_share[tid] = max(1, round(id_to_share[tid] * factor))
                    remainder = 100 - sum(id_to_share.values())
                    if remainder != 0:
                        largest = max(id_to_share, key=id_to_share.get)
                        id_to_share[largest] = max(1, id_to_share[largest] + remainder)

                results[soc_code] = {
                    "soc_code": soc_code,
                    "job_title": entry["job_title"],
                    "task_count": len(id_to_share),
                    "time_shares": id_to_share,
                    "warnings": [f"Sum was {raw_sum}"] if raw_sum != 100 else [],
                }

                # Show top instruction task
                top_tid = max(id_to_share, key=id_to_share.get)
                top_desc = next(
                    t["task_description"][:60]
                    for t in entry["tasks"]
                    if t["task_id"] == top_tid
                )
                print(
                    f"  {soc_code}: {entry['job_title']} — "
                    f"top={id_to_share[top_tid]}% ({top_desc})"
                )
                success = True
                break

            except anthropic.RateLimitError:
                wait = 30 * (attempt + 1)
                print(f"  [{soc_code}] Rate limited, waiting {wait}s...")
                time.sleep(wait)

            except Exception as e:
                print(f"  [{soc_code}] Error: {e}")
                time.sleep(5)

        if not success:
            failed.append(soc_code)
            print(f"  {soc_code}: FAILED")

    print(f"\nDone: {len(results)} complete, {len(failed)} failed")
    if failed:
        print(f"Failed: {failed}")

    # Update results file
    with open(RESULTS_FILE) as f:
        all_results = json.load(f)

    all_by_soc = {r["soc_code"]: r for r in all_results}
    for soc_code, result in results.items():
        all_by_soc[soc_code] = result

    sorted_results = sorted(all_by_soc.values(), key=lambda r: r["soc_code"])
    with open(RESULTS_FILE, "w") as f:
        json.dump(sorted_results, f, indent=2)
    print("Updated time_share_results.json")

    # Update workbook
    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb["3 Tasks"]
    updated = 0
    for row_idx in range(2, ws.max_row + 1):
        soc_code = ws.cell(row=row_idx, column=1).value
        task_id = ws.cell(row=row_idx, column=3).value
        emp_k = ws.cell(row=row_idx, column=9).value
        if soc_code in results and task_id:
            shares = results[soc_code]["time_shares"]
            if task_id in shares:
                pct = shares[task_id]
                ws.cell(row=row_idx, column=6).value = pct
                if emp_k is not None:
                    ws.cell(row=row_idx, column=10).value = round(emp_k * pct / 100, 2)
                updated += 1
    wb.save(WORKBOOK)
    wb.close()
    print(f"Updated {updated} rows in workbook")


if __name__ == "__main__":
    main()
