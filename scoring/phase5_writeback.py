"""Phase 5: Write final scores back to the 4 Results tab in jobs-data-v3.xlsx."""
import json
import openpyxl

with open('scoring/final_scores.json') as f:
    scores = json.load(f)

# Clean titles: API sometimes appends " (Sector Name)" to titles
import re
def clean_title(t):
    """Strip trailing parenthetical sector names."""
    return re.sub(r'\s*\([^)]*\)\s*$', '', t).strip()

# Build lookup by (soc_code, custom_title) and also by clean title
score_map = {}
score_by_title = {}
score_by_clean_title = {}
for s in scores:
    raw_title = s['custom_title']
    cleaned = clean_title(raw_title)
    key = (s.get('soc_code', ''), raw_title)
    score_map[key] = s
    score_by_title[raw_title] = s
    score_by_clean_title[cleaned] = s

wb = openpyxl.load_workbook('jobs-data-v3.xlsx')
ws = wb['4 Results']

# Get header positions
headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
col_ws = headers['workflow_simplicity']  # column 8
col_xscale = headers['x_scale']          # column 11
col_xsub = headers['x_sub']              # column 12

print(f"Writing to columns: workflow_simplicity={col_ws}, x_scale={col_xscale}, x_sub={col_xsub}")

matched = 0
unmatched = []
for r in range(2, ws.max_row + 1):
    soc = ws.cell(r, 1).value
    title = ws.cell(r, 2).value
    key = (soc, title)

    if key in score_map:
        s = score_map[key]
        ws.cell(r, col_ws).value = s['workflow_simplicity']
        ws.cell(r, col_xscale).value = s['x_scale']
        ws.cell(r, col_xsub).value = s['x_sub']
        matched += 1
    else:
        # Try title-only match, then cleaned title match
        s = score_by_title.get(title) or score_by_clean_title.get(title)
        if s:
            ws.cell(r, col_ws).value = s['workflow_simplicity']
            ws.cell(r, col_xscale).value = s['x_scale']
            ws.cell(r, col_xsub).value = s['x_sub']
            matched += 1
        else:
            unmatched.append(title)

print(f"Matched: {matched}/{ws.max_row - 1}")
if unmatched:
    print(f"Unmatched ({len(unmatched)}): {unmatched[:10]}...")

wb.save('jobs-data-v3.xlsx')
print("Saved jobs-data-v3.xlsx")
