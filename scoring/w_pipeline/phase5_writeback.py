"""Phase 5: Write final results to '5 Results' tab in jobs-data-v3.xlsx.

Combines:
- tc_adj (from tc_adj_results.json, computed from task data)
- w (from final_w_scores.json, LLM-scored)

Computes:
- a = tc_adj * w
- S(a) = a^k / (a^k + (1-a)^k)  where k=0.8

Reads d_max, E, T, R from frictions tabs and computes displacement.
"""
import json
import math
import os
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(SCRIPT_DIR))

WORKBOOK = os.path.join(PROJECT_ROOT, "jobs-data-v3.xlsx")
TC_ADJ_FILE = os.path.join(PROJECT_ROOT, "scoring", "tc_adj_results.json")
W_SCORES_FILE = os.path.join(SCRIPT_DIR, "final_w_scores.json")
DMAX_FILE = os.path.join(PROJECT_ROOT, "scoring", "dmax_results.json")
SIGMOID_K = 0.8


def load_dmax(filepath=None):
    """Load d_max values from dmax_results.json.

    Returns dict: sector_name -> d_max (float).
    """
    filepath = filepath or DMAX_FILE
    with open(filepath) as f:
        data = json.load(f)
    return {entry["sector"]: entry["d_max"] for entry in data}


def sigmoid(a, k=SIGMOID_K):
    """S-shaped transform: S(a) = a^k / (a^k + (1-a)^k)."""
    if a <= 0:
        return 0.0
    if a >= 1:
        return 1.0
    ak = a ** k
    one_minus_ak = (1 - a) ** k
    return ak / (ak + one_minus_ak)


def compute_T_18mo(T1, T2, T3, T4, alpha=1.2):
    """Compute T(18mo) from raw T sub-scores.

    D = avg(T1, T2, T3, T4)
    t0 = 1.5*D - 0.5*T4 - 0.5
    T(18mo) = 1 / (1 + exp(-alpha * (1.5 - t0)))
    """
    if any(x is None for x in [T1, T2, T3, T4]):
        return None
    D = (T1 + T2 + T3 + T4) / 4
    t0 = 1.5 * D - 0.5 * T4 - 0.5
    return 1 / (1 + math.exp(-alpha * (1.5 - t0)))


def compute_R(f1, f2, f3):
    """Compute R from raw friction sub-scores.

    F = f1 + f2 + f3
    R = 1 - 0.7*(F-3)/9
    """
    if any(x is None for x in [f1, f2, f3]):
        return None
    F = f1 + f2 + f3
    return 1 - 0.7 * (F - 3) / 9


# Frictions tab uses human-readable names; Jobs tab uses snake_case codes.
# Map from Jobs-tab code -> Frictions-tab name.
OCC_GROUP_NORMALIZE = {
    "Core": "Core",
    "G1_Exec_Management": "G1 Executive & Management",
    "G2_HR_People": "G2 HR & People Ops",
    "G3_Finance_Accounting": "G3 Finance & Accounting",
    "G4_IT_Digital": "G4 IT & Digital",
    "G5_Marketing_Creative": "G5 Marketing & Creative",
    "G6_Sales_BizDev": "G6 Sales & Business Dev",
    "G7_Legal_Compliance": "G7 Legal & Compliance",
    "G8_Procurement_Supply": "G8 Procurement & Supply Chain",
    "G9_Admin_Office": "G9 Admin & Office Support",
}


def load_frictions(wb):
    """Load friction parameters from 4L and 4H tabs.

    Reads raw T1-T4, f1-f3, E from ALL rows (Core + occupation groups),
    computes T(18mo) and R.
    Returns dict: scenario -> (sector, occ_group) -> {E, T_18mo, R}.
    Keys use the frictions-tab occ_group names (human-readable).
    """
    frictions = {"low": {}, "high": {}}

    for scenario, tab_name in [("low", "4L Frictions Low"), ("high", "4H Frictions High")]:
        try:
            ws = wb[tab_name]
        except KeyError:
            print(f"WARNING: Tab '{tab_name}' not found.")
            continue

        # Data starts at row 5, headers at row 4
        # Columns: 1=Sector_ID, 2=Sector, 3=Occ_Group, 4=Emp, 5=Wage,
        #          6=T1, 7=T2, 8=T3, 9=T4, 13=f1, 14=f2, 15=f3, 18=E
        for r in range(5, ws.max_row + 1):
            sector = ws.cell(r, 2).value
            occ_group = ws.cell(r, 3).value
            if not sector or not occ_group:
                continue

            T1 = ws.cell(r, 6).value
            T2 = ws.cell(r, 7).value
            T3 = ws.cell(r, 8).value
            T4 = ws.cell(r, 9).value
            f1 = ws.cell(r, 13).value
            f2 = ws.cell(r, 14).value
            f3 = ws.cell(r, 15).value
            E = ws.cell(r, 18).value

            # Skip rows with no scores at all
            if all(x is None for x in [T1, T2, T3, T4, f1, f2, f3, E]):
                continue

            T_18mo = compute_T_18mo(T1, T2, T3, T4)
            R = compute_R(f1, f2, f3)

            frictions[scenario][(sector, occ_group)] = {
                "E": E if E is not None else 1.0,
                "T_18mo": round(T_18mo, 4) if T_18mo is not None else 0.5,
                "R": round(R, 4) if R is not None else 1.0,
            }

        scored = len(frictions[scenario])
        print(f"  {scenario}: loaded {scored} (sector, occ_group) friction rows from '{tab_name}'")

    return frictions


def main():
    # Load d_max from JOLTS data
    dmax_scores = load_dmax()
    print(f"Loaded d_max for {len(dmax_scores)} sectors from {DMAX_FILE}")

    # Load inputs
    with open(TC_ADJ_FILE) as f:
        tc_adj_data = json.load(f)
    tc_adj_map = {d["soc"]: d for d in tc_adj_data}

    with open(W_SCORES_FILE) as f:
        w_scores = json.load(f)
    w_map = {s["soc_code"]: s["w"] for s in w_scores}
    # Also build title-based lookup for merged SOC entries
    w_by_title = {s["title"]: s["w"] for s in w_scores}

    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb["5 Results"]

    # Load sector mapping from Staffing Patterns
    # Columns: 1=Sector_ID, 2=Sector, 3=Occupation_Group, 4=SOC_Code,
    #          5=SOC_Title, 6=Employment(K), 7=Staffing_Share, 8=Occ_Industry_Share,
    #          9=Median_Wage, 10=Projected_Change
    ws_sp = wb["2 Jobs"]
    soc_rows = {}  # soc -> list of (sector, occ_group, title, emp, occ_share, wage)
    for r in range(2, ws_sp.max_row + 1):
        soc = ws_sp.cell(r, 4).value
        if not soc:
            continue
        sector = ws_sp.cell(r, 2).value
        occ_group = ws_sp.cell(r, 3).value
        title = ws_sp.cell(r, 5).value
        emp = ws_sp.cell(r, 6).value
        occ_share = ws_sp.cell(r, 8).value or 0
        wage = ws_sp.cell(r, 9).value
        soc_rows.setdefault(soc, []).append((sector, occ_group, title, emp, occ_share, wage))

    soc_meta = {}
    for soc, rows in soc_rows.items():
        best = max(rows, key=lambda x: x[4])  # highest occ_share = primary sector
        total_emp = sum(r[3] or 0 for r in rows)
        soc_meta[soc] = {
            "title": best[2],
            "sector": best[0],
            "occ_group": best[1],
            "employment_K": total_emp,
            "wage": best[5],
        }

    # Load frictions
    frictions = load_frictions(wb)

    # Capture max row before writing so we can clear leftover old data
    old_max_row = ws.max_row

    # Define new headers
    new_headers = [
        "SOC_Code", "Job_Title", "Sector", "Occupation_Group",
        "Employment_2024_K", "Median_Wage",
        "tc_adj_mod", "tc_adj_sig",
        "w",
        "a_mod", "a_sig",
        "S_mod", "S_sig",
        "d_max", "E",
        "T_18mo_low", "T_18mo_high",
        "R_low", "R_high",
        "d_mod_low", "d_mod_high",
        "d_sig_low", "d_sig_high",
        "displaced_K_mod_low", "displaced_K_mod_high",
        "displaced_K_sig_low", "displaced_K_sig_high",
    ]

    # Write headers
    for col, header in enumerate(new_headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Clear any extra old columns
    for col in range(len(new_headers) + 1, ws.max_column + 1):
        ws.cell(row=1, column=col, value=None)

    # Build rows for ALL tc_adj SOCs (357 = Staffing alignment).
    # SOCs without w scores get tc_adj but null displacement columns.
    all_socs = sorted(tc_adj_map.keys())
    w_matched = 0
    w_missing = []

    row = 2
    total_displaced = {"mod_high": 0, "sig_high": 0}
    friction_misses = []  # (soc, sector, occ_group) tuples with no friction match

    for soc in all_socs:
        tc = tc_adj_map[soc]

        meta = soc_meta.get(soc, {})
        # For merged SOCs (comma-separated), try first individual code
        if not meta and ", " in soc:
            first_soc = soc.split(", ")[0]
            meta = soc_meta.get(first_soc, {})
        sector = meta.get("sector", "Unknown")
        occ_group = meta.get("occ_group", "")

        tc_adj_mod = tc["tc_adj_mod"]
        tc_adj_sig = tc["tc_adj_sig"]

        # Try SOC code match first, then title match for w
        w = w_map.get(soc)
        if w is None:
            w = w_by_title.get(tc["title"])

        if w is not None:
            w_matched += 1
            a_mod = round(tc_adj_mod * w, 4)
            a_sig = round(tc_adj_sig * w, 4)
            s_mod = round(sigmoid(a_mod), 4)
            s_sig = round(sigmoid(a_sig), 4)

            # Look up frictions by (sector, occ_group) — no fallback
            # Normalize occ_group from Jobs-tab code to Frictions-tab name
            fric_occ_group = OCC_GROUP_NORMALIZE.get(occ_group, occ_group)
            high_frictions = frictions.get("high", {})
            f_high = high_frictions.get((sector, fric_occ_group))
            if f_high is None:
                friction_misses.append((soc, sector, occ_group))
                # Still compute with None markers so we can see the gap
                f_high = {"E": None, "T_18mo": None, "R": None}

            d_max = dmax_scores.get(sector, 1.0)
            E = f_high["E"]

            if E is not None and f_high["T_18mo"] is not None and f_high["R"] is not None:
                d_mod_high = round(d_max * s_mod * E * f_high["T_18mo"] * f_high["R"], 4)
                d_sig_high = round(d_max * s_sig * E * f_high["T_18mo"] * f_high["R"], 4)
            else:
                d_mod_high = None
                d_sig_high = None

            emp = meta.get("employment_K", 0) or 0

            if d_mod_high is not None:
                displaced_mod_high = round(emp * d_mod_high, 2)
                displaced_sig_high = round(emp * d_sig_high, 2)
                total_displaced["mod_high"] += displaced_mod_high
                total_displaced["sig_high"] += displaced_sig_high
            else:
                displaced_mod_high = None
                displaced_sig_high = None

            values = [
                soc, meta.get("title", tc["title"]), sector, occ_group,
                meta.get("employment_K", 0), meta.get("wage"),
                tc_adj_mod, tc_adj_sig,
                w,
                a_mod, a_sig,
                s_mod, s_sig,
                d_max, E,
                None, f_high["T_18mo"],  # T_18mo_low=None (not yet scored)
                None, f_high["R"],  # R_low=None (not yet scored)
                None, d_mod_high,  # d_mod_low=None
                None, d_sig_high,  # d_sig_low=None
                None, displaced_mod_high,  # displaced_K_mod_low=None
                None, displaced_sig_high,  # displaced_K_sig_low=None
            ]
        else:
            w_missing.append(f"{soc}: {tc['title']}")
            # Write tc_adj and metadata, leave w and downstream columns null
            values = [
                soc, meta.get("title", tc["title"]), sector, meta.get("occ_group", ""),
                meta.get("employment_K", 0), meta.get("wage"),
                tc_adj_mod, tc_adj_sig,
                None,  # w
                None, None,  # a_mod, a_sig
                None, None,  # S_mod, S_sig
                None, None,  # d_max, E
                None, None,  # T_18mo_low, T_18mo_high
                None, None,  # R_low, R_high
                None, None,  # d_mod_low, d_mod_high
                None, None,  # d_sig_low, d_sig_high
                None, None,  # displaced_K_mod_low, displaced_K_mod_high
                None, None,  # displaced_K_sig_low, displaced_K_sig_high
            ]

        for col, val in enumerate(values, 1):
            ws.cell(row=row, column=col, value=val)

        row += 1

    # Delete any remaining old data rows
    if row <= old_max_row:
        ws.delete_rows(row, old_max_row - row + 1)

    print(f"w matched: {w_matched}/{len(all_socs)}")
    if w_missing:
        print(f"w missing ({len(w_missing)}): {w_missing[:5]}...")

    if friction_misses:
        print(f"\n*** FRICTION MISSES: {len(friction_misses)} SOCs had no (sector, occ_group) match ***")
        print(f"  These SOCs have NULL displacement values.")
        for soc, sec, og in friction_misses:
            print(f"    {soc}: sector='{sec}', occ_group='{og}'")
    else:
        print(f"\nAll SOCs matched to friction (sector, occ_group) entries.")

    # ── Summary tables to the right of main data ──────────────────────────
    from collections import defaultdict

    # Accumulate by sector and occ_group
    by_sector = defaultdict(lambda: {"emp": 0, "disp_mod": 0, "disp_sig": 0})
    by_group = defaultdict(lambda: {"emp": 0, "disp_mod": 0, "disp_sig": 0})

    for r in range(2, row):
        sector = ws.cell(r, 3).value
        occ_group = ws.cell(r, 4).value
        emp = ws.cell(r, 5).value or 0
        disp_mod = ws.cell(r, 25).value or 0
        disp_sig = ws.cell(r, 27).value or 0

        by_sector[sector]["emp"] += emp
        by_sector[sector]["disp_mod"] += disp_mod
        by_sector[sector]["disp_sig"] += disp_sig
        by_group[occ_group]["emp"] += emp
        by_group[occ_group]["disp_mod"] += disp_mod
        by_group[occ_group]["disp_sig"] += disp_sig

    SUMMARY_START_COL = 29  # 1 col gap after main table

    # Table 1: By Sector
    summary_headers = ["Category", "Employment_K", "Displaced_Mod_K", "Displaced_Sig_K",
                        "d_rate_mod_pct", "d_rate_sig_pct"]
    ws.cell(row=1, column=SUMMARY_START_COL, value="DISPLACEMENT BY SECTOR")
    for ci, h in enumerate(summary_headers):
        ws.cell(row=2, column=SUMMARY_START_COL + ci, value=h)

    sr = 3
    grand = {"emp": 0, "disp_mod": 0, "disp_sig": 0}
    for sector in sorted(by_sector.keys()):
        d = by_sector[sector]
        grand["emp"] += d["emp"]
        grand["disp_mod"] += d["disp_mod"]
        grand["disp_sig"] += d["disp_sig"]
        rate_mod = round(d["disp_mod"] / d["emp"] * 100, 2) if d["emp"] > 0 else 0
        rate_sig = round(d["disp_sig"] / d["emp"] * 100, 2) if d["emp"] > 0 else 0
        vals = [sector, round(d["emp"], 1), round(d["disp_mod"], 1),
                round(d["disp_sig"], 1), rate_mod, rate_sig]
        for ci, v in enumerate(vals):
            ws.cell(row=sr, column=SUMMARY_START_COL + ci, value=v)
        sr += 1

    # Grand total row
    rate_mod = round(grand["disp_mod"] / grand["emp"] * 100, 2) if grand["emp"] > 0 else 0
    rate_sig = round(grand["disp_sig"] / grand["emp"] * 100, 2) if grand["emp"] > 0 else 0
    for ci, v in enumerate(["TOTAL", round(grand["emp"], 1), round(grand["disp_mod"], 1),
                             round(grand["disp_sig"], 1), rate_mod, rate_sig]):
        ws.cell(row=sr, column=SUMMARY_START_COL + ci, value=v)
    sr += 2  # blank row

    # Table 2: By Occupation Group
    ws.cell(row=sr, column=SUMMARY_START_COL, value="DISPLACEMENT BY OCCUPATION GROUP")
    sr += 1
    for ci, h in enumerate(summary_headers):
        ws.cell(row=sr, column=SUMMARY_START_COL + ci, value=h)
    sr += 1

    for group in sorted(by_group.keys()):
        d = by_group[group]
        rate_mod = round(d["disp_mod"] / d["emp"] * 100, 2) if d["emp"] > 0 else 0
        rate_sig = round(d["disp_sig"] / d["emp"] * 100, 2) if d["emp"] > 0 else 0
        vals = [group, round(d["emp"], 1), round(d["disp_mod"], 1),
                round(d["disp_sig"], 1), rate_mod, rate_sig]
        for ci, v in enumerate(vals):
            ws.cell(row=sr, column=SUMMARY_START_COL + ci, value=v)
        sr += 1

    # Grand total row
    for ci, v in enumerate(["TOTAL", round(grand["emp"], 1), round(grand["disp_mod"], 1),
                             round(grand["disp_sig"], 1), rate_mod, rate_sig]):
        ws.cell(row=sr, column=SUMMARY_START_COL + ci, value=v)

    # Clear any old summary data below
    for r in range(sr + 1, ws.max_row + 1):
        for c in range(SUMMARY_START_COL, SUMMARY_START_COL + len(summary_headers)):
            ws.cell(row=r, column=c, value=None)

    wb.save(WORKBOOK)
    rows_written = row - 2
    print(f"\nWritten {rows_written} rows to '5 Results' tab")
    print(f"Summary tables: {len(by_sector)} sectors, {len(by_group)} occupation groups")
    print(f"Saved {WORKBOOK}")

    print(f"\nDisplacement totals (high friction, thousands of workers):")
    print(f"  Moderate tech scenario:    {total_displaced['mod_high']:.1f}K")
    print(f"  Significant tech scenario: {total_displaced['sig_high']:.1f}K")


if __name__ == "__main__":
    main()
