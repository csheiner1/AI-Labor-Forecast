"""Multi-agent task autonomy scorer.

Reads 5,382 tasks from the workbook, splits 310 SOCs across 6 parallel agents
by sector grouping, scores aut_score_mod + aut_score_sig via Claude Opus,
and writes results back to the 3 Tasks tab.

Usage:
    python3 scoring/task_autonomy/run.py [--dry-run] [--agent X] [--soc 15-1252]

Options:
    --dry-run   Print plan without making API calls
    --agent X   Run only agent X (A-F) for testing
    --soc CODE  Run only one SOC for testing
"""

import argparse
import json
import os
import sys
import threading
import time
import anthropic
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from prompts import SYSTEM_PROMPT, build_scoring_prompt

# ── Config ───────────────────────────────────────────────────────────────────

WORKBOOK = "jobs-data-v3.xlsx"
RESULTS_FILE = "scoring/task_autonomy/results.json"
MODEL = "claude-opus-4-6"
TEMPERATURE = 0.2
MAX_TOKENS = 8192
MAX_RETRIES = 3
MAX_WORKERS = 6  # one per agent

VALID_SCORES = {0.0, 0.25, 0.5, 0.75, 1.0}
VALID_CONFIDENCE = {"high", "low"}

# ── Agent sector assignments ─────────────────────────────────────────────────

AGENT_SECTORS = {
    "A": ["Healthcare & Life Sciences"],
    "B": ["Education & Academia"],
    "C": ["Government & Public Administration"],
    "D": [
        "Finance & Banking", "Insurance", "Accounting & Tax Firms",
        "Law Firms & Legal Services", "Management Consulting Firms",
        "Real Estate & Property",
    ],
    "E": ["Technology & Software", "Architecture & Engineering Firms"],
    "F": [
        "Advertising & PR Agencies", "Media Publishing & Entertainment",
        "Manufacturing", "Retail Trade", "Staffing & Recruitment Agencies",
        "Construction", "Transportation & Logistics", "Energy & Utilities",
        "Wholesale Trade", "Accommodation & Food Services",
    ],
}

# Reverse: sector -> agent
SECTOR_TO_AGENT = {}
for agent, sectors in AGENT_SECTORS.items():
    for s in sectors:
        SECTOR_TO_AGENT[s] = agent


# ── Data loading ─────────────────────────────────────────────────────────────

def load_tasks():
    """Read tasks from 3 Tasks tab, grouped by SOC_Code.

    Returns dict: soc_code -> {job_title, employment_K, tasks: [...]}.
    """
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb["3 Tasks"]

    soc_entries = {}
    for row in ws.iter_rows(min_row=2, max_col=12, values_only=True):
        soc_code = row[0]
        if not soc_code:
            continue

        if soc_code not in soc_entries:
            soc_entries[soc_code] = {
                "soc_code": soc_code,
                "job_title": row[1],
                "employment_K": row[8],
                "tasks": [],
            }

        soc_entries[soc_code]["tasks"].append({
            "task_id": row[2],
            "task_description": row[3],
            "task_type": row[4],
            "time_share_pct": row[5],
            "importance": row[6],
            "gwa": row[7],
        })

    wb.close()
    return soc_entries


def load_primary_sectors():
    """For each SOC, find the sector where it has the highest employment.

    Returns dict: soc_code -> sector_name.
    """
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb["2 Staffing Patterns"]

    soc_sector_emp = defaultdict(lambda: defaultdict(float))
    for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
        if not row[0] or not row[3]:
            continue
        sector = row[1]
        soc = row[3]
        emp = float(row[5]) if row[5] else 0
        soc_sector_emp[soc][sector] += emp

    wb.close()

    primary = {}
    for soc, sectors in soc_sector_emp.items():
        primary[soc] = max(sectors, key=sectors.get)
    return primary


def assign_agents(soc_entries, primary_sectors):
    """Assign each SOC to an agent based on its primary sector.

    Returns dict: agent_id -> [soc_code, ...].
    """
    agent_socs = defaultdict(list)
    unassigned = []

    for soc in sorted(soc_entries.keys()):
        sector = primary_sectors.get(soc)
        if sector and sector in SECTOR_TO_AGENT:
            agent = SECTOR_TO_AGENT[sector]
        else:
            # Fallback: assign to agent F (catch-all)
            agent = "F"
            if sector:
                unassigned.append((soc, sector))

        agent_socs[agent].append(soc)

    if unassigned:
        print(f"  {len(unassigned)} SOCs assigned to Agent F (unmapped sector):")
        for soc, sector in unassigned[:5]:
            print(f"    {soc}: {sector}")

    return dict(agent_socs)


# ── Checkpoint / resume ──────────────────────────────────────────────────────

def load_checkpoint():
    """Load completed results. Returns dict: soc_code -> scores list."""
    if os.path.exists(RESULTS_FILE):
        with open(RESULTS_FILE) as f:
            data = json.load(f)
        return {entry["soc_code"]: entry for entry in data}
    return {}


_checkpoint_lock = threading.Lock()


def save_checkpoint(completed):
    """Save results as sorted list (thread-safe)."""
    with _checkpoint_lock:
        sorted_list = sorted(completed.values(), key=lambda x: x["soc_code"])
        os.makedirs(os.path.dirname(RESULTS_FILE), exist_ok=True)
        with open(RESULTS_FILE, "w") as f:
            json.dump(sorted_list, f, indent=2)


# ── API scoring ──────────────────────────────────────────────────────────────

def score_soc(client, soc_entry):
    """Score one SOC's tasks via API. Returns result dict or None."""
    soc_code = soc_entry["soc_code"]
    prompt = build_scoring_prompt(
        soc_code,
        soc_entry["job_title"],
        soc_entry["employment_K"],
        soc_entry["tasks"],
    )

    for attempt in range(MAX_RETRIES):
        try:
            response = client.messages.create(
                model=MODEL,
                temperature=TEMPERATURE,
                max_tokens=MAX_TOKENS,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": prompt}],
            )

            text = response.content[0].text.strip()

            # Extract JSON
            if "```json" in text:
                text = text.split("```json")[1].split("```")[0]
            elif "```" in text:
                text = text.split("```")[1].split("```")[0]

            scores = json.loads(text.strip())
            validated = validate_scores(scores, soc_entry)
            if validated is None:
                if attempt < MAX_RETRIES - 1:
                    print(f"    [{soc_code}] Validation failed, retrying...")
                    time.sleep(2)
                    continue
                return None

            return {
                "soc_code": soc_code,
                "job_title": soc_entry["job_title"],
                "task_count": len(validated),
                "scores": validated,
                "usage": {
                    "input_tokens": response.usage.input_tokens,
                    "output_tokens": response.usage.output_tokens,
                },
            }

        except json.JSONDecodeError as e:
            print(f"    [{soc_code}] JSON parse error: {e}")
            if attempt < MAX_RETRIES - 1:
                time.sleep(2)
                continue
            return None

        except anthropic.RateLimitError:
            wait = 30 * (attempt + 1)
            print(f"    [{soc_code}] Rate limited, waiting {wait}s...")
            time.sleep(wait)
            continue

        except anthropic.APIError as e:
            print(f"    [{soc_code}] API error: {e}")
            if attempt < MAX_RETRIES - 1:
                time.sleep(5)
                continue
            return None

    return None


def validate_scores(scores, soc_entry):
    """Validate and fix API response. Returns list of score dicts or None."""
    if not isinstance(scores, list):
        print(f"    [{soc_entry['soc_code']}] Response not a list")
        return None

    expected_ids = {t["task_id"] for t in soc_entry["tasks"]}

    if len(scores) != len(expected_ids):
        print(f"    [{soc_entry['soc_code']}] Got {len(scores)} scores, "
              f"expected {len(expected_ids)}")
        # Allow if all expected IDs are present
        returned_ids = {s.get("task_id") for s in scores}
        if not expected_ids.issubset(returned_ids):
            return None

    validated = []
    for s in scores:
        tid = s.get("task_id")
        if tid not in expected_ids:
            continue  # skip extra entries

        mod = s.get("aut_score_mod")
        sig = s.get("aut_score_sig")

        # Fix scores to nearest valid value
        if isinstance(mod, (int, float)) and mod not in VALID_SCORES:
            mod = min(VALID_SCORES, key=lambda x: abs(x - mod))
        if isinstance(sig, (int, float)) and sig not in VALID_SCORES:
            sig = min(VALID_SCORES, key=lambda x: abs(x - sig))

        # Enforce sig >= mod
        if isinstance(mod, (int, float)) and isinstance(sig, (int, float)):
            if sig < mod:
                sig = mod

        # Confidence flags (default to "high" if missing)
        conf_mod = s.get("confidence_mod", "high")
        conf_sig = s.get("confidence_sig", "high")
        if conf_mod not in VALID_CONFIDENCE:
            conf_mod = "high"
        if conf_sig not in VALID_CONFIDENCE:
            conf_sig = "high"

        if mod not in VALID_SCORES or sig not in VALID_SCORES:
            print(f"    [{tid}] Invalid scores: mod={mod}, sig={sig}")
            return None

        validated.append({
            "task_id": tid,
            "aut_score_mod": mod,
            "confidence_mod": conf_mod,
            "aut_score_sig": sig,
            "confidence_sig": conf_sig,
        })

    # Check all expected IDs were scored
    scored_ids = {v["task_id"] for v in validated}
    missing = expected_ids - scored_ids
    if missing:
        print(f"    [{soc_entry['soc_code']}] Missing task_ids: {missing}")
        return None

    return validated


# ── Agent runner ─────────────────────────────────────────────────────────────

def run_agent(agent_id, soc_codes, soc_entries, completed):
    """Run scoring for one agent's SOC assignments. Returns list of results."""
    remaining = [soc for soc in soc_codes if soc not in completed]
    if not remaining:
        print(f"  Agent {agent_id}: all {len(soc_codes)} SOCs already complete")
        return []

    print(f"  Agent {agent_id}: {len(remaining)} SOCs to score "
          f"({len(soc_codes) - len(remaining)} already done)")

    client = anthropic.Anthropic()
    results = []

    for i, soc in enumerate(remaining):
        entry = soc_entries[soc]
        result = score_soc(client, entry)

        if result:
            results.append(result)
            # Incremental checkpoint: save immediately so progress survives crashes
            completed[soc] = result
            save_checkpoint(completed)
            status = f"{result['task_count']} tasks"
        else:
            status = "FAILED"

        print(f"    Agent {agent_id} [{i+1}/{len(remaining)}] "
              f"{soc}: {entry['job_title'][:35]} — {status}")

        # Brief pause between calls
        if i < len(remaining) - 1:
            time.sleep(0.5)

    return results


# ── Writeback ────────────────────────────────────────────────────────────────

def writeback(completed):
    """Write aut_score_mod and aut_score_sig to the 3 Tasks tab."""
    # Build task_id -> scores lookup
    score_map = {}
    for entry in completed.values():
        for s in entry["scores"]:
            score_map[s["task_id"]] = s

    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb["3 Tasks"]

    rows_written = 0
    for row_idx in range(2, ws.max_row + 1):
        task_id = ws.cell(row=row_idx, column=3).value
        if task_id and task_id in score_map:
            s = score_map[task_id]
            ws.cell(row=row_idx, column=11).value = s["aut_score_mod"]
            ws.cell(row=row_idx, column=12).value = s["aut_score_sig"]
            rows_written += 1

    wb.save(WORKBOOK)
    print(f"\nWriteback: {rows_written} rows updated in {WORKBOOK}")
    return rows_written


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Multi-agent task autonomy scorer")
    parser.add_argument("--dry-run", action="store_true", help="Print plan only")
    parser.add_argument("--agent", type=str, help="Run only agent X (A-F)")
    parser.add_argument("--soc", type=str, help="Run only one SOC code")
    parser.add_argument("--writeback-only", action="store_true",
                        help="Skip scoring, just write existing results to workbook")
    args = parser.parse_args()

    if not args.dry_run and not args.writeback_only:
        if not os.environ.get("ANTHROPIC_API_KEY"):
            print("ERROR: ANTHROPIC_API_KEY not set.")
            sys.exit(1)

    print("Loading tasks from workbook...")
    soc_entries = load_tasks()
    total_tasks = sum(len(e["tasks"]) for e in soc_entries.values())
    print(f"  {len(soc_entries)} SOCs, {total_tasks} tasks")

    print("Loading primary sectors...")
    primary_sectors = load_primary_sectors()

    print("Assigning agents...")
    agent_socs = assign_agents(soc_entries, primary_sectors)

    # Print plan
    print(f"\n{'='*60}")
    print(f"  TASK AUTONOMY SCORING — AGENT ASSIGNMENTS")
    print(f"{'='*60}")
    total_socs = 0
    total_agent_tasks = 0
    for agent_id in sorted(agent_socs.keys()):
        socs = agent_socs[agent_id]
        n_tasks = sum(len(soc_entries[s]["tasks"]) for s in socs)
        sectors = ", ".join(AGENT_SECTORS[agent_id])
        print(f"  Agent {agent_id}: {len(socs):3d} SOCs, {n_tasks:5d} tasks  "
              f"[{sectors[:60]}]")
        total_socs += len(socs)
        total_agent_tasks += n_tasks
    print(f"  {'─'*56}")
    print(f"  Total:  {total_socs:3d} SOCs, {total_agent_tasks:5d} tasks")
    print(f"  Model:  {MODEL}")
    print(f"{'='*60}\n")

    if args.dry_run:
        print("Dry run — exiting.")
        return

    # Load checkpoint
    completed = load_checkpoint()
    if completed:
        scored_tasks = sum(len(e["scores"]) for e in completed.values())
        print(f"Resuming: {len(completed)} SOCs already scored ({scored_tasks} tasks)")

    if args.writeback_only:
        if not completed:
            print("No results to write back.")
            return
        writeback(completed)
        return

    # Filter to specific agent or SOC if requested
    if args.soc:
        if args.soc not in soc_entries:
            print(f"SOC {args.soc} not found in workbook.")
            sys.exit(1)
        agent_socs = {"X": [args.soc]}
        print(f"Single SOC mode: {args.soc}")

    if args.agent:
        if args.agent not in agent_socs:
            print(f"Agent {args.agent} not found. Available: {sorted(agent_socs.keys())}")
            sys.exit(1)
        agent_socs = {args.agent: agent_socs[args.agent]}
        print(f"Single agent mode: Agent {args.agent}")

    # Run agents in parallel
    print(f"Launching {len(agent_socs)} agent(s)...\n")
    all_new_results = []

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(run_agent, aid, socs, soc_entries, completed): aid
            for aid, socs in agent_socs.items()
        }

        for future in as_completed(futures):
            agent_id = futures[future]
            try:
                results = future.result()
                all_new_results.extend(results)
                print(f"\n  Agent {agent_id} finished: {len(results)} SOCs scored")
            except Exception as e:
                print(f"\n  Agent {agent_id} ERROR: {e}")

    # Merge new results into completed
    for r in all_new_results:
        completed[r["soc_code"]] = r

    # Save checkpoint
    save_checkpoint(completed)

    # Summary
    scored_tasks = sum(len(e["scores"]) for e in completed.values())
    total_input = sum(e.get("usage", {}).get("input_tokens", 0) for e in completed.values())
    total_output = sum(e.get("usage", {}).get("output_tokens", 0) for e in completed.values())
    low_confidence = sum(
        1 for e in completed.values() for s in e["scores"]
        if s.get("confidence_mod") == "low" or s.get("confidence_sig") == "low"
    )

    print(f"\n{'='*60}")
    print(f"  SCORING COMPLETE")
    print(f"{'='*60}")
    print(f"  SOCs scored:       {len(completed)}/{len(soc_entries)}")
    print(f"  Tasks scored:      {scored_tasks}/{total_tasks}")
    print(f"  Low confidence:    {low_confidence} tasks flagged for audit")
    if total_input > 0:
        cost = total_input * 15 / 1_000_000 + total_output * 75 / 1_000_000
        print(f"  Tokens:            {total_input:,} in + {total_output:,} out")
        print(f"  Est. cost:         ${cost:.2f}")
    print(f"{'='*60}\n")

    # Failed SOCs
    failed = [soc for soc in soc_entries if soc not in completed]
    if failed:
        print(f"{len(failed)} SOCs not scored:")
        for soc in failed[:10]:
            print(f"  {soc}: {soc_entries[soc]['job_title']}")
        if len(failed) > 10:
            print(f"  ... and {len(failed) - 10} more")
        print("Re-run to retry failed SOCs.\n")

    # Writeback
    if len(completed) == len(soc_entries):
        print("All SOCs complete. Writing to workbook...")
        writeback(completed)
    else:
        print(f"{len(soc_entries) - len(completed)} SOCs remaining — "
              f"skipping writeback until all complete.")
        print("Use --writeback-only to force partial writeback.")


if __name__ == "__main__":
    main()
