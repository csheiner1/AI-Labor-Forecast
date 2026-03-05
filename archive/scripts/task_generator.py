"""
Multi-Agent Task Generator
Reads jobs from Tab "2 Jobs" of jobs-data.xlsx, generates O*NET-style tasks
using a 3-agent pipeline, and writes output to a Tasks tab.

Prerequisites:
  pip install anthropic openpyxl
  export ANTHROPIC_API_KEY=your_key

Usage:
  python task_generator.py --input jobs-data.xlsx --sheet "2 Jobs" --output jobs-data.xlsx
  python task_generator.py --input jobs-data.xlsx --sheet "2 Jobs" --limit 5  # test with 5 jobs
"""

import argparse
import json
import time
import openpyxl
from anthropic import Anthropic

client = Anthropic()
MODEL = "claude-sonnet-4-20250514"

# ---------------------------------------------------------------------------
# AGENT 1: TASK GENERATOR
# Produces raw 5-7 tasks per job in structured JSON
# ---------------------------------------------------------------------------
GENERATOR_SYSTEM = """You are an occupational analyst specializing in job task decomposition.
You produce task statements in O*NET format: verb-object-context, one sentence each.

Rules:
- Generate exactly 5-7 tasks per occupation
- Tasks should collectively cover ~90% of the role's core activities
- Use active verbs (analyze, develop, prepare, coordinate, evaluate, implement, manage)
- Each task must be distinct — no overlapping responsibilities
- Include a mix of strategic/analytical tasks and operational/communication tasks
- Rate importance honestly: not everything is a 5

Respond ONLY with valid JSON, no markdown fences, no preamble."""

def generate_tasks(job_title: str, soc_code: str, function_name: str, sector: str) -> list[dict]:
    """Agent 1: Generate raw tasks for a single job."""
    prompt = f"""Decompose this role into its 5-7 most important tasks:

Job Title: {job_title}
SOC Code: {soc_code}
Function: {function_name}
Primary Sector: {sector}

Return a JSON array where each element has:
{{
  "task": "O*NET-style task statement (verb-object-context, one sentence)",
  "importance": <float 1.0-5.0>,
  "relevance": <int 0-100, % of workers in this role who perform this task>,
  "frequency": "<daily|weekly|monthly|quarterly|annually>"
}}

Order by importance descending."""

    resp = client.messages.create(
        model=MODEL,
        max_tokens=1500,
        system=GENERATOR_SYSTEM,
        messages=[{"role": "user", "content": prompt}],
    )
    text = resp.content[0].text.strip()
    # Strip markdown fences if model includes them
    if text.startswith("```"):
        text = text.split("\n", 1)[1].rsplit("```", 1)[0]
    return json.loads(text)


# ---------------------------------------------------------------------------
# AGENT 2: QUALITY REVIEWER
# Validates format, checks for overlap, enforces O*NET style
# ---------------------------------------------------------------------------
REVIEWER_SYSTEM = """You are a senior occupational data quality reviewer.
You review AI-generated task statements against O*NET standards.

Your job:
1. Check each task follows verb-object-context format
2. Flag and fix any vague or overlapping tasks
3. Verify importance ratings are calibrated (not all 4.5+)
4. Ensure tasks collectively represent the role's core work
5. Assign each task a Task Type: "Core" if relevance >= 67 AND importance >= 3.0, else "Supplemental"
6. Assign the closest Generalized Work Activity (GWA) category from this list:
   - Analyzing Data/Information
   - Making Decisions and Solving Problems
   - Processing Information
   - Communicating with Supervisors, Peers, or Subordinates
   - Communicating with People Outside the Organization
   - Organizing, Planning, and Prioritizing Work
   - Updating and Using Relevant Knowledge
   - Establishing and Maintaining Interpersonal Relationships
   - Documenting/Recording Information
   - Evaluating Information to Determine Compliance
   - Developing Objectives and Strategies
   - Interacting With Computers
   - Getting Information
   - Monitoring Processes, Materials, or Surroundings
   - Coordinating the Work and Activities of Others
   - Training and Teaching Others
   - Thinking Creatively
   - Selling or Influencing Others
   - Performing Administrative Activities

Respond ONLY with valid JSON, no markdown fences, no preamble."""

def review_tasks(job_title: str, soc_code: str, raw_tasks: list[dict]) -> list[dict]:
    """Agent 2: Review and enrich tasks."""
    prompt = f"""Review these AI-generated tasks for: {job_title} ({soc_code})

Raw tasks:
{json.dumps(raw_tasks, indent=2)}

Return a JSON array with the reviewed tasks. Each element:
{{
  "task": "corrected task statement if needed",
  "importance": <float, adjusted if miscalibrated>,
  "relevance": <int>,
  "frequency": "<daily|weekly|monthly|quarterly|annually>",
  "task_type": "Core or Supplemental",
  "gwa": "closest GWA category from the list",
  "review_note": "what you changed, or 'approved' if no changes"
}}"""

    resp = client.messages.create(
        model=MODEL,
        max_tokens=2000,
        system=REVIEWER_SYSTEM,
        messages=[{"role": "user", "content": prompt}],
    )
    text = resp.content[0].text.strip()
    if text.startswith("```"):
        text = text.split("\n", 1)[1].rsplit("```", 1)[0]
    return json.loads(text)


# ---------------------------------------------------------------------------
# AGENT 3: FORMATTER
# Calculates weights and produces final structured rows
# ---------------------------------------------------------------------------
def format_tasks(soc_code: str, job_title: str, function_id: str,
                 function_name: str, reviewed_tasks: list[dict]) -> list[dict]:
    """Agent 3: Calculate weights, assign IDs, produce final rows."""
    total_importance = sum(t["importance"] for t in reviewed_tasks)

    rows = []
    for i, t in enumerate(reviewed_tasks, start=1):
        task_id = f"CL-{soc_code.replace('-', '')}-{i:03d}"
        weight = round(t["importance"] / total_importance, 4) if total_importance > 0 else 0

        rows.append({
            "SOC_Code": soc_code,
            "Job_Title": job_title,
            "Function_ID": function_id,
            "Function_Name": function_name,
            "Task_ID": task_id,
            "Task_Description": t["task"],
            "Task_Source": "Claude-generated",
            "Task_Type": t.get("task_type", ""),
            "Importance": t["importance"],
            "Relevance": t["relevance"],
            "Frequency": t["frequency"],
            "Task_Weight": weight,
            "GWA": t.get("gwa", ""),
            "Review_Note": t.get("review_note", ""),
            "Automatability_Score": "",  # blank for your scoring
            "Confidence": "Low",         # until validated
        })
    return rows


# ---------------------------------------------------------------------------
# PIPELINE ORCHESTRATOR
# ---------------------------------------------------------------------------
def process_job(job: dict) -> list[dict]:
    """Run the full 3-agent pipeline for one job."""
    title = job["title"]
    soc = job["soc_code"]
    func_id = job.get("function_id", "")
    func_name = job.get("function_name", "")
    sector = job.get("sector", "")

    print(f"  [Agent 1] Generating tasks for: {title} ({soc})")
    raw = generate_tasks(title, soc, func_name, sector)

    print(f"  [Agent 2] Reviewing {len(raw)} tasks...")
    reviewed = review_tasks(title, soc, raw)

    print(f"  [Agent 3] Formatting final output...")
    final = format_tasks(soc, title, func_id, func_name, reviewed)

    return final


def read_jobs_from_workbook(filepath: str, sheet_name: str) -> list[dict]:
    """Read job list from the specified sheet of the workbook.

    Expected columns in jobs-data.xlsx "2 Jobs" tab:
      SOC_Code | SOC_Title | Custom_Title | Delta_Sector_ID | Delta_Sector |
      Function_ID | Function_Name | ...
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]

    jobs = []
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Build column index map from actual header names
    col_map = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        hl = str(h).strip()
        # Map exact column names from our workbook
        if hl == "SOC_Code":
            col_map["soc_code"] = i
        elif hl == "Custom_Title":
            col_map["title"] = i
        elif hl == "SOC_Title":
            col_map["soc_title"] = i
        elif hl == "Function_ID":
            col_map["function_id"] = i
        elif hl == "Function_Name":
            col_map["function_name"] = i
        elif hl == "Delta_Sector":
            col_map["sector"] = i
        elif hl == "Delta_Sector_ID":
            col_map["sector_id"] = i

    # Fallback: if Custom_Title not found, try any "title" column
    if "title" not in col_map:
        for i, h in enumerate(headers):
            if h and "title" in str(h).lower():
                col_map["title"] = i
                break

    for row in ws.iter_rows(min_row=2, values_only=True):
        soc = row[col_map.get("soc_code", 0)]
        if soc is None:
            continue
        title = str(row[col_map.get("title", 2)] or "").strip()
        # Use Custom_Title if available, fall back to SOC_Title
        if not title and "soc_title" in col_map:
            title = str(row[col_map["soc_title"]] or "").strip()
        jobs.append({
            "soc_code": str(soc).strip(),
            "title": title,
            "function_id": str(row[col_map.get("function_id", 5)] or "").strip(),
            "function_name": str(row[col_map.get("function_name", 6)] or "").strip(),
            "sector": str(row[col_map.get("sector", 4)] or "").strip(),
        })

    wb.close()
    print(f"Loaded {len(jobs)} jobs from '{sheet_name}'")
    return jobs


def write_tasks_to_workbook(filepath: str, all_tasks: list[dict]):
    """Write tasks to a new 'Tasks' tab in the workbook."""
    wb = openpyxl.load_workbook(filepath)

    # Remove existing Tasks tab if present
    if "Tasks" in wb.sheetnames:
        del wb["Tasks"]

    ws = wb.create_sheet("Tasks")

    # Headers
    columns = [
        "SOC_Code", "Job_Title", "Function_ID", "Function_Name",
        "Task_ID", "Task_Description", "Task_Source", "Task_Type",
        "Importance", "Relevance", "Frequency", "Task_Weight",
        "GWA", "Review_Note", "Automatability_Score", "Confidence"
    ]

    # Format headers
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col_i, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True)

    # Freeze top row
    ws.freeze_panes = "A2"

    # Yellow fill for blank scoring columns
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Data rows
    for row_i, task in enumerate(all_tasks, start=2):
        for col_i, key in enumerate(columns, start=1):
            val = task.get(key, "")
            cell = ws.cell(row=row_i, column=col_i, value=val)
            # Yellow fill for Automatability_Score column
            if key == "Automatability_Score":
                cell.fill = yellow_fill

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    wb.save(filepath)
    print(f"\nWrote {len(all_tasks)} tasks to 'Tasks' tab in {filepath}")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Generate O*NET-style tasks via multi-agent pipeline")
    parser.add_argument("--input", default="jobs-data.xlsx", help="Path to workbook (.xlsx)")
    parser.add_argument("--sheet", default="2 Jobs", help="Sheet name with job list")
    parser.add_argument("--output", default=None, help="Output path (defaults to input file)")
    parser.add_argument("--limit", type=int, default=None, help="Process only first N jobs (for testing)")
    parser.add_argument("--delay", type=float, default=1.0, help="Seconds between API calls (rate limiting)")
    args = parser.parse_args()

    output = args.output or args.input

    # Read jobs
    jobs = read_jobs_from_workbook(args.input, args.sheet)
    if args.limit:
        jobs = jobs[:args.limit]
        print(f"Limited to first {args.limit} jobs for testing")

    # Process each job through the pipeline
    all_tasks = []
    for i, job in enumerate(jobs, start=1):
        print(f"\n[{i}/{len(jobs)}] Processing: {job['title']} ({job['soc_code']})")
        try:
            tasks = process_job(job)
            all_tasks.extend(tasks)
            print(f"  Generated {len(tasks)} tasks")
        except Exception as e:
            print(f"  ERROR: {e}")
            continue

        if args.delay and i < len(jobs):
            time.sleep(args.delay)

    # Write output
    write_tasks_to_workbook(output, all_tasks)

    # Summary
    print(f"\n{'='*60}")
    print(f"SUMMARY")
    print(f"{'='*60}")
    print(f"Jobs processed: {len(jobs)}")
    print(f"Total tasks generated: {len(all_tasks)}")
    if jobs:
        print(f"Avg tasks per job: {len(all_tasks)/len(jobs):.1f}")
    print(f"Output: {output} -> 'Tasks' tab")


if __name__ == "__main__":
    main()
