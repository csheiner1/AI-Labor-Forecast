"""
Agent 2: NIOEM Parser
Parses the BLS Employment Projections workbook (occupation.xlsx) and downloads
industry-occupation matrix data from the BLS API for Delta Sector industries.

Outputs:
  - occupations_master.csv  (from Table 1.2)
  - nioem_long.csv          (industry × occupation matrix in long format)
"""

import csv
import json
import os
import re
import sys
import time
import urllib.request
import urllib.error
from html.parser import HTMLParser

import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BLS_FILE = os.path.join(BASE_DIR, "occupation.xlsx")
LOOKUP_SECTORS = os.path.join(BASE_DIR, "lookup_sectors.csv")

# ─── Manual overrides for NAICS → BLS NEM code mapping ─────────────────────
# BLS uses its own NEM codes that don't always match standard NAICS.
# This handles 2022 NAICS revisions, composite codes, and government.
NAICS_TO_NEM_OVERRIDES = {
    # Finance: BLS composite code 5220A1 covers NAICS 5221 + 5223
    '52211': ('5220A1', 'Credit intermediation and related activities (5221,3)'),
    '52212': ('5220A1', 'Credit intermediation and related activities (5221,3)'),
    '52213': ('5220A1', 'Credit intermediation and related activities (5221,3)'),
    '5223':  ('5220A1', 'Credit intermediation and related activities (5221,3)'),
    # Securities & investments: BLS aggregate 523000
    '5231':  ('523000', 'Securities, commodity contracts, and other financial investments'),
    '5232':  ('523000', 'Securities, commodity contracts, and other financial investments'),
    '5239':  ('523000', 'Securities, commodity contracts, and other financial investments'),
    # Technology: NAICS 2022 revisions
    '5112':  ('513200', 'Software publishers'),
    '5182':  ('518000', 'Computing infrastructure providers, data processing, web hosting'),
    '51822': ('519000', 'Web search portals, libraries, archives, and other info services'),
    # Media: NAICS 2022 revisions
    '5111':  ('513110', 'Newspaper publishers'),
    # Marketing research (part of 5419 → check if 541900 exists)
    '54191': ('541900', 'Other professional, scientific, and technical services'),
    # Oil & Gas
    '2111':  ('211000', 'Oil and gas extraction'),
    # Architecture (BLS aggregates arch + engineering)
    '54131': ('541300', 'Architectural, engineering, and related services'),
    # Government: BLS uses special NEM codes
    '921':   ('999100', 'Federal government, excluding postal service'),
    '922':   ('999200', 'State government'),  # Justice is split across state/local
    '926':   ('999100', 'Federal government, excluding postal service'),
    '928':   ('999100', 'Federal government, excluding postal service'),
    # ── Manufacturing (Sector 15) ──────────────────────────────────────────
    # Excludes 3344 (Sector 2: Technology) and 3391 (Sector 3: Healthcare)
    '311':   ('311000', 'Food manufacturing'),
    '312':   ('312000', 'Beverage and tobacco product manufacturing'),
    '325':   ('325000', 'Chemical manufacturing'),
    '326':   ('326000', 'Plastics and rubber products manufacturing'),
    '331':   ('331000', 'Primary metal manufacturing'),
    '332':   ('332000', 'Fabricated metal product manufacturing'),
    '333':   ('333000', 'Machinery manufacturing'),
    '335':   ('335000', 'Electrical equipment, appliance, and component manufacturing'),
    '336':   ('336000', 'Transportation equipment manufacturing'),
    # ── Retail Trade (Sector 16) ───────────────────────────────────────────
    '441':   ('441000', 'Motor vehicle and parts dealers'),
    '445':   ('445000', 'Food and beverage retailers'),
    '452':   ('452000', 'General merchandise stores'),
    '455':   ('455000', 'Health and personal care retailers'),
    '456':   ('456000', 'Clothing, clothing accessories, shoe, and jewelry retailers'),
    # ── Construction (Sector 17) ───────────────────────────────────────────
    '236':   ('236000', 'Construction of buildings'),
    '237':   ('237000', 'Heavy and civil engineering construction'),
    '238':   ('238000', 'Specialty trade contractors'),
    # ── Transportation & Warehousing (Sector 18) ──────────────────────────
    '481':   ('481000', 'Air transportation'),
    '482':   ('482000', 'Rail transportation'),
    '484':   ('484000', 'Truck transportation'),
    '485':   ('485000', 'Transit and ground passenger transportation'),
    '488':   ('488000', 'Support activities for transportation'),
    '492':   ('492000', 'Couriers and messengers'),
    '493':   ('493000', 'Warehousing and storage'),
    # ── Wholesale Trade (Sector 19) ────────────────────────────────────────
    '423':   ('423000', 'Merchant wholesalers, durable goods'),
    '424':   ('424000', 'Merchant wholesalers, nondurable goods'),
    '4251':  ('425100', 'Wholesale trade agents and brokers'),
    # ── Accommodation & Food Services (Sector 20) ─────────────────────────
    '721':   ('721000', 'Accommodation'),
    '722':   ('722000', 'Food services and drinking places'),
}

# ─── Helper: clean BLS values ──────────────────────────────────────────────

def clean_value(val):
    """Strip BLS notation from a cell value."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    # Remove footnote markers like [1], (1), etc.
    s = re.sub(r'\[\d+\]', '', s)
    s = re.sub(r'\(\d+\)', '', s)
    s = s.strip()
    if s in ('—', '-', '--', '–', '*', '', 'N/A', 'n/a'):
        return None
    # Try to convert to float
    try:
        return float(s.replace(',', ''))
    except ValueError:
        return s


def clean_str(val):
    """Clean a string value."""
    if val is None:
        return ""
    s = str(val).strip()
    s = re.sub(r'\[\d+\]', '', s)
    return s.strip()


# ─── Part 1: Parse Table 1.2 (Occupations Master) ──────────────────────────

def parse_table_1_2():
    print("\n--- Parsing Table 1.2: Occupational Projections ---")
    wb = openpyxl.load_workbook(BLS_FILE, read_only=True, data_only=True)
    ws = wb['Table 1.2']

    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
        vals = list(row)
        title = clean_str(vals[0])
        soc_code = clean_str(vals[1])
        occ_type = clean_str(vals[2])

        if not soc_code or soc_code == '—':
            continue

        rows.append({
            'SOC_Code': soc_code,
            'SOC_Title': title.strip(),
            'Occupation_Type': occ_type,
            'Employment_2024': clean_value(vals[3]),
            'Employment_2034': clean_value(vals[4]),
            'Employment_Dist_Pct_2024': clean_value(vals[5]),
            'Employment_Dist_Pct_2034': clean_value(vals[6]),
            'Change_Numeric': clean_value(vals[7]),
            'Change_Percent': clean_value(vals[8]),
            'Percent_Self_Employed': clean_value(vals[9]),
            'Occupational_Openings': clean_value(vals[10]),
            'Median_Annual_Wage': clean_value(vals[11]),
            'Typical_Education': clean_str(vals[12]) if vals[12] else None,
            'Work_Experience': clean_str(vals[13]) if vals[13] else None,
            'OJT_Training': clean_str(vals[14]) if vals[14] else None,
        })

    wb.close()

    # Write occupations_master.csv
    out_path = os.path.join(BASE_DIR, "occupations_master.csv")
    fields = ['SOC_Code', 'SOC_Title', 'Occupation_Type', 'Employment_2024', 'Employment_2034',
              'Employment_Dist_Pct_2024', 'Employment_Dist_Pct_2034', 'Change_Numeric', 'Change_Percent',
              'Percent_Self_Employed', 'Occupational_Openings', 'Median_Annual_Wage',
              'Typical_Education', 'Work_Experience', 'OJT_Training']
    with open(out_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)

    total = len(rows)
    line_items = sum(1 for r in rows if r['Occupation_Type'] == 'Line item')
    summaries = sum(1 for r in rows if r['Occupation_Type'] == 'Summary')
    print(f"  Total occupations: {total}")
    print(f"  Line items: {line_items}, Summaries: {summaries}")
    print(f"  Written to: {out_path}")

    # Top 10 by employment
    sorted_rows = sorted([r for r in rows if r['Occupation_Type'] == 'Line item' and r['Employment_2024']],
                         key=lambda x: x['Employment_2024'], reverse=True)
    print("\n  Top 10 occupations by 2024 employment (thousands):")
    for r in sorted_rows[:10]:
        print(f"    {r['SOC_Code']}  {r['Employment_2024']:>10,.1f}  {r['SOC_Title'][:50]}")

    return rows


# ─── Part 2: Build NEM code → NAICS mapping from Table 1.9 ─────────────────

def parse_table_1_9():
    """Parse Table 1.9 to get industry NEM codes and NAICS mappings."""
    print("\n--- Parsing Table 1.9: Industry List ---")
    wb = openpyxl.load_workbook(BLS_FILE, read_only=True, data_only=True)
    ws = wb['Table 1.9']

    industries = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        vals = list(row)
        title = clean_str(vals[0])
        nem_code = clean_str(vals[1])
        ind_type = clean_str(vals[2])
        naics = clean_str(vals[3])

        if not nem_code:
            continue

        industries.append({
            'Title': title.strip(),
            'NEM_Code': nem_code,
            'Industry_Type': ind_type,
            'NAICS_Code': naics if naics != '—' else None,
        })

    wb.close()
    print(f"  Total industries: {len(industries)}")
    print(f"  With NAICS codes: {sum(1 for i in industries if i['NAICS_Code'])}")
    return industries


# ─── Part 3: Load Delta Sector lookup and find matching NEM codes ───────────

def load_sectors_and_find_nem_codes(industries):
    """Load lookup_sectors.csv and find matching NEM codes for each NAICS."""
    print("\n--- Matching Delta Sector NAICS to BLS NEM codes ---")

    sectors = []
    with open(LOOKUP_SECTORS, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            sectors.append(row)

    # Build NEM code → industry info lookup
    nem_lookup = {ind['NEM_Code']: ind for ind in industries}

    best_matches = {}
    unmatched_naics = []

    for sec in sectors:
        naics = sec['NAICS_Code']
        if naics in best_matches:
            continue  # Already matched this NAICS

        # Strategy 1: Check manual overrides first
        if naics in NAICS_TO_NEM_OVERRIDES:
            nem_code, nem_title = NAICS_TO_NEM_OVERRIDES[naics]
            # Verify this NEM code exists in BLS data (or trust the override)
            ind_info = nem_lookup.get(nem_code)
            best_matches[naics] = {
                'NAICS_Code': naics,
                'Delta_Sector_ID': sec['Delta_Sector_ID'],
                'Delta_Sector': sec['Delta_Sector'],
                'Delta_Sub_Industry': sec['Delta_Sub_Industry'],
                'NEM_Code': nem_code,
                'NEM_Title': ind_info['Title'].strip() if ind_info else nem_title,
                'Industry_Type': ind_info['Industry_Type'] if ind_info else 'Override',
                'BLS_NAICS': ind_info['NAICS_Code'] if ind_info else naics,
            }
            continue

        # Strategy 2: Automated matching against Table 1.9
        found = False
        candidates = []
        for ind in industries:
            nem = ind['NEM_Code']
            ind_naics = ind['NAICS_Code'] or ''

            # Exact match or padded match
            if (ind_naics == naics or ind_naics == naics + '0' or
                ind_naics == naics + '00' or ind_naics == naics + '000'):
                candidates.append({
                    'NAICS_Code': naics,
                    'Delta_Sector_ID': sec['Delta_Sector_ID'],
                    'Delta_Sector': sec['Delta_Sector'],
                    'Delta_Sub_Industry': sec['Delta_Sub_Industry'],
                    'NEM_Code': nem,
                    'NEM_Title': ind['Title'].strip(),
                    'Industry_Type': ind['Industry_Type'],
                    'BLS_NAICS': ind_naics,
                    'match_quality': 1,
                })
                found = True

            # Prefix match (our code is prefix of BLS code)
            elif ind_naics.startswith(naics) and len(ind_naics) - len(naics) <= 3:
                candidates.append({
                    'NAICS_Code': naics,
                    'Delta_Sector_ID': sec['Delta_Sector_ID'],
                    'Delta_Sector': sec['Delta_Sector'],
                    'Delta_Sub_Industry': sec['Delta_Sub_Industry'],
                    'NEM_Code': nem,
                    'NEM_Title': ind['Title'].strip(),
                    'Industry_Type': ind['Industry_Type'],
                    'BLS_NAICS': ind_naics,
                    'match_quality': 2,
                })
                found = True

            # Composite NAICS (like "5221,3")
            elif ',' in ind_naics:
                parts = ind_naics.split(',')
                base = parts[0][:len(naics)]
                if base == naics:
                    candidates.append({
                        'NAICS_Code': naics,
                        'Delta_Sector_ID': sec['Delta_Sector_ID'],
                        'Delta_Sector': sec['Delta_Sector'],
                        'Delta_Sub_Industry': sec['Delta_Sub_Industry'],
                        'NEM_Code': nem,
                        'NEM_Title': ind['Title'].strip(),
                        'Industry_Type': ind['Industry_Type'],
                        'BLS_NAICS': ind_naics,
                        'match_quality': 3,
                    })
                    found = True

        if candidates:
            # Pick best: prefer Line item, then best match quality, then closest NAICS length
            candidates.sort(key=lambda x: (
                x['match_quality'],
                0 if x['Industry_Type'] == 'Line item' else 1,
                abs(len(x['BLS_NAICS']) - len(naics))
            ))
            best_matches[naics] = candidates[0]
        else:
            unmatched_naics.append(sec)

    print(f"  Matched NAICS codes: {len(best_matches)} / {len(sectors)} rows ({len(set(m['NEM_Code'] for m in best_matches.values()))} unique NEM codes)")
    if unmatched_naics:
        print(f"  Unmatched NAICS codes ({len(unmatched_naics)}):")
        for u in unmatched_naics:
            print(f"    {u['NAICS_Code']}: {u['NAICS_Title']} ({u['Delta_Sector']})")

    print("\n  NEM codes to download:")
    # Show unique NEM codes (many NAICS map to same NEM)
    seen_nem = {}
    for naics, m in sorted(best_matches.items()):
        nem = m['NEM_Code']
        if nem not in seen_nem:
            seen_nem[nem] = []
        seen_nem[nem].append(naics)
    for nem, naics_list in sorted(seen_nem.items()):
        m = best_matches[naics_list[0]]
        print(f"    NEM {nem:8s}  ← NAICS {','.join(naics_list):20s}  {m['NEM_Title'][:50]}")

    return best_matches, sectors


# ─── Part 4: Download matrix data from BLS API ─────────────────────────────

class BLSMatrixParser(HTMLParser):
    """Parse the BLS projections HTML table to extract occupation × employment data."""

    def __init__(self):
        super().__init__()
        self.in_table = False
        self.in_thead = False
        self.in_tbody = False
        self.in_row = False
        self.in_cell = False
        self.in_header = False
        self.current_row = []
        self.current_cell = ""
        self.headers = []
        self.rows = []
        self.table_count = 0
        self.in_link = False

    def handle_starttag(self, tag, attrs):
        attrs_dict = dict(attrs)
        if tag == 'table':
            self.table_count += 1
            self.in_table = True
        elif tag == 'thead' and self.in_table:
            self.in_thead = True
        elif tag == 'tbody' and self.in_table:
            self.in_tbody = True
        elif tag == 'tr' and self.in_table:
            self.in_row = True
            self.current_row = []
        elif tag in ('td', 'th') and self.in_row:
            self.in_cell = True
            self.in_header = (tag == 'th')
            self.current_cell = ""
        elif tag == 'a' and self.in_cell:
            self.in_link = True

    def handle_endtag(self, tag):
        if tag == 'table':
            self.in_table = False
        elif tag == 'thead':
            self.in_thead = False
        elif tag == 'tbody':
            self.in_tbody = False
        elif tag == 'tr' and self.in_row:
            self.in_row = False
            if self.in_thead and self.current_row:
                self.headers = self.current_row
            elif self.in_tbody and self.current_row:
                self.rows.append(self.current_row)
            elif self.current_row and not self.headers:
                self.headers = self.current_row
            elif self.current_row:
                self.rows.append(self.current_row)
        elif tag in ('td', 'th') and self.in_cell:
            self.in_cell = False
            self.current_row.append(self.current_cell.strip())
        elif tag == 'a':
            self.in_link = False

    def handle_data(self, data):
        if self.in_cell:
            self.current_cell += data


def download_industry_matrix(nem_code, max_retries=3):
    """Download the occupation breakdown for a single BLS industry."""
    url = f"https://data.bls.gov/projections/nationalMatrix?queryParams={nem_code}&ioType=i"

    for attempt in range(max_retries):
        try:
            req = urllib.request.Request(url, headers={
                'User-Agent': 'Mozilla/5.0 (research-project)',
                'Accept': 'text/html',
            })
            with urllib.request.urlopen(req, timeout=30) as resp:
                html = resp.read().decode('utf-8', errors='replace')

            parser = BLSMatrixParser()
            parser.feed(html)

            if not parser.rows:
                # Try a simpler regex-based parsing for the data
                return parse_matrix_html_regex(html, nem_code)

            return parse_matrix_from_table(parser, nem_code)

        except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError) as e:
            if attempt < max_retries - 1:
                time.sleep(2 * (attempt + 1))
            else:
                print(f"    FAILED to download {nem_code}: {e}")
                return []


def parse_matrix_from_table(parser, nem_code):
    """Parse the HTML table parser results into structured data."""
    results = []
    for row in parser.rows:
        if len(row) < 4:
            continue

        # Typical columns: Title, Code, OccType, Employment2024, Employment2034, Change, ChangePct
        title = row[0].strip() if row[0] else ""
        code = row[1].strip() if len(row) > 1 else ""
        occ_type = row[2].strip() if len(row) > 2 else ""

        # SOC code validation
        if not re.match(r'\d{2}-\d{4}', code):
            continue

        emp_2024 = clean_value(row[3]) if len(row) > 3 else None
        emp_2034 = clean_value(row[4]) if len(row) > 4 else None

        if emp_2024 is not None and emp_2024 != 0:
            # Normalize occupation type
            if occ_type.lower().strip() == 'line item':
                occ_type = 'Line item'
            elif occ_type.lower().strip() == 'summary':
                occ_type = 'Summary'
            results.append({
                'SOC_Code': code,
                'SOC_Title': title,
                'Occupation_Type': occ_type,
                'NEM_Code': nem_code,
                'Employment_2024': emp_2024,
                'Employment_2034': emp_2034,
            })

    return results


def parse_matrix_html_regex(html, nem_code):
    """Fallback regex-based parser for BLS matrix pages."""
    results = []

    # Look for table rows with SOC codes
    # Pattern: <td>title</td><td>XX-XXXX</td><td>type</td><td>number</td><td>number</td>...
    row_pattern = re.compile(
        r'<tr[^>]*>.*?<td[^>]*>(.*?)</td>.*?<td[^>]*>([\d]{2}-[\d]{4})</td>.*?<td[^>]*>(.*?)</td>.*?<td[^>]*>(.*?)</td>.*?<td[^>]*>(.*?)</td>',
        re.DOTALL
    )

    for match in row_pattern.finditer(html):
        title = re.sub(r'<[^>]+>', '', match.group(1)).strip()
        soc_code = match.group(2).strip()
        occ_type = re.sub(r'<[^>]+>', '', match.group(3)).strip()
        emp_2024_str = re.sub(r'<[^>]+>', '', match.group(4)).strip()
        emp_2034_str = re.sub(r'<[^>]+>', '', match.group(5)).strip()

        emp_2024 = clean_value(emp_2024_str)
        emp_2034 = clean_value(emp_2034_str)

        if emp_2024 is not None and emp_2024 != 0:
            # Normalize occupation type
            if occ_type.lower().strip() == 'line item':
                occ_type = 'Line item'
            elif occ_type.lower().strip() == 'summary':
                occ_type = 'Summary'
            results.append({
                'SOC_Code': soc_code,
                'SOC_Title': title,
                'Occupation_Type': occ_type,
                'NEM_Code': nem_code,
                'Employment_2024': emp_2024,
                'Employment_2034': emp_2034,
            })

    return results


def verify_nem_codes(nem_codes):
    """Test each NEM code URL before full download. Returns set of valid codes."""
    print("\n--- Verifying NEM codes ---")
    valid = set()
    invalid = []
    for nem_code in sorted(nem_codes):
        url = f"https://data.bls.gov/projections/nationalMatrix?queryParams={nem_code}&ioType=i"
        try:
            req = urllib.request.Request(url, headers={
                'User-Agent': 'Mozilla/5.0 (research-project)',
                'Accept': 'text/html',
            })
            with urllib.request.urlopen(req, timeout=15) as resp:
                html = resp.read().decode('utf-8', errors='replace')
                # Check if page has actual data (SOC codes)
                if re.search(r'\d{2}-\d{4}', html):
                    valid.add(nem_code)
                    print(f"  {nem_code}: OK")
                else:
                    invalid.append(nem_code)
                    print(f"  {nem_code}: NO DATA (page exists but no occupations)")
        except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError) as e:
            invalid.append(nem_code)
            print(f"  {nem_code}: FAILED ({e})")
        time.sleep(0.3)

    print(f"  Valid: {len(valid)}/{len(nem_codes)}")
    if invalid:
        print(f"  Invalid NEM codes: {invalid}")
    return valid


def download_all_matrices(best_matches, sectors_list):
    """Download matrix data for all matched industries."""
    print("\n--- Downloading industry-occupation matrices from BLS ---")

    all_results = []
    nem_codes_done = set()

    # Build NEM → list of sector infos (multiple NAICS may map to one NEM)
    nem_to_sectors = {}
    for naics, m in best_matches.items():
        nem = m['NEM_Code']
        if nem not in nem_to_sectors:
            nem_to_sectors[nem] = m  # Use first one for metadata

    total = len(nem_to_sectors)
    for idx, (nem_code, info) in enumerate(sorted(nem_to_sectors.items()), 1):
        if nem_code in nem_codes_done:
            continue
        nem_codes_done.add(nem_code)

        print(f"  [{idx}/{total}] Downloading NEM {nem_code}: {info['NEM_Title'][:40]}...", end="", flush=True)
        rows = download_industry_matrix(nem_code)

        if rows:
            # Enrich with sector info from all NAICS that map to this NEM
            for r in rows:
                r['NAICS_Code'] = info['NAICS_Code']
                r['Delta_Sector_ID'] = info['Delta_Sector_ID']
                r['Delta_Sector'] = info['Delta_Sector']
                r['Delta_Sub_Industry'] = info['Delta_Sub_Industry']
            all_results.extend(rows)
            line_items = sum(1 for r in rows if r['Occupation_Type'] == 'Line item')
            print(f" {len(rows)} rows ({line_items} line items)")
        else:
            print(f" NO DATA")

        # Be nice to BLS servers
        time.sleep(0.5)

    return all_results


# ─── Main ───────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("AGENT 2: NIOEM Parser")
    print("=" * 60)

    # Part 1: Parse Table 1.2
    occ_data = parse_table_1_2()

    # Part 2: Parse Table 1.9 for industry list
    industries = parse_table_1_9()

    # Part 3: Match Delta Sectors to NEM codes
    best_matches, sectors_list = load_sectors_and_find_nem_codes(industries)

    # Part 3.5: Verify NEM codes before downloading
    unique_nem_codes = set(m['NEM_Code'] for m in best_matches.values())
    valid_nem = verify_nem_codes(unique_nem_codes)

    # Part 4: Download matrix data
    all_matrix = download_all_matrices(best_matches, sectors_list)

    # Write nioem_long.csv
    print("\n--- Writing nioem_long.csv ---")
    out_path = os.path.join(BASE_DIR, "nioem_long.csv")
    fields = ['SOC_Code', 'SOC_Title', 'Occupation_Type', 'NAICS_Code', 'Delta_Sector_ID',
              'Delta_Sector', 'Delta_Sub_Industry', 'NEM_Code', 'Employment_2024', 'Employment_2034']
    with open(out_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for r in all_matrix:
            writer.writerow({k: r.get(k) for k in fields})

    print(f"  Total rows: {len(all_matrix)}")
    unique_occ = set(r['SOC_Code'] for r in all_matrix)
    unique_ind = set(r['NEM_Code'] for r in all_matrix)
    print(f"  Unique occupations: {len(unique_occ)}")
    print(f"  Unique industries: {len(unique_ind)}")

    # Total employment - check occupation types
    occ_types = {}
    for r in all_matrix:
        ot = r.get('Occupation_Type', 'Unknown')
        occ_types[ot] = occ_types.get(ot, 0) + 1
    print(f"  Occupation type breakdown: {occ_types}")

    # Employment by line items AND all (in case type detection is different)
    total_emp_line = sum(r['Employment_2024'] for r in all_matrix
                         if r['Employment_2024'] and r['Occupation_Type'] == 'Line item')
    total_emp_all = sum(r['Employment_2024'] for r in all_matrix
                        if r['Employment_2024'] and re.match(r'\d{2}-\d{4}$', r['SOC_Code']))
    total_emp_detail = sum(r['Employment_2024'] for r in all_matrix
                           if r['Employment_2024'] and not r['SOC_Code'].endswith('0000') and not r['SOC_Code'].endswith('000'))
    print(f"  Total employment (Line items, thousands): {total_emp_line:,.1f}")
    print(f"  Total employment (detailed SOC XX-XXXX, thousands): {total_emp_all:,.1f}")
    print(f"  Total employment (non-summary SOC, thousands): {total_emp_detail:,.1f}")

    # Sample rows
    print("\n  Sample rows from nioem_long.csv:")
    for r in all_matrix[:5]:
        print(f"    {r['SOC_Code']} | {r['SOC_Title'][:30]:30s} | {r['Delta_Sector'][:25]:25s} | {r.get('Employment_2024', 'N/A')}")

    # Diagnostics
    print("\n" + "=" * 60)
    print("DIAGNOSTIC SUMMARY")
    print("=" * 60)
    total_occ = len(occ_data)
    line_items_occ = sum(1 for r in occ_data if r['Occupation_Type'] == 'Line item')
    print(f"Occupations Master: {total_occ} total, {line_items_occ} line items")
    print(f"NIOEM Long: {len(all_matrix)} rows, {len(unique_occ)} occupations × {len(unique_ind)} industries")
    print(f"Non-zero cells: {sum(1 for r in all_matrix if r['Employment_2024'] and r['Employment_2024'] > 0)}")

    print("\nAgent 2 complete.")


if __name__ == "__main__":
    main()
