"""
phase_removals.py – Phases 2, 3, 6, 7
Removes sectors, trims jobs, and cleans up all related sheets in the workbook.
"""

from openpyxl import Workbook


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _delete_rows(ws, row_indices):
    """Delete rows by 1-based index, bottom-to-top to avoid shifting."""
    for r in sorted(row_indices, reverse=True):
        ws.delete_rows(r)


def _delete_cols(ws, col_indices):
    """Delete columns by 1-based index, right-to-left to avoid shifting."""
    for c in sorted(col_indices, reverse=True):
        ws.delete_cols(c)


def _find_col_index(ws, header_row, text):
    """Return the 1-based column index whose header contains *text*."""
    for c in range(1, ws.max_column + 1):
        val = ws.cell(header_row, c).value
        if val is not None and text in str(val):
            return c
    return None


def _recalc_matrix_row_totals(ws):
    """Recalculate the last column ('Row Total') for both sections of Matrix."""
    max_col = ws.max_column
    for r in range(3, ws.max_row + 1):
        fid = ws.cell(r, 1).value
        if fid is None:
            # Could be a blank row or label row – check if there's numeric data
            if ws.cell(r, 2).value == 'COLUMN TOTAL':
                total = 0
                for c in range(3, max_col):
                    v = ws.cell(r, c).value
                    if isinstance(v, (int, float)):
                        total += v
                ws.cell(r, max_col).value = round(total, 1)
            continue
        if str(fid).strip() in ('Function ID', 'NORMALIZED (% of row total)'):
            continue
        # Data row – sum cols 3..(max_col-1) -> max_col
        total = 0
        for c in range(3, max_col):
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)):
                total += v
        ws.cell(r, max_col).value = round(total, 2) if isinstance(total, float) else total


# ---------------------------------------------------------------------------
# Phase 2 – Remove Sectors 8 (Staffing) and 20 (Accommodation)
# ---------------------------------------------------------------------------

def _phase2(wb):
    print("=== Phase 2: Remove Sectors 8 (Staffing) & 20 (Accommodation) ===")

    # ---- Collect job titles from 2 Jobs BEFORE deletion ----
    ws_jobs = wb['2 Jobs']
    staffing_titles = set()
    accommodation_titles = set()
    for r in range(2, ws_jobs.max_row + 1):
        sid = ws_jobs.cell(r, 4).value  # Sector_ID (string)
        title = ws_jobs.cell(r, 3).value  # Custom_Title
        if sid is not None and str(sid).strip() == '8' and title:
            staffing_titles.add(title)
        elif sid is not None and str(sid).strip() == '20' and title:
            accommodation_titles.add(title)
    all_removed_titles = staffing_titles | accommodation_titles
    print(f"  Collected {len(staffing_titles)} staffing job titles, {len(accommodation_titles)} accommodation job titles")

    # ---- 1A Industries (Sector_ID is INTEGER) ----
    ws = wb['1A Industries']
    rows_to_del = []
    for r in range(2, ws.max_row + 1):
        sid = ws.cell(r, 3).value
        if sid in (8, 20):
            rows_to_del.append(r)
    _delete_rows(ws, rows_to_del)
    print(f"  1A Industries: deleted {len(rows_to_del)} rows")

    # ---- 1A Summary (Sector_ID is INTEGER) ----
    ws = wb['1A Summary']
    rows_to_del = []
    for r in range(2, ws.max_row + 1):
        sid = ws.cell(r, 1).value
        if sid in (8, 20):
            rows_to_del.append(r)
    _delete_rows(ws, rows_to_del)
    print(f"  1A Summary: deleted {len(rows_to_del)} rows")

    # ---- Matrix – delete columns for sectors 8 and 20 ----
    ws = wb['Matrix']
    col8 = _find_col_index(ws, 2, '8 Staffing')
    col20 = _find_col_index(ws, 2, '20 Accommodation')
    cols_to_del = [c for c in (col8, col20) if c is not None]
    # Also check normalized header row
    col8n = _find_col_index(ws, 23, '8 Staffing')
    col20n = _find_col_index(ws, 23, '20 Accommodation')
    # Columns should be same index for both sections; delete once removes both
    _delete_cols(ws, cols_to_del)
    _recalc_matrix_row_totals(ws)
    print(f"  Matrix: deleted {len(cols_to_del)} columns")

    # ---- 2B Job_Industry – delete columns for Staffing and Accommodation ----
    ws = wb['2B Job_Industry']
    col8 = _find_col_index(ws, 2, 'Staffing')
    col20 = _find_col_index(ws, 2, 'Accommodation')
    cols_to_del = [c for c in (col8, col20) if c is not None]
    _delete_cols(ws, cols_to_del)
    print(f"  2B Job_Industry: deleted {len(cols_to_del)} columns")

    # ---- Staffing Patterns (Sector_ID is STRING) ----
    ws = wb['Staffing Patterns']
    rows_to_del = []
    for r in range(2, ws.max_row + 1):
        sid = ws.cell(r, 1).value
        if sid is not None and str(sid).strip() in ('8', '20'):
            rows_to_del.append(r)
    _delete_rows(ws, rows_to_del)
    print(f"  Staffing Patterns: deleted {len(rows_to_del)} rows")

    # ---- Lookup_Sectors (Sector_ID is STRING) ----
    ws = wb['Lookup_Sectors']
    rows_to_del = []
    for r in range(2, ws.max_row + 1):
        sid = ws.cell(r, 1).value
        if sid is not None and str(sid).strip() in ('8', '20'):
            rows_to_del.append(r)
    _delete_rows(ws, rows_to_del)
    print(f"  Lookup_Sectors: deleted {len(rows_to_del)} rows")

    # ---- Lookup_Jobs (Sector_ID is STRING) ----
    ws = wb['Lookup_Jobs']
    rows_to_del = []
    for r in range(2, ws.max_row + 1):
        sid = ws.cell(r, 2).value  # col B = Sector_ID
        if sid is not None and str(sid).strip() in ('8', '20'):
            rows_to_del.append(r)
    _delete_rows(ws, rows_to_del)
    print(f"  Lookup_Jobs: deleted {len(rows_to_del)} rows")

    # ---- 2 Jobs (Sector_ID is STRING) ----
    ws_jobs = wb['2 Jobs']
    rows_to_del = []
    for r in range(2, ws_jobs.max_row + 1):
        sid = ws_jobs.cell(r, 4).value
        if sid is not None and str(sid).strip() in ('8', '20'):
            rows_to_del.append(r)
    _delete_rows(ws_jobs, rows_to_del)
    print(f"  2 Jobs: deleted {len(rows_to_del)} rows")

    # ---- 3 Tasks – delete by matching Job_Title (col B) ----
    ws = wb['3 Tasks']
    rows_to_del = []
    for r in range(2, ws.max_row + 1):
        title = ws.cell(r, 2).value  # Job_Title
        if title is not None and title in all_removed_titles:
            rows_to_del.append(r)
    _delete_rows(ws, rows_to_del)
    print(f"  3 Tasks: deleted {len(rows_to_del)} rows")

    # ---- Industry Frictions ----
    ws = wb['Industry Frictions']
    rows_to_del = []
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 2).value
        if val is not None:
            v = str(val).strip()
            if v == 'Staffing & Recruitment Agencies' or 'Accommodation' in v:
                rows_to_del.append(r)
    _delete_rows(ws, rows_to_del)
    print(f"  Industry Frictions: deleted {len(rows_to_del)} rows")


# ---------------------------------------------------------------------------
# Phase 3 – Remove Construction (17), migrate 2 jobs to Sector 14
# ---------------------------------------------------------------------------

def _phase3(wb):
    print("\n=== Phase 3: Remove Construction (17), migrate 2 jobs ===")

    # Step 3a – Migrate Construction Manager and Construction Estimator to Sector 14
    migrate_socs = {'11-9021', '13-1051'}
    migrate_titles = set()

    # 2 Jobs
    ws = wb['2 Jobs']
    migrated = 0
    for r in range(2, ws.max_row + 1):
        soc = ws.cell(r, 1).value
        sid = ws.cell(r, 4).value
        if soc in migrate_socs and sid is not None and str(sid).strip() == '17':
            ws.cell(r, 4).value = '14'  # Sector_ID
            ws.cell(r, 5).value = 'Architecture & Engineering Firms'  # Sector
            ws.cell(r, 6).value = '17'  # Function_ID
            ws.cell(r, 7).value = 'Engineering & Architecture'  # Function_Name
            title = ws.cell(r, 3).value
            if title:
                migrate_titles.add(title)
            migrated += 1
    print(f"  Step 3a: migrated {migrated} jobs in 2 Jobs")

    # Lookup_Jobs – set Sector_ID='14'
    ws = wb['Lookup_Jobs']
    migrated = 0
    for r in range(2, ws.max_row + 1):
        soc = ws.cell(r, 3).value  # SOC_Code col C
        sid = ws.cell(r, 2).value  # Sector_ID col B
        if soc in migrate_socs and sid is not None and str(sid).strip() == '17':
            ws.cell(r, 2).value = '14'
            migrated += 1
    print(f"  Step 3a: migrated {migrated} jobs in Lookup_Jobs")

    # 3 Tasks – update Function_ID and Function_Name for migrated titles
    ws = wb['3 Tasks']
    updated = 0
    for r in range(2, ws.max_row + 1):
        title = ws.cell(r, 2).value
        if title in migrate_titles:
            ws.cell(r, 3).value = '17'  # Function_ID
            ws.cell(r, 4).value = 'Engineering & Architecture'  # Function_Name
            updated += 1
    print(f"  Step 3a: updated {updated} tasks in 3 Tasks")

    # Step 3b – Collect remaining Construction job titles, then delete
    ws_jobs = wb['2 Jobs']
    construction_titles = set()
    construction_rows = []
    for r in range(2, ws_jobs.max_row + 1):
        sid = ws_jobs.cell(r, 4).value
        if sid is not None and str(sid).strip() == '17':
            title = ws_jobs.cell(r, 3).value
            if title:
                construction_titles.add(title)
            construction_rows.append(r)
    _delete_rows(ws_jobs, construction_rows)
    print(f"  Step 3b: deleted {len(construction_rows)} remaining jobs from 2 Jobs ({construction_titles})")

    # 3 Tasks – delete tasks for remaining construction jobs
    ws = wb['3 Tasks']
    rows_to_del = []
    for r in range(2, ws.max_row + 1):
        title = ws.cell(r, 2).value
        if title is not None and title in construction_titles:
            rows_to_del.append(r)
    _delete_rows(ws, rows_to_del)
    print(f"  Step 3b: deleted {len(rows_to_del)} tasks from 3 Tasks")

    # 1A Industries (INTEGER)
    ws = wb['1A Industries']
    rows_to_del = [r for r in range(2, ws.max_row + 1) if ws.cell(r, 3).value == 17]
    _delete_rows(ws, rows_to_del)
    print(f"  1A Industries: deleted {len(rows_to_del)} rows")

    # 1A Summary (INTEGER)
    ws = wb['1A Summary']
    rows_to_del = [r for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value == 17]
    _delete_rows(ws, rows_to_del)
    print(f"  1A Summary: deleted {len(rows_to_del)} rows")

    # Staffing Patterns (STRING)
    ws = wb['Staffing Patterns']
    rows_to_del = [r for r in range(2, ws.max_row + 1)
                   if ws.cell(r, 1).value is not None and str(ws.cell(r, 1).value).strip() == '17']
    _delete_rows(ws, rows_to_del)
    print(f"  Staffing Patterns: deleted {len(rows_to_del)} rows")

    # Lookup_Sectors (STRING)
    ws = wb['Lookup_Sectors']
    rows_to_del = [r for r in range(2, ws.max_row + 1)
                   if ws.cell(r, 1).value is not None and str(ws.cell(r, 1).value).strip() == '17']
    _delete_rows(ws, rows_to_del)
    print(f"  Lookup_Sectors: deleted {len(rows_to_del)} rows")

    # Lookup_Jobs (STRING, col B)
    ws = wb['Lookup_Jobs']
    rows_to_del = [r for r in range(2, ws.max_row + 1)
                   if ws.cell(r, 2).value is not None and str(ws.cell(r, 2).value).strip() == '17']
    _delete_rows(ws, rows_to_del)
    print(f"  Lookup_Jobs: deleted {len(rows_to_del)} rows")

    # Matrix – delete Construction column
    ws = wb['Matrix']
    col17 = _find_col_index(ws, 2, '17 Construction')
    if col17 is None:
        col17 = _find_col_index(ws, 2, 'Construction')
    if col17:
        _delete_cols(ws, [col17])
        _recalc_matrix_row_totals(ws)
        print(f"  Matrix: deleted column {col17}")

    # 2B Job_Industry – delete Construction column
    ws = wb['2B Job_Industry']
    col17 = _find_col_index(ws, 2, 'Construction')
    if col17:
        _delete_cols(ws, [col17])
        print(f"  2B Job_Industry: deleted column {col17}")

    # Industry Frictions
    ws = wb['Industry Frictions']
    rows_to_del = []
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 2).value
        if val is not None and 'Construction' in str(val):
            rows_to_del.append(r)
    _delete_rows(ws, rows_to_del)
    print(f"  Industry Frictions: deleted {len(rows_to_del)} rows")


# ---------------------------------------------------------------------------
# Phase 6 – Trim Manufacturing (15): remove 2 blue-collar jobs
# ---------------------------------------------------------------------------

def _phase6(wb):
    print("\n=== Phase 6: Trim Manufacturing (15) – remove 2 jobs ===")

    trim_jobs = {
        ('51-1011', '15'),  # Production Supervisor
        ('43-5071', '15'),  # Shipping, receiving, and inventory clerks
    }

    # Collect titles first
    ws = wb['2 Jobs']
    titles_to_remove = set()
    rows_to_del = []
    for r in range(2, ws.max_row + 1):
        soc = ws.cell(r, 1).value
        sid = ws.cell(r, 4).value
        if soc is not None and sid is not None and (str(soc).strip(), str(sid).strip()) in trim_jobs:
            title = ws.cell(r, 3).value
            if title:
                titles_to_remove.add(title)
            rows_to_del.append(r)
    _delete_rows(ws, rows_to_del)
    print(f"  2 Jobs: deleted {len(rows_to_del)} rows – {titles_to_remove}")

    # Lookup_Jobs
    ws = wb['Lookup_Jobs']
    rows_to_del = []
    for r in range(2, ws.max_row + 1):
        soc = ws.cell(r, 3).value
        sid = ws.cell(r, 2).value
        if soc is not None and sid is not None and (str(soc).strip(), str(sid).strip()) in trim_jobs:
            rows_to_del.append(r)
    _delete_rows(ws, rows_to_del)
    print(f"  Lookup_Jobs: deleted {len(rows_to_del)} rows")

    # 3 Tasks
    ws = wb['3 Tasks']
    rows_to_del = []
    for r in range(2, ws.max_row + 1):
        title = ws.cell(r, 2).value
        if title is not None and title in titles_to_remove:
            rows_to_del.append(r)
    _delete_rows(ws, rows_to_del)
    print(f"  3 Tasks: deleted {len(rows_to_del)} rows")


# ---------------------------------------------------------------------------
# Phase 7 – Trim Retail (16): remove 7 jobs
# ---------------------------------------------------------------------------

def _phase7(wb):
    print("\n=== Phase 7: Trim Retail (16) – remove 7 jobs ===")

    trim_socs = {
        '33-9099',  # Loss Prevention Manager
        '27-1023',  # Floral designers
        '29-1051',  # Pharmacists
        '29-2052',  # Pharmacy technicians
        '29-2092',  # Hearing aid specialists
        '41-1011',  # First-line supervisors of retail sales workers
        '41-9099',  # Sales and related workers, all other
    }

    # Collect titles first
    ws = wb['2 Jobs']
    titles_to_remove = set()
    rows_to_del = []
    for r in range(2, ws.max_row + 1):
        soc = ws.cell(r, 1).value
        sid = ws.cell(r, 4).value
        if soc is not None and sid is not None and str(soc).strip() in trim_socs and str(sid).strip() == '16':
            title = ws.cell(r, 3).value
            if title:
                titles_to_remove.add(title)
            rows_to_del.append(r)
    _delete_rows(ws, rows_to_del)
    print(f"  2 Jobs: deleted {len(rows_to_del)} rows – {titles_to_remove}")

    # Lookup_Jobs
    ws = wb['Lookup_Jobs']
    rows_to_del = []
    for r in range(2, ws.max_row + 1):
        soc = ws.cell(r, 3).value
        sid = ws.cell(r, 2).value
        if soc is not None and sid is not None and str(soc).strip() in trim_socs and str(sid).strip() == '16':
            rows_to_del.append(r)
    _delete_rows(ws, rows_to_del)
    print(f"  Lookup_Jobs: deleted {len(rows_to_del)} rows")

    # 3 Tasks
    ws = wb['3 Tasks']
    rows_to_del = []
    for r in range(2, ws.max_row + 1):
        title = ws.cell(r, 2).value
        if title is not None and title in titles_to_remove:
            rows_to_del.append(r)
    _delete_rows(ws, rows_to_del)
    print(f"  3 Tasks: deleted {len(rows_to_del)} rows")


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def apply_removals(wb):
    """Apply Phases 2, 3, 6, 7 removals to an openpyxl Workbook in place."""
    _phase2(wb)
    _phase3(wb)
    _phase6(wb)
    _phase7(wb)
    print("\n=== All removal phases complete ===")
