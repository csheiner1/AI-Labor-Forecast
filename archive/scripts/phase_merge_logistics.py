"""
Phase 5 — Merge Transportation & Warehousing (18) + Wholesale Trade (19)
         into "Logistics & Distribution" (keep Sector_ID 18, remove 19).
"""

from openpyxl import Workbook

NEW_NAME = "Logistics & Distribution"

# Jobs to KEEP (move wholesale ones to sector 18)
KEEP_CUSTOM_TITLES_SECTOR18 = {
    "Logistics Coordinator",
    "Logistics Manager",
    "Warehouse Manager",
}
KEEP_CUSTOM_TITLES_FROM_WHOLESALE = {
    "Sales representatives, wholesale and manufacturing, technical and scientific products",
    "First-line supervisors of non-retail sales workers",
}

# Jobs to REMOVE from Transport (sector 18)
REMOVE_FROM_TRANSPORT = {
    "Fleet Manager",
    "Dispatch Coordinator",
    "Occupational health and safety technicians",
    "Reservation and transportation ticket agents and travel clerks",
    "Cargo and freight agents",
    "Weighers, measurers, checkers, and samplers, recordkeeping",
}

# Jobs to REMOVE from Wholesale (sector 19)
REMOVE_FROM_WHOLESALE = {
    "Distribution Manager",
    "Wholesale Account Manager",
    "Farmers, ranchers, and other agricultural managers",
    "Fashion designers",
    "Merchandise displayers and window trimmers",
    "Designers, all other",
    "Order clerks",
}


def _safe(v):
    """Return numeric value, treating None as 0."""
    return v if v is not None else 0


def apply_logistics_merge(wb: Workbook):
    """Merge Transport & Warehousing + Wholesale Trade into Logistics & Distribution."""

    # ── Step 5a: Rename Sector 18 everywhere ──────────────────────────
    print("  5a  Renaming sector 18 → Logistics & Distribution …")

    # Sheets where Sector_ID is INTEGER (column index 2 = col C for 1A Industries, col A for 1A Summary)
    ws_ind = wb["1A Industries"]
    for r in range(2, ws_ind.max_row + 1):
        if ws_ind.cell(r, 3).value == 18:  # Sector_ID col C
            ws_ind.cell(r, 4).value = NEW_NAME  # Sector col D

    ws_sum = wb["1A Summary"]
    for r in range(2, ws_sum.max_row + 1):
        if ws_sum.cell(r, 1).value == 18:  # Sector_ID col A
            ws_sum.cell(r, 2).value = NEW_NAME  # Sector col B

    # Sheets where Sector_ID is STRING
    str_sheets_sector = {
        "2 Jobs": (4, 5),        # Sector_ID col D(4), Sector col E(5)
        "Staffing Patterns": (1, 2),  # Sector_ID col A(1), Sector col B(2)
        "Lookup_Sectors": (1, 2),     # Sector_ID col A(1), Sector col B(2)
        "Lookup_Jobs": (2, None),     # Sector_ID col B(2), no Sector column
    }
    for sheet_name, (sid_col, sec_col) in str_sheets_sector.items():
        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, sid_col).value) == "18":
                if sec_col is not None:
                    ws.cell(r, sec_col).value = NEW_NAME

    # ── Step 5b: Merge NAICS/lookup data ──────────────────────────────
    print("  5b  Merging NAICS / lookup / summary data …")

    # 1A Industries: change sector 19 rows → sector 18
    for r in range(2, ws_ind.max_row + 1):
        if ws_ind.cell(r, 3).value == 19:
            ws_ind.cell(r, 3).value = 18
            ws_ind.cell(r, 4).value = NEW_NAME

    # Lookup_Sectors: change '19' → '18'
    ws_lu = wb["Lookup_Sectors"]
    for r in range(2, ws_lu.max_row + 1):
        if str(ws_lu.cell(r, 1).value) == "19":
            ws_lu.cell(r, 1).value = "18"
            ws_lu.cell(r, 2).value = NEW_NAME

    # 1A Summary: merge row 19 into row 18, delete row 19
    transport_row = wholesale_row = None
    for r in range(2, ws_sum.max_row + 1):
        sid = ws_sum.cell(r, 1).value
        if sid == 18:
            transport_row = r
        elif sid == 19:
            wholesale_row = r

    if transport_row and wholesale_row:
        emp_t = _safe(ws_sum.cell(transport_row, 4).value)  # 6456.6
        emp_w = _safe(ws_sum.cell(wholesale_row, 4).value)  # 5625.2
        naics_t = _safe(ws_sum.cell(transport_row, 3).value)
        naics_w = _safe(ws_sum.cell(wholesale_row, 3).value)
        wage_t = _safe(ws_sum.cell(transport_row, 5).value)
        wage_w = _safe(ws_sum.cell(wholesale_row, 5).value)

        combined_emp = emp_t + emp_w  # 12081.8
        combined_naics = naics_t + naics_w  # 10
        combined_wage = (emp_t * wage_t + emp_w * wage_w) / combined_emp if combined_emp else 0

        ws_sum.cell(transport_row, 3).value = combined_naics
        ws_sum.cell(transport_row, 4).value = round(combined_emp, 1)
        ws_sum.cell(transport_row, 5).value = combined_wage

        ws_sum.delete_rows(wholesale_row)
        print(f"       Summary: emp={combined_emp:.1f}K, NAICS={combined_naics}, wage=${combined_wage:,.0f}")

    # ── Step 5c: Merge Matrix columns ─────────────────────────────────
    print("  5c  Merging Matrix columns …")
    _merge_matrix(wb["Matrix"])

    # ── Step 5d: Merge 2B Job_Industry columns ────────────────────────
    print("  5d  Merging 2B Job_Industry columns …")
    _merge_2b(wb["2B Job_Industry"])

    # ── Step 5e: Merge Staffing Patterns ──────────────────────────────
    print("  5e  Merging Staffing Patterns …")
    _merge_staffing(wb["Staffing Patterns"])

    # ── Step 5f: Clean up jobs ────────────────────────────────────────
    print("  5f  Cleaning up jobs …")
    _clean_jobs(wb)

    # ── Step 5g: Merge Industry Frictions ─────────────────────────────
    print("  5g  Merging Industry Frictions …")
    _merge_frictions(wb["Industry Frictions"])

    print("  ✓  Phase 5 (Logistics merge) complete.")


# ──────────────────────────────────────────────────────────────────────
# 5c  Matrix merge
# ──────────────────────────────────────────────────────────────────────
def _merge_matrix(ws):
    """Merge Transport + Wholesale columns in Matrix (employment + normalized)."""
    # Find column indices from header row 2
    transport_col = wholesale_col = row_total_col = None
    for c in range(1, ws.max_column + 1):
        hdr = str(ws.cell(2, c).value or "")
        if "18 Transportation" in hdr or "18 Logistics" in hdr:
            transport_col = c
        elif "19 Wholesale" in hdr:
            wholesale_col = c
        elif hdr == "Row Total":
            row_total_col = c

    if not (transport_col and wholesale_col):
        print("       WARNING: could not find Transport/Wholesale columns in Matrix")
        return

    # Employment section: rows 3-20 (data rows 3-19, total row 20)
    _merge_col_section(ws, transport_col, wholesale_col, row_total_col,
                       data_start=3, data_end=20)

    # Rename header in employment section
    ws.cell(2, transport_col).value = f"18 {NEW_NAME}"

    # Normalized section: find headers again (row 23), data rows 24-40
    norm_row_total_col = None
    for c in range(1, ws.max_column + 1):
        hdr = str(ws.cell(23, c).value or "")
        if hdr == "Row Total":
            norm_row_total_col = c
            break

    # The transport/wholesale columns are at the same position in the normalized section
    _merge_col_section(ws, transport_col, wholesale_col, norm_row_total_col,
                       data_start=24, data_end=40)

    # Rename header in normalized section
    ws.cell(23, transport_col).value = f"18 {NEW_NAME}"

    # Delete the wholesale column (applies to both sections at once)
    ws.delete_cols(wholesale_col)
    print(f"       Deleted wholesale column {wholesale_col}")


def _merge_col_section(ws, t_col, w_col, total_col, data_start, data_end):
    """Add wholesale values into transport column for a range of rows, recalc row totals."""
    for r in range(data_start, data_end + 1):
        tv = _safe(ws.cell(r, t_col).value)
        wv = _safe(ws.cell(r, w_col).value)
        ws.cell(r, t_col).value = round(tv + wv, 10)

    # Recalculate Row Total for each row (sum of all industry columns)
    if total_col:
        # After merging, industry data cols are 3 .. total_col-1
        # (wholesale col still exists at this point, but its value is stale — however
        # we've already added it into transport, so skip wholesale)
        for r in range(data_start, data_end + 1):
            row_sum = 0
            for c in range(3, total_col):
                if c == w_col:
                    continue  # skip wholesale (already folded in)
                row_sum += _safe(ws.cell(r, c).value)
            ws.cell(r, total_col).value = round(row_sum, 10)


# ──────────────────────────────────────────────────────────────────────
# 5d  2B Job_Industry merge
# ──────────────────────────────────────────────────────────────────────
def _merge_2b(ws):
    """Merge Transport + Wholesale columns in 2B Job_Industry."""
    transport_col = wholesale_col = None
    for c in range(1, ws.max_column + 1):
        hdr = str(ws.cell(2, c).value or "")
        if "Transportation" in hdr:
            transport_col = c
        elif "Wholesale" in hdr:
            wholesale_col = c

    if not (transport_col and wholesale_col):
        print("       WARNING: could not find Transport/Wholesale columns in 2B")
        return

    # Merge values for all data rows (row 3 onwards)
    for r in range(3, ws.max_row + 1):
        tv = _safe(ws.cell(r, transport_col).value)
        wv = _safe(ws.cell(r, wholesale_col).value)
        ws.cell(r, transport_col).value = round(tv + wv, 10)

    # Rename header
    ws.cell(2, transport_col).value = NEW_NAME

    # Delete wholesale column
    ws.delete_cols(wholesale_col)
    print(f"       Deleted wholesale column {wholesale_col}")


# ──────────────────────────────────────────────────────────────────────
# 5e  Staffing Patterns merge
# ──────────────────────────────────────────────────────────────────────
def _merge_staffing(ws):
    """Change sector 19 → 18, then combine duplicate SOC entries."""
    # Reassign sector 19 → 18
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 1).value) == "19":
            ws.cell(r, 1).value = "18"
            ws.cell(r, 2).value = NEW_NAME

    # Collect all sector-18 rows, keyed by SOC_Code
    soc_data = {}  # soc_code → { title, total_emp, rows }
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 1).value) == "18":
            soc = ws.cell(r, 3).value  # SOC_Code col C
            title = ws.cell(r, 4).value  # SOC_Title col D
            emp = _safe(ws.cell(r, 5).value)  # Employment col E
            if soc not in soc_data:
                soc_data[soc] = {"title": title, "total_emp": 0, "rows": []}
            soc_data[soc]["total_emp"] += emp
            soc_data[soc]["rows"].append(r)

    # Total sector employment
    total_sector_emp = sum(d["total_emp"] for d in soc_data.values())

    # Find duplicates and resolve
    rows_to_delete = []
    for soc, d in soc_data.items():
        if len(d["rows"]) > 1:
            # Keep the first row, delete the rest
            keep_row = d["rows"][0]
            ws.cell(keep_row, 5).value = round(d["total_emp"], 10)
            share = (d["total_emp"] / total_sector_emp * 100) if total_sector_emp else 0
            ws.cell(keep_row, 6).value = round(share, 2)
            rows_to_delete.extend(d["rows"][1:])
        else:
            # Recalculate share for single entries too
            keep_row = d["rows"][0]
            share = (d["total_emp"] / total_sector_emp * 100) if total_sector_emp else 0
            ws.cell(keep_row, 6).value = round(share, 2)

    # Delete duplicate rows in reverse order
    for r in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r)

    dup_count = len(rows_to_delete)
    print(f"       Combined {dup_count} duplicate SOC rows, total sector emp={total_sector_emp:.1f}K")


# ──────────────────────────────────────────────────────────────────────
# 5f  Job cleanup
# ──────────────────────────────────────────────────────────────────────
def _clean_jobs(wb):
    """Move kept wholesale jobs to sector 18, remove specified jobs, clean tasks."""
    ws_jobs = wb["2 Jobs"]
    ws_lookup = wb["Lookup_Jobs"]
    ws_tasks = wb["3 Tasks"]

    # --- Move wholesale keepers to sector 18 ---
    for r in range(2, ws_jobs.max_row + 1):
        sid = str(ws_jobs.cell(r, 4).value or "")
        custom = ws_jobs.cell(r, 3).value or ""
        if sid == "19" and custom in KEEP_CUSTOM_TITLES_FROM_WHOLESALE:
            ws_jobs.cell(r, 4).value = "18"
            ws_jobs.cell(r, 5).value = NEW_NAME

    for r in range(2, ws_lookup.max_row + 1):
        sid = str(ws_lookup.cell(r, 2).value or "")
        custom = ws_lookup.cell(r, 1).value or ""
        if sid == "19" and custom in KEEP_CUSTOM_TITLES_FROM_WHOLESALE:
            ws_lookup.cell(r, 2).value = "18"

    # --- Collect custom titles of jobs to REMOVE (needed for task matching) ---
    titles_to_remove = set()
    all_remove = REMOVE_FROM_TRANSPORT | REMOVE_FROM_WHOLESALE

    for r in range(2, ws_jobs.max_row + 1):
        custom = ws_jobs.cell(r, 3).value or ""
        if custom in all_remove:
            titles_to_remove.add(custom)

    print(f"       Removing {len(titles_to_remove)} jobs and their tasks")

    # --- Delete from 3 Tasks (match Job_Title col B = Custom_Title from 2 Jobs) ---
    task_rows_to_delete = []
    for r in range(2, ws_tasks.max_row + 1):
        job_title = ws_tasks.cell(r, 2).value or ""  # Job_Title col B (index 2)
        if job_title in titles_to_remove:
            task_rows_to_delete.append(r)

    for r in sorted(task_rows_to_delete, reverse=True):
        ws_tasks.delete_rows(r)
    print(f"       Deleted {len(task_rows_to_delete)} task rows from 3 Tasks")

    # --- Delete from 2 Jobs ---
    job_rows_to_delete = []
    for r in range(2, ws_jobs.max_row + 1):
        custom = ws_jobs.cell(r, 3).value or ""
        if custom in all_remove:
            job_rows_to_delete.append(r)

    for r in sorted(job_rows_to_delete, reverse=True):
        ws_jobs.delete_rows(r)
    print(f"       Deleted {len(job_rows_to_delete)} rows from 2 Jobs")

    # --- Delete from Lookup_Jobs ---
    lookup_rows_to_delete = []
    for r in range(2, ws_lookup.max_row + 1):
        custom = ws_lookup.cell(r, 1).value or ""  # Custom_Title col A
        if custom in all_remove:
            lookup_rows_to_delete.append(r)

    for r in sorted(lookup_rows_to_delete, reverse=True):
        ws_lookup.delete_rows(r)
    print(f"       Deleted {len(lookup_rows_to_delete)} rows from Lookup_Jobs")


# ──────────────────────────────────────────────────────────────────────
# 5g  Industry Frictions merge
# ──────────────────────────────────────────────────────────────────────
def _merge_frictions(ws):
    """Merge Transport + Wholesale rows in Industry Frictions."""
    # Row 4 is the header row; data starts at row 5
    # Col A=ID, B=Name, C=Emp, D=Avg Wage, E-H=T scores, ...
    transport_row = wholesale_row = None
    for r in range(5, ws.max_row + 1):
        name = str(ws.cell(r, 2).value or "")
        if "Transportation" in name:
            transport_row = r
        elif "Wholesale" in name:
            wholesale_row = r

    if not (transport_row and wholesale_row):
        print("       WARNING: could not find Transport/Wholesale rows in Industry Frictions")
        return

    emp_t = _safe(ws.cell(transport_row, 3).value)
    emp_w = _safe(ws.cell(wholesale_row, 3).value)
    wage_t = _safe(ws.cell(transport_row, 4).value)
    wage_w = _safe(ws.cell(wholesale_row, 4).value)

    combined_emp = emp_t + emp_w
    combined_wage = (emp_t * wage_t + emp_w * wage_w) / combined_emp if combined_emp else 0

    # Update transport row
    ws.cell(transport_row, 2).value = NEW_NAME
    ws.cell(transport_row, 3).value = round(combined_emp, 1)
    ws.cell(transport_row, 4).value = round(combined_wage, 0)

    # Keep T/R/E scores from transport row (already there), leave blank if blank

    # Delete wholesale row
    ws.delete_rows(wholesale_row)
    print(f"       Frictions: emp={combined_emp:.1f}K, wage=${combined_wage:,.0f}")
