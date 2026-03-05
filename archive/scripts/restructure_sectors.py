#!/usr/bin/env python3
"""
Sector Restructuring Script for jobs-data-clean.xlsx

Transforms 20 sectors → 17 sectors:
- Removes: Staffing (8), Construction (17), Accommodation (20)
- Splits: Insurance from Finance (new Sector 21)
- Merges: Transport (18) + Wholesale (19) → Logistics & Distribution (18)
- Trims: Manufacturing (15), Retail (16) blue-collar jobs
"""

import shutil
import openpyxl
import os

os.chdir(os.path.dirname(os.path.abspath(__file__)))

# Phase 1: Backup
print("=" * 60)
print("PHASE 1: Backup")
print("=" * 60)
shutil.copy2("jobs-data-clean.xlsx", "jobs-data-clean-backup.xlsx")
print("Backed up to jobs-data-clean-backup.xlsx")

# Load workbook
print("\nLoading workbook...")
wb = openpyxl.load_workbook("jobs-data-clean.xlsx")
print(f"Sheets: {wb.sheetnames}")

# Phase 2-3, 6-7: Removals and trims
print("\n" + "=" * 60)
print("PHASES 2-3, 6-7: Removals, migrations, trims")
print("=" * 60)
from phase_removals import apply_removals
apply_removals(wb)

# Phase 4: Insurance split
print("\n" + "=" * 60)
print("PHASE 4: Insurance split from Finance")
print("=" * 60)
from phase_insurance import apply_insurance_split
apply_insurance_split(wb)

# Phase 5: Transport + Wholesale merge
print("\n" + "=" * 60)
print("PHASE 5: Transport + Wholesale → Logistics & Distribution")
print("=" * 60)
from phase_merge_logistics import apply_logistics_merge
apply_logistics_merge(wb)

# Phase 8: Save
print("\n" + "=" * 60)
print("PHASE 8: Save")
print("=" * 60)
output_file = "jobs-data-clean-v2.xlsx"
wb.save(output_file)
print(f"Saved to {output_file}")

# Verification
print("\n" + "=" * 60)
print("VERIFICATION")
print("=" * 60)
wb2 = openpyxl.load_workbook(output_file, read_only=True)
for s in wb2.sheetnames:
    ws = wb2[s]
    rows = sum(1 for _ in ws.iter_rows())
    print(f"  {s}: {rows} rows")
wb2.close()

print("\nDone! Review jobs-data-clean-v2.xlsx")
