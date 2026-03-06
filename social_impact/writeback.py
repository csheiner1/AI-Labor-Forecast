"""Write merged social impact data to '6 Social Impact' tab in the workbook.

Creates the tab if it doesn't exist. Overwrites all data if it does.
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from social_impact.config import WORKBOOK, MERGED_OUTPUT


# Column order matching the approved design
COLUMNS = [
    "SOC_Code",
    "Job_Title",
    "Pct_Female",
    "Pct_White",
    "Pct_Black",
    "Pct_Asian",
    "Pct_Hispanic",
    "Median_Age",
    "Pct_Over_55",
    "Pct_Bachelors_Plus",
    "Pct_Graduate_Deg",
    "Typical_Entry_Ed",
    "Pct_Foreign_Born",
    "Union_Rate_Pct",
    "Edu_Partisan_Lean",
    "Top_State_1",
    "Top_State_2",
    "Top_State_3",
    "Top_Metro_LQ",
]


def writeback(data=None):
    """Write social impact data to the workbook.

    Args:
        data: list of dicts (if None, loads from merged_social_data.json)
    """
    if data is None:
        with open(MERGED_OUTPUT) as f:
            data = json.load(f)
        print(f"Loaded {len(data)} records from {MERGED_OUTPUT}")

    print(f"Writing to {WORKBOOK}...")
    wb = openpyxl.load_workbook(WORKBOOK)

    # Create or clear the tab
    tab_name = "6 Social Impact"
    if tab_name in wb.sheetnames:
        del wb[tab_name]
    ws = wb.create_sheet(tab_name)

    # Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")

    # Write headers
    for col, header in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Write data
    for i, record in enumerate(data, 2):
        for col, field in enumerate(COLUMNS, 1):
            val = record.get(field)
            ws.cell(row=i, column=col, value=val)

    # Set column widths
    widths = {
        "SOC_Code": 12, "Job_Title": 40, "Pct_Female": 11, "Pct_White": 10,
        "Pct_Black": 10, "Pct_Asian": 10, "Pct_Hispanic": 12, "Median_Age": 11,
        "Pct_Over_55": 11, "Pct_Bachelors_Plus": 15, "Pct_Graduate_Deg": 14,
        "Typical_Entry_Ed": 25, "Pct_Foreign_Born": 14, "Union_Rate_Pct": 13,
        "Edu_Partisan_Lean": 15, "Top_State_1": 18, "Top_State_2": 18,
        "Top_State_3": 18, "Top_Metro_LQ": 35,
    }
    for col, field in enumerate(COLUMNS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = widths.get(field, 12)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(WORKBOOK)
    print(f"  Written {len(data)} rows to '{tab_name}' tab")
    print(f"  {len(COLUMNS)} columns: {', '.join(COLUMNS)}")


if __name__ == "__main__":
    writeback()
