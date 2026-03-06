"""Census occupation code <-> SOC crosswalk.

BLS file: '2018 Census occupation classification titles and code list'
Format: XLSX with columns like Census_Code, Census_Title, SOC_Code, SOC_Title
Many-to-many: one Census code can map to multiple SOC codes (broad groupings),
and multiple Census codes can share the same SOC.
"""
import os
import openpyxl
from collections import defaultdict

from social_impact.config import DATA_CACHE


def load_crosswalk(crosswalk_path=None):
    """Parse the Census->SOC crosswalk file.

    Returns:
        census_to_soc: dict mapping Census code (str) -> list of SOC codes (str)
        soc_to_census: dict mapping SOC code (str) -> list of Census codes (str)
        census_titles: dict mapping Census code (str) -> Census occupation title (str)
                       Used to match against CPSAAT occupation text.
    """
    if crosswalk_path is None:
        crosswalk_path = os.path.join(DATA_CACHE,
            "2018-census-occupation-classification-titles-and-code-list.xlsx")

    wb = openpyxl.load_workbook(crosswalk_path, read_only=True)
    ws = wb.active

    # Read all rows into a list for random access (read_only worksheets
    # only support forward iteration via iter_rows)
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        print("  WARNING: Crosswalk file is empty")
        return {}, {}, {}

    n_cols = len(all_rows[0])

    # Find header row -- look for a row containing 'Census' and 'SOC'
    header_row = None
    census_col = None
    census_title_col = None
    soc_col = None
    for r_idx, row in enumerate(all_rows[:20]):
        row_vals = [str(v or "").strip().lower() for v in row]
        for i, val in enumerate(row_vals):
            if "census" in val and "code" in val:
                census_col = i
            if ("occupation" in val and "title" in val) or ("census" in val and "title" in val):
                census_title_col = i
            if "soc" in val and "code" in val:
                soc_col = i
        if census_col is not None and soc_col is not None:
            header_row = r_idx
            # If no title column found yet, use the first column (common layout)
            if census_title_col is None:
                census_title_col = 0
            break

    if header_row is None:
        # Fallback: try common column positions
        # Typically: col 0 = Occupation title, col 1 = Census code, col 2 = SOC code
        print("  WARNING: Could not find header row, using fallback columns")
        census_title_col = 0
        census_col = 1
        soc_col = 2
        header_row = 6  # typical header row in this file

    census_to_soc = defaultdict(list)
    soc_to_census = defaultdict(list)
    census_titles = {}

    for row in all_rows[header_row + 1:]:
        if len(row) <= max(census_col, soc_col):
            continue
        census_raw = row[census_col]
        soc_raw = row[soc_col]

        if not census_raw or not soc_raw:
            continue

        census_code = str(census_raw).strip()
        soc_code = str(soc_raw).strip()

        # Normalize SOC: remove .00 suffix if present
        if soc_code.endswith(".00"):
            soc_code = soc_code[:-3]

        # Skip non-numeric codes (headers, totals)
        if not census_code[0].isdigit():
            continue

        # Capture Census title
        if census_title_col is not None and len(row) > census_title_col:
            title_raw = row[census_title_col]
            if title_raw:
                census_titles[census_code] = str(title_raw).strip()

        if soc_code not in census_to_soc[census_code]:
            census_to_soc[census_code].append(soc_code)
        if census_code not in soc_to_census[soc_code]:
            soc_to_census[soc_code].append(census_code)

    print(f"  Crosswalk: {len(census_to_soc)} Census codes -> {len(soc_to_census)} SOC codes, "
          f"{len(census_titles)} Census titles")
    return dict(census_to_soc), dict(soc_to_census), dict(census_titles)


def load_project_socs():
    """Load SOC codes from the workbook's 4 Results tab.

    Returns unique SOCs. The workbook may have multiple rows per SOC
    (different sector assignments); we keep the first row's metadata
    (title, sector, employment, wage) and skip subsequent duplicates.

    Returns:
        dict: soc_code -> {title, sector, employment_K, median_wage}
    """
    from social_impact.config import WORKBOOK
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    ws = wb["4 Results"]

    socs = {}
    first = True
    for row in ws.iter_rows(max_col=5, values_only=True):
        if first:
            first = False
            # Validate that first column is the SOC code header
            hdr = str(row[0] or "").strip().lower()
            if "soc" not in hdr and "code" not in hdr:
                print(f"  WARNING: Unexpected first header '{row[0]}', expected SOC_Code column")
            continue
        soc = row[0]
        if not soc:
            continue
        if soc not in socs:
            socs[soc] = {
                "title": row[1],
                "sector": row[2],
                "employment_K": row[3],
                "median_wage": row[4],
            }

    wb.close()
    print(f"  Project SOCs: {len(socs)}")
    return socs


def build_soc_lookup(project_socs, soc_to_census):
    """Build a lookup: for each project SOC, find the Census codes to pull from.

    Handles merged SOCs (comma-separated like "13-2051, 13-2052") by trying
    each individual code.

    Returns:
        dict: project_soc -> list of Census codes
    """
    lookup = {}
    matched = 0
    unmatched = []

    for soc in project_socs:
        # Handle comma-separated merged SOCs
        individual_socs = [s.strip() for s in soc.split(",")]
        census_codes = []
        for s in individual_socs:
            if s in soc_to_census:
                census_codes.extend(soc_to_census[s])
        if census_codes:
            lookup[soc] = list(set(census_codes))
            matched += 1
        else:
            unmatched.append(soc)

    print(f"  SOC->Census lookup: {matched}/{len(project_socs)} matched")
    if unmatched:
        print(f"  Unmatched ({len(unmatched)}): {unmatched[:10]}")

    return lookup
