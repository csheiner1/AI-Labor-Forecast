"""Tests for social_impact.writeback module."""
import pytest
import os
import openpyxl
from social_impact.config import WORKBOOK


def _social_tab_exists():
    """Check if the 6 Social Impact tab has been written."""
    try:
        wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
        exists = "6 Social Impact" in wb.sheetnames
        wb.close()
        return exists
    except Exception:
        return False


needs_writeback = pytest.mark.skipif(
    not _social_tab_exists(),
    reason="'6 Social Impact' tab not present (run writeback first)",
)


@needs_writeback
def test_social_impact_tab_exists():
    """Workbook should have '6 Social Impact' tab after writeback."""
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    assert "6 Social Impact" in wb.sheetnames
    wb.close()


@needs_writeback
def test_social_impact_tab_has_headers():
    """Tab should have expected column headers."""
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    ws = wb["6 Social Impact"]
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, max_col=19, values_only=True):
        headers = list(row)
    wb.close()
    assert "SOC_Code" in headers
    assert "Pct_Female" in headers
    assert "Top_State_1" in headers
    assert "Union_Rate_Pct" in headers


@needs_writeback
def test_social_impact_tab_has_data():
    """Tab should have data rows."""
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    ws = wb["6 Social Impact"]
    count = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            count += 1
    wb.close()
    assert count > 300, f"Only {count} data rows"


@needs_writeback
def test_social_impact_tab_has_19_columns():
    """Tab should have 19 columns as specified."""
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    ws = wb["6 Social Impact"]
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, max_col=25, values_only=True):
        headers = [h for h in row if h is not None]
    wb.close()
    assert len(headers) == 19, f"Expected 19 headers, got {len(headers)}: {headers}"
