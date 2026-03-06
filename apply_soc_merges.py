"""
Apply SOC merge groups to the 2 Staffing Patterns tab.
Merges functionally-identical SOCs within each sector.
"""
import openpyxl
from copy import copy

MERGE_GROUPS = {
    'Software Developers & Programmers': {
        'socs': ['15-1251', '15-1252', '15-1254'],
    },
    'Network Engineers & Administrators': {
        'socs': ['15-1241', '15-1244'],
    },
    'Database Administrators & Architects': {
        'socs': ['15-1242', '15-1243'],
    },
    'IT Support Specialists': {
        'socs': ['15-1231', '15-1232'],
    },
    'Data Scientists & Statisticians': {
        'socs': ['15-2031', '15-2041', '15-2051'],
    },
    'Drafters': {
        'socs': ['17-3011', '17-3012', '17-3013', '17-3019'],
    },
    'Engineering Technologists & Technicians': {
        'socs': ['17-3021', '17-3022', '17-3023', '17-3024', '17-3025',
                 '17-3026', '17-3027', '17-3028', '17-3029'],
    },
    'Surveyors & Cartographers': {
        'socs': ['17-1021', '17-1022', '17-3031'],
    },
    'Physicians & Surgeons': {
        'socs': ['29-1211', '29-1212', '29-1213', '29-1214', '29-1215',
                 '29-1216', '29-1217', '29-1218', '29-1221', '29-1222',
                 '29-1223', '29-1224', '29-1229', '29-1241', '29-1242',
                 '29-1243', '29-1249'],
    },
    'Dentists': {
        'socs': ['29-1021', '29-1022', '29-1023', '29-1024', '29-1029'],
    },
    'Nurses': {
        'socs': ['29-1141', '29-1151', '29-1161', '29-1171', '29-2061'],
    },
    'Therapists': {
        'socs': ['29-1122', '29-1125', '29-1126', '29-1127', '29-1128', '29-1129'],
    },
    'Diagnostic Imaging Technologists': {
        'socs': ['29-2031', '29-2032', '29-2033', '29-2034', '29-2035', '29-2099'],
    },
    'Postsecondary Teachers': {
        'socs': ['25-1011', '25-1021', '25-1022', '25-1031', '25-1032',
                 '25-1041', '25-1042', '25-1043', '25-1051', '25-1052',
                 '25-1053', '25-1054', '25-1061', '25-1062', '25-1063',
                 '25-1064', '25-1065', '25-1066', '25-1067', '25-1071',
                 '25-1072', '25-1081', '25-1082', '25-1111', '25-1112',
                 '25-1113', '25-1121', '25-1122', '25-1124', '25-1125',
                 '25-1126', '25-1191', '25-1192', '25-1193', '25-1194',
                 '25-1199'],
    },
    'Elementary & Middle School Teachers': {
        'socs': ['25-2021', '25-2022', '25-2031', '25-2032'],
    },
    'Secondary & Special Education Teachers': {
        'socs': ['25-2051', '25-2052', '25-2059'],
    },
    'Social Workers': {
        'socs': ['21-1021', '21-1022', '21-1023', '21-1029'],
    },
    'Counselors': {
        'socs': ['21-1013', '21-1015', '21-1018', '21-1019'],
    },
    'Secretaries & Administrative Assistants': {
        'socs': ['43-6011', '43-6012', '43-6013', '43-6014'],
    },
    'Communications Equipment Operators': {
        'socs': ['43-2011', '43-2021', '43-2099'],
    },
    'Biological Scientists': {
        'socs': ['19-1021', '19-1022', '19-1029', '19-1099'],
    },
    'Science Technicians': {
        'socs': ['19-4012', '19-4013', '19-4021', '19-4031', '19-4042', '19-4099'],
    },
    'Writers & Authors': {
        'socs': ['27-3042', '27-3043'],
    },
    'Legal Support Workers': {
        'socs': ['23-2011', '23-2093', '23-2099'],
    },
}

# Build reverse lookup: SOC -> (group_name, soc_list)
SOC_TO_GROUP = {}
for gname, gdef in MERGE_GROUPS.items():
    for soc in gdef['socs']:
        SOC_TO_GROUP[soc] = gname


def format_soc_list(socs_present):
    """Format list of SOC codes for the SOC_Code cell."""
    return ', '.join(sorted(socs_present))


def main():
    wb = openpyxl.load_workbook('jobs-data-v3.xlsx')
    ws = wb['2 Staffing Patterns']

    # Read headers
    headers = [c.value for c in ws[1]]
    h = {headers[i]: i for i in range(len(headers)) if headers[i]}
    print(f"Headers: {headers[:10]}")

    # Read all data rows
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[h['Sector_ID']] or not row[h['SOC_Code']]:
            continue
        rows.append({
            'sector_id': int(row[h['Sector_ID']]) if isinstance(row[h['Sector_ID']], (int, float)) else int(row[h['Sector_ID']]),
            'sector': row[h['Sector']],
            'occ_group': row[h['Occupation_Group']],
            'soc': str(row[h['SOC_Code']]).strip(),
            'title': row[h['SOC_Title']],
            'emp': float(row[h['Employment (Thousands)']]) if row[h['Employment (Thousands)']] else 0,
            'staff_share': float(row[h['Staffing_Share_Pct']]) if row[h['Staffing_Share_Pct']] else 0,
            'ois': float(row[h['Occupation_Industry_Share_Pct']]) if row[h['Occupation_Industry_Share_Pct']] else 0,
            'wage': float(row[h['Median_Wage']]) if row[h['Median_Wage']] else 0,
            'chg': float(row[h['Projected_Change_Pct']]) if row[h['Projected_Change_Pct']] else 0,
        })

    print(f"Read {len(rows)} data rows")

    # Precompute total merged employment per merge group across all sectors
    # (needed for correct OIS% computation)
    group_total_emp = {}
    for gname, gdef in MERGE_GROUPS.items():
        soc_set = set(gdef['socs'])
        total = sum(r['emp'] for r in rows if r['soc'] in soc_set)
        group_total_emp[gname] = total

    # Process: group rows by sector, apply merges
    from collections import defaultdict
    sector_rows = defaultdict(list)
    for r in rows:
        sector_rows[r['sector_id']].append(r)

    output_rows = []
    merge_count = 0
    rows_eliminated = 0

    for sid in sorted(sector_rows.keys()):
        srows = sector_rows[sid]

        # Find which merge groups are active in this sector (2+ SOCs present)
        group_members = defaultdict(list)
        for r in srows:
            if r['soc'] in SOC_TO_GROUP:
                gname = SOC_TO_GROUP[r['soc']]
                group_members[gname].append(r)

        # SOCs that will be merged (skip individually)
        merged_socs = set()
        for gname, members in group_members.items():
            if len(members) >= 2:
                for m in members:
                    merged_socs.add(m['soc'])

        # Add non-merged rows as-is
        for r in srows:
            if r['soc'] not in merged_socs:
                output_rows.append(r)

        # Add merged rows
        for gname, members in sorted(group_members.items()):
            if len(members) < 2:
                continue  # only 1 SOC present, no merge needed

            merge_count += 1
            rows_eliminated += len(members) - 1

            total_emp = sum(m['emp'] for m in members)
            total_staff = sum(m['staff_share'] for m in members)

            # Correct OIS: merged emp in this sector / total merged emp across all sectors
            denom = group_total_emp[gname]
            correct_ois = (total_emp / denom * 100) if denom > 0 else 0

            # Employment-weighted wage and change
            if total_emp > 0:
                wtd_wage = sum(m['emp'] * m['wage'] for m in members) / total_emp
                wtd_chg = sum(m['emp'] * m['chg'] for m in members) / total_emp
            else:
                wtd_wage = members[0]['wage']
                wtd_chg = members[0]['chg']

            # Occupation_Group: Core wins if any member is Core
            has_core = any(m['occ_group'] == 'Core' for m in members)
            if has_core:
                occ_group = 'Core'
            else:
                # Take from highest-employment member
                occ_group = max(members, key=lambda m: m['emp'])['occ_group']

            # SOC code: list all constituent SOCs
            soc_list = format_soc_list([m['soc'] for m in members])

            output_rows.append({
                'sector_id': sid,
                'sector': members[0]['sector'],
                'occ_group': occ_group,
                'soc': soc_list,
                'title': gname,
                'emp': round(total_emp, 1),
                'staff_share': round(total_staff, 2),
                'ois': round(correct_ois, 2),
                'wage': round(wtd_wage, 0),
                'chg': round(wtd_chg, 1),
            })

    # Sort output: by sector_id, then by employment descending
    output_rows.sort(key=lambda r: (r['sector_id'], -r['emp']))

    print(f"Merges applied: {merge_count}")
    print(f"Rows eliminated: {rows_eliminated}")
    print(f"Output rows: {len(output_rows)} (was {len(rows)})")

    # Write back to sheet
    # Clear existing data rows
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, 11):
            ws.cell(row=row_idx, column=col_idx).value = None

    # Write new rows
    for i, r in enumerate(output_rows):
        row_idx = i + 2
        ws.cell(row=row_idx, column=1).value = r['sector_id']
        ws.cell(row=row_idx, column=2).value = r['sector']
        ws.cell(row=row_idx, column=3).value = r['occ_group']
        ws.cell(row=row_idx, column=4).value = r['soc']
        ws.cell(row=row_idx, column=5).value = r['title']
        ws.cell(row=row_idx, column=6).value = r['emp']
        ws.cell(row=row_idx, column=7).value = r['staff_share']
        ws.cell(row=row_idx, column=8).value = r['ois']
        ws.cell(row=row_idx, column=9).value = r['wage']
        ws.cell(row=row_idx, column=10).value = r['chg']

    wb.save('jobs-data-v3.xlsx')
    print("Saved jobs-data-v3.xlsx")

    # Verification
    print(f"\n--- Verification ---")
    # Spot check: Software Devs in Tech (sector 3)
    tech_sw = [r for r in output_rows if r['sector_id'] == 3 and 'Software' in r['title']]
    for r in tech_sw:
        print(f"  Sector 3: SOC={r['soc']}  Title={r['title']}  Emp={r['emp']}K  OIS={r['ois']}%  Group={r['occ_group']}")

    # Spot check: Physicians in Healthcare (sector 4)
    hc_phys = [r for r in output_rows if r['sector_id'] == 4 and 'Physicians' in r['title']]
    for r in hc_phys:
        print(f"  Sector 4: SOC={r['soc'][:60]}...  Title={r['title']}  Emp={r['emp']}K  Group={r['occ_group']}")

    # Spot check: Nurses in Healthcare
    hc_nurses = [r for r in output_rows if r['sector_id'] == 4 and 'Nurses' in r['title']]
    for r in hc_nurses:
        print(f"  Sector 4: SOC={r['soc']}  Title={r['title']}  Emp={r['emp']}K  Group={r['occ_group']}")

    # Spot check: Postsecondary Teachers in Education (sector 11)
    ed_post = [r for r in output_rows if r['sector_id'] == 11 and 'Postsecondary' in r['title']]
    for r in ed_post:
        print(f"  Sector 11: SOC={r['soc'][:60]}...  Title={r['title']}  Emp={r['emp']}K  Group={r['occ_group']}")

    # Check: Network Engineers in Tech should be Core
    tech_net = [r for r in output_rows if r['sector_id'] == 3 and 'Network Engineers' in r['title']]
    for r in tech_net:
        print(f"  Sector 3: SOC={r['soc']}  Title={r['title']}  Emp={r['emp']}K  Group={r['occ_group']}")

    # Check: Legal Support in Law should be Core
    law_legal = [r for r in output_rows if r['sector_id'] == 5 and 'Legal Support' in r['title']]
    for r in law_legal:
        print(f"  Sector 5: SOC={r['soc']}  Title={r['title']}  Emp={r['emp']}K  Group={r['occ_group']}")


if __name__ == '__main__':
    main()
