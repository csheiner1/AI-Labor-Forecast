#!/usr/bin/env python3
"""
Deep-dive analysis of TECHNOLOGY & SOFTWARE sector
under Significant capability / High friction scenario.
Read-only — does NOT modify the workbook.
"""

import openpyxl
from collections import defaultdict

WB_PATH = "jobs-data-v3.xlsx"

# ── Load 5 Results ────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(WB_PATH, read_only=True, data_only=True)
ws = wb["5 Results"]

rows_5 = []
for row in ws.iter_rows(min_row=2, max_row=400, values_only=True):
    soc       = row[0]   # col A
    if soc is None:
        continue
    rec = dict(
        soc           = str(soc).strip(),
        title         = str(row[1] or "").strip(),
        sector        = str(row[2] or "").strip(),
        occ_group     = str(row[3] or "").strip(),
        emp_k         = float(row[4] or 0),
        median_wage   = float(row[5] or 0),
        tc_adj_mod    = float(row[6] or 0),
        tc_adj_sig    = float(row[7] or 0),
        w             = float(row[8] or 0),
        a_mod         = float(row[9] or 0),
        a_sig         = float(row[10] or 0),
        S_mod         = float(row[11] or 0),
        S_sig         = float(row[12] or 0),
        d_max         = float(row[13] or 0),
        E             = float(row[14] or 0),
        T_18mo_high   = float(row[16] or 0),
        R_high        = float(row[18] or 0),
        d_mod_high    = float(row[20] or 0),
        d_sig_high    = float(row[22] or 0),
        disp_mod_high = float(row[24] or 0),
        disp_sig_high = float(row[26] or 0),
    )
    rows_5.append(rec)

# ── Filter to Technology & Software ───────────────────────────────────────
tech = [r for r in rows_5 if r["sector"] == "Technology & Software"]
tech.sort(key=lambda r: r["disp_sig_high"], reverse=True)

total_emp = sum(r["emp_k"] for r in tech)
total_disp_sig = sum(r["disp_sig_high"] for r in tech)
total_disp_mod = sum(r["disp_mod_high"] for r in tech)

print("=" * 120)
print("TECHNOLOGY & SOFTWARE — DEEP DIVE")
print("Scenario: Significant Capability / High Friction")
print("=" * 120)
print(f"\nSector totals: {len(tech)} SOCs | {total_emp:,.1f}K employed | "
      f"{total_disp_sig:,.1f}K displaced (sig) | {total_disp_mod:,.1f}K displaced (mod)")
print(f"Sector displacement rate (sig): {total_disp_sig / total_emp * 100:.2f}%")
print()

# ── 1. Full job listing ──────────────────────────────────────────────────
print("-" * 120)
print("ALL JOBS — sorted by Displaced_K (Significant, High Friction) descending")
print("-" * 120)
hdr = f"{'SOC':<10} {'Job Title':<48} {'Occ Group':<26} {'Emp K':>7} {'Disp K':>7} {'Rate%':>6} {'w':>5} {'a_sig':>6} {'S_sig':>6}"
print(hdr)
print("-" * len(hdr))
for r in tech:
    rate = r["disp_sig_high"] / r["emp_k"] * 100 if r["emp_k"] > 0 else 0
    print(f"{r['soc']:<10} {r['title'][:47]:<48} {r['occ_group'][:25]:<26} "
          f"{r['emp_k']:>7.1f} {r['disp_sig_high']:>7.2f} {rate:>6.1f} "
          f"{r['w']:>5.2f} {r['a_sig']:>6.3f} {r['S_sig']:>6.3f}")

# ── 2. Subtotals by occupation group ─────────────────────────────────────
print()
print("=" * 100)
print("SUBTOTALS BY OCCUPATION GROUP")
print("=" * 100)
occ_groups = defaultdict(lambda: {"emp": 0, "disp_sig": 0, "disp_mod": 0, "count": 0,
                                    "a_sig_sum": 0, "w_sum": 0})
for r in tech:
    g = occ_groups[r["occ_group"]]
    g["emp"] += r["emp_k"]
    g["disp_sig"] += r["disp_sig_high"]
    g["disp_mod"] += r["disp_mod_high"]
    g["count"] += 1
    g["a_sig_sum"] += r["a_sig"] * r["emp_k"]
    g["w_sum"] += r["w"] * r["emp_k"]

print(f"{'Occ Group':<30} {'SOCs':>5} {'Emp K':>8} {'Disp K sig':>11} {'Rate%':>7} {'Wtd a_sig':>10} {'Wtd w':>7}")
print("-" * 80)
for grp_name in sorted(occ_groups.keys()):
    g = occ_groups[grp_name]
    rate = g["disp_sig"] / g["emp"] * 100 if g["emp"] > 0 else 0
    wa = g["a_sig_sum"] / g["emp"] if g["emp"] > 0 else 0
    ww = g["w_sum"] / g["emp"] if g["emp"] > 0 else 0
    print(f"{grp_name[:29]:<30} {g['count']:>5} {g['emp']:>8.1f} {g['disp_sig']:>11.2f} "
          f"{rate:>7.1f} {wa:>10.3f} {ww:>7.2f}")
print(f"{'TOTAL':<30} {len(tech):>5} {total_emp:>8.1f} {total_disp_sig:>11.2f} "
      f"{total_disp_sig/total_emp*100:>7.1f}")

# ── 3. Pipeline component breakdown ──────────────────────────────────────
print()
print("=" * 120)
print("PIPELINE COMPONENT BREAKDOWN — What drives displacement?")
print("=" * 120)
print()
print("Full formula: displacement_rate = d_max * S(a_sig) * E * T_18mo * R")
print("              where a_sig = tc_adj_sig * w")
print()

# Show distributions of each component
import statistics

tc_adjs = [r["tc_adj_sig"] for r in tech]
ws = [r["w"] for r in tech]
a_sigs = [r["a_sig"] for r in tech]
S_sigs = [r["S_sig"] for r in tech]
d_maxs = [r["d_max"] for r in tech]
Es = [r["E"] for r in tech]
Ts = [r["T_18mo_high"] for r in tech]
Rs = [r["R_high"] for r in tech]
rates = [r["d_sig_high"] for r in tech]

def show_dist(name, vals):
    vals_s = sorted(vals)
    print(f"  {name:<14}  min={min(vals):.3f}  p25={vals_s[len(vals_s)//4]:.3f}  "
          f"median={statistics.median(vals):.3f}  p75={vals_s[3*len(vals_s)//4]:.3f}  "
          f"max={max(vals):.3f}  mean={statistics.mean(vals):.3f}")

show_dist("tc_adj_sig", tc_adjs)
show_dist("w", ws)
show_dist("a_sig", a_sigs)
show_dist("S(a_sig)", S_sigs)
show_dist("d_max", d_maxs)
show_dist("E", Es)
show_dist("T_18mo_high", Ts)
show_dist("R_high", Rs)
show_dist("d_sig_high", rates)

# Sensitivity: which component matters most?
# Compare: if we set each component to its sector average, how much does displacement change?
print()
print("COMPONENT CONTRIBUTION ANALYSIS")
print("For each job, decompose the displacement rate into multiplicative contributions:")
print("  d = d_max * S(a) * E * T * R")
print()
print("Sector-wide values (all jobs share same d_max for this sector):")
print(f"  d_max = {tech[0]['d_max']:.4f}  (JOLTS-based, same for all SOCs in sector)")
print()

# Group by unique (E, T, R) combos to show friction variation
friction_combos = defaultdict(list)
for r in tech:
    key = (round(r["E"], 4), round(r["T_18mo_high"], 4), round(r["R_high"], 4))
    friction_combos[key].append(r["occ_group"])

print("Friction parameters by occupation group (E, T_18mo_high, R_high):")
for (e, t, rr), occ_groups_list in sorted(friction_combos.items()):
    unique_groups = sorted(set(occ_groups_list))
    etr = e * t * rr
    print(f"  E={e:.3f}  T={t:.4f}  R={rr:.4f}  →  E*T*R={etr:.4f}  | {', '.join(unique_groups)}")

# ── 4. Software / IT SOCs deep dive ──────────────────────────────────────
print()
print("=" * 120)
print("15-xxxx SOCs (Computer & Mathematical) IN TECH SECTOR — Full Pipeline Values")
print("=" * 120)
print()

sw_socs = [r for r in tech if r["soc"].startswith("15-")]
sw_socs.sort(key=lambda r: r["disp_sig_high"], reverse=True)

if not sw_socs:
    print("  No 15-xxxx SOCs found in Technology & Software sector.")
else:
    print(f"{'SOC':<10} {'Title':<45} {'Emp K':>7} {'tc_adj':>7} {'w':>5} {'a_sig':>6} "
          f"{'S_sig':>6} {'d_max':>6} {'E':>5} {'T_18':>6} {'R':>6} {'d_sig%':>7} {'Disp K':>7}")
    print("-" * 130)
    for r in sw_socs:
        print(f"{r['soc']:<10} {r['title'][:44]:<45} {r['emp_k']:>7.1f} {r['tc_adj_sig']:>7.3f} "
              f"{r['w']:>5.2f} {r['a_sig']:>6.3f} {r['S_sig']:>6.3f} {r['d_max']:>6.3f} "
              f"{r['E']:>5.2f} {r['T_18mo_high']:>6.4f} {r['R_high']:>6.3f} "
              f"{r['d_sig_high']*100:>7.2f} {r['disp_sig_high']:>7.2f}")
    total_sw_emp = sum(r["emp_k"] for r in sw_socs)
    total_sw_disp = sum(r["disp_sig_high"] for r in sw_socs)
    print(f"{'':>10} {'SUBTOTAL':<45} {total_sw_emp:>7.1f} {'':>7} {'':>5} {'':>6} "
          f"{'':>6} {'':>6} {'':>5} {'':>6} {'':>6} {'':>7} {total_sw_disp:>7.2f}")
    print(f"\n  15-xxxx share of sector displacement: {total_sw_disp/total_disp_sig*100:.1f}%")
    print(f"  15-xxxx share of sector employment:   {total_sw_emp/total_emp*100:.1f}%")

# Highlight key dev SOCs
key_dev_socs = ["15-1251", "15-1252", "15-1254", "15-1256"]
print()
print("KEY SOFTWARE DEVELOPER SOCs:")
for soc_code in key_dev_socs:
    matches = [r for r in rows_5 if r["soc"] == soc_code]
    if matches:
        r = matches[0]
        print(f"\n  {r['soc']} — {r['title']}")
        print(f"    Sector: {r['sector']}  |  Occ Group: {r['occ_group']}")
        print(f"    Employment: {r['emp_k']:.1f}K  |  Median Wage: ${r['median_wage']:,.0f}")
        print(f"    tc_adj_sig: {r['tc_adj_sig']:.4f}  |  w: {r['w']:.2f}  |  a_sig: {r['a_sig']:.4f}")
        print(f"    S(a_sig): {r['S_sig']:.4f}")
        print(f"    d_max: {r['d_max']:.4f}  |  E: {r['E']:.2f}  |  T_18mo: {r['T_18mo_high']:.4f}  |  R: {r['R_high']:.4f}")
        print(f"    d_sig_high: {r['d_sig_high']:.4f}  ({r['d_sig_high']*100:.2f}%)")
        print(f"    Displaced: {r['disp_sig_high']:.2f}K")
    else:
        print(f"\n  {soc_code} — NOT FOUND in 5 Results")

# ── 5. Task-level detail for 15-1252 (Software Developers) ───────────────
print()
print("=" * 120)
print("TASK-LEVEL DETAIL: 15-1251/15-1252 (Software Developers & Programmers)")
print("=" * 120)

ws3 = wb["3 Tasks"]
tasks = []
for row in ws3.iter_rows(min_row=2, max_row=7000, values_only=True):
    soc = row[0]
    if soc is None:
        continue
    soc_str = str(soc).strip()
    # Tasks are keyed as "15-1251, 15-1252" in the workbook
    if "15-1252" in soc_str or "15-1251" in soc_str:
        task = dict(
            soc         = soc_str,
            title       = str(row[1] or "").strip(),
            task_id     = str(row[2] or "").strip(),
            description = str(row[3] or "").strip(),
            task_type   = str(row[4] or "").strip(),
            time_share  = float(row[5] or 0),
            importance  = float(row[6] or 0),
            frequency   = str(row[7] or "").strip(),
            gwa         = str(row[8] or "").strip(),
            aut_mod     = float(row[11] or 0) if row[11] is not None else None,
            aut_sig     = float(row[12] or 0) if row[12] is not None else None,
        )
        tasks.append(task)

if not tasks:
    print("  No tasks found.")
else:
    tasks.sort(key=lambda t: t["time_share"], reverse=True)
    print(f"\n{len(tasks)} tasks found | SOC label: {tasks[0]['soc']}")
    print()
    print(f"{'#':>2} {'Task ID':<28} {'Type':<7} {'Time%':>6} {'Imp':>4} {'Freq':<8} "
          f"{'Aut_Mod':>8} {'Aut_Sig':>8} {'Description'}")
    print("-" * 140)
    for i, t in enumerate(tasks, 1):
        aut_mod_str = f"{t['aut_mod']:.2f}" if t['aut_mod'] is not None else "n/a"
        aut_sig_str = f"{t['aut_sig']:.2f}" if t['aut_sig'] is not None else "n/a"
        desc = t['description'][:60] + "..." if len(t['description']) > 60 else t['description']
        print(f"{i:>2} {t['task_id']:<28} {t['task_type']:<7} {t['time_share']:>6.1f} "
              f"{t['importance']:>4.0f} {t['frequency']:<8} {aut_mod_str:>8} {aut_sig_str:>8} {desc}")

    # Compute tc_adj from raw task data to verify
    print()
    print("TASK-LEVEL STATISTICS (computed from raw task data):")
    total_ts = sum(t["time_share"] for t in tasks)
    print(f"  Total time_share: {total_ts:.1f}%")

    # Weighted mean autonomy (importance * time_share weighted)
    tasks_with_aut = [t for t in tasks if t["aut_sig"] is not None]
    if tasks_with_aut:
        weight_sum = sum(t["importance"] * t["time_share"] for t in tasks_with_aut)
        tc_mean_sig = sum(t["importance"] * t["time_share"] * t["aut_sig"] for t in tasks_with_aut) / weight_sum if weight_sum > 0 else 0
        tc_mean_mod = sum(t["importance"] * t["time_share"] * t["aut_mod"] for t in tasks_with_aut) / weight_sum if weight_sum > 0 else 0

        # z = time share of zero-autonomy tasks
        z_sig = sum(t["time_share"] for t in tasks_with_aut if t["aut_sig"] == 0) / total_ts
        z_mod = sum(t["time_share"] for t in tasks_with_aut if t["aut_mod"] == 0) / total_ts

        # h = time share of high-autonomy (>= 0.65) tasks
        h_sig = sum(t["time_share"] for t in tasks_with_aut if t["aut_sig"] >= 0.65) / total_ts
        h_mod = sum(t["time_share"] for t in tasks_with_aut if t["aut_mod"] >= 0.65) / total_ts

        # tc_adj = tc_mean * (1-z)^0.5 * min(1, h/tc_mean)
        tc_adj_sig = tc_mean_sig * (1 - z_sig)**0.5 * min(1, h_sig / tc_mean_sig) if tc_mean_sig > 0 else 0
        tc_adj_mod = tc_mean_mod * (1 - z_mod)**0.5 * min(1, h_mod / tc_mean_mod) if tc_mean_mod > 0 else 0

        print(f"\n  Significant scenario:")
        print(f"    tc_mean (imp*ts weighted): {tc_mean_sig:.4f}")
        print(f"    z (zero-autonomy share):   {z_sig:.4f}")
        print(f"    h (high-autonomy share):   {h_sig:.4f}")
        print(f"    tc_adj = {tc_mean_sig:.4f} * {(1-z_sig)**0.5:.4f} * min(1, {h_sig:.4f}/{tc_mean_sig:.4f})")
        print(f"           = {tc_adj_sig:.4f}")

        print(f"\n  Moderate scenario:")
        print(f"    tc_mean (imp*ts weighted): {tc_mean_mod:.4f}")
        print(f"    z (zero-autonomy share):   {z_mod:.4f}")
        print(f"    h (high-autonomy share):   {h_mod:.4f}")
        print(f"    tc_adj = {tc_mean_mod:.4f} * {(1-z_mod)**0.5:.4f} * min(1, {h_mod:.4f}/{tc_mean_mod:.4f})")
        print(f"           = {tc_adj_mod:.4f}")

        # Autonomy distribution
        print()
        print("  AUTONOMY DISTRIBUTION (Significant):")
        buckets = {"0.00": 0, "0.01-0.24": 0, "0.25-0.49": 0, "0.50-0.64": 0, "0.65-0.84": 0, "0.85-1.00": 0}
        for t in tasks_with_aut:
            a = t["aut_sig"]
            ts = t["time_share"]
            if a == 0:
                buckets["0.00"] += ts
            elif a < 0.25:
                buckets["0.01-0.24"] += ts
            elif a < 0.50:
                buckets["0.25-0.49"] += ts
            elif a < 0.65:
                buckets["0.50-0.64"] += ts
            elif a < 0.85:
                buckets["0.65-0.84"] += ts
            else:
                buckets["0.85-1.00"] += ts
        for bucket, ts in buckets.items():
            bar = "#" * int(ts / total_ts * 50)
            print(f"    {bucket:<12}  {ts:>5.1f}% time  ({ts/total_ts*100:>5.1f}%)  {bar}")

# ── 6. Cross-sector comparison for context ────────────────────────────────
print()
print("=" * 100)
print("CONTEXT: Technology & Software vs. All Sectors")
print("=" * 100)

sector_stats = defaultdict(lambda: {"emp": 0, "disp_sig": 0, "count": 0})
for r in rows_5:
    s = sector_stats[r["sector"]]
    s["emp"] += r["emp_k"]
    s["disp_sig"] += r["disp_sig_high"]
    s["count"] += 1

print(f"\n{'Sector':<35} {'SOCs':>5} {'Emp K':>10} {'Disp K sig':>11} {'Rate%':>7}")
print("-" * 70)
for sname in sorted(sector_stats.keys(), key=lambda s: sector_stats[s]["disp_sig"], reverse=True):
    s = sector_stats[sname]
    rate = s["disp_sig"] / s["emp"] * 100 if s["emp"] > 0 else 0
    marker = " <<<" if sname == "Technology & Software" else ""
    print(f"{sname[:34]:<35} {s['count']:>5} {s['emp']:>10.1f} {s['disp_sig']:>11.2f} {rate:>7.2f}{marker}")

grand_emp = sum(s["emp"] for s in sector_stats.values())
grand_disp = sum(s["disp_sig"] for s in sector_stats.values())
print(f"{'ALL SECTORS':<35} {sum(s['count'] for s in sector_stats.values()):>5} "
      f"{grand_emp:>10.1f} {grand_disp:>11.2f} {grand_disp/grand_emp*100:>7.2f}")

tech_s = sector_stats["Technology & Software"]
print(f"\nTech & Software share of all displacement: {tech_s['disp_sig']/grand_disp*100:.1f}%")
print(f"Tech & Software share of all employment:   {tech_s['emp']/grand_emp*100:.1f}%")

wb.close()
print("\nDone.")
