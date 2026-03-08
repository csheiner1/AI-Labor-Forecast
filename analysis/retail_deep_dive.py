"""Deep-dive analysis of Retail Trade sector under Significant / High friction scenario.

Reads the 5 Results tab from jobs-data-v3.xlsx (read-only) and prints:
1. All Retail Trade jobs sorted by displaced_K_sig_high descending
2. Subtotals by occupation group
3. Drivers of Retail's displacement rate
4. Comparison of Retail's d_max, E, T, R vs overall averages
"""
import os
import openpyxl
from collections import defaultdict

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
WORKBOOK = os.path.join(PROJECT_ROOT, "jobs-data-v3.xlsx")

# Column indices (1-based) from phase5_writeback.py header definition:
# 1=SOC_Code, 2=Job_Title, 3=Sector, 4=Occupation_Group,
# 5=Employment_2024_K, 6=Median_Wage,
# 7=tc_adj_mod, 8=tc_adj_sig, 9=w,
# 10=a_mod, 11=a_sig, 12=S_mod, 13=S_sig,
# 14=d_max, 15=E,
# 16=T_18mo_low, 17=T_18mo_high,
# 18=R_low, 19=R_high,
# 20=d_mod_low, 21=d_mod_high,
# 22=d_sig_low, 23=d_sig_high,
# 24=displaced_K_mod_low, 25=displaced_K_mod_high,
# 26=displaced_K_sig_low, 27=displaced_K_sig_high

COL = {
    "soc": 1, "title": 2, "sector": 3, "occ_group": 4,
    "emp_K": 5, "wage": 6,
    "tc_adj_mod": 7, "tc_adj_sig": 8, "w": 9,
    "a_mod": 10, "a_sig": 11, "S_mod": 12, "S_sig": 13,
    "d_max": 14, "E": 15,
    "T_18mo_low": 16, "T_18mo_high": 17,
    "R_low": 18, "R_high": 19,
    "d_mod_low": 20, "d_mod_high": 21,
    "d_sig_low": 22, "d_sig_high": 23,
    "disp_K_mod_low": 24, "disp_K_mod_high": 25,
    "disp_K_sig_low": 26, "disp_K_sig_high": 27,
}


def val(ws, row, col_name):
    """Read cell value by column name."""
    return ws.cell(row, COL[col_name]).value


def num(v, default=0):
    """Coerce to float, defaulting None to default."""
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def main():
    print("=" * 100)
    print("RETAIL TRADE DEEP DIVE — Significant Capability / High Friction Scenario")
    print("=" * 100)

    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb["5 Results"]

    # ── Read all data ──────────────────────────────────────────────────
    all_jobs = []
    for r in range(2, ws.max_row + 1):
        soc = val(ws, r, "soc")
        if not soc:
            continue
        row_data = {k: val(ws, r, k) for k in COL}
        all_jobs.append(row_data)

    retail_jobs = [j for j in all_jobs if j["sector"] == "Retail Trade"]

    if not retail_jobs:
        print("ERROR: No jobs found with Sector == 'Retail Trade'")
        # Show all unique sectors for debugging
        sectors = sorted(set(j["sector"] for j in all_jobs if j["sector"]))
        print(f"Available sectors ({len(sectors)}): {sectors}")
        return

    # Sort by displaced_K_sig_high descending
    retail_jobs.sort(key=lambda j: num(j["disp_K_sig_high"]), reverse=True)

    # ── Section 1: All Retail Trade jobs ───────────────────────────────
    print(f"\n{'─' * 100}")
    print(f"1. ALL RETAIL TRADE JOBS ({len(retail_jobs)} SOCs) — sorted by Displaced K (sig/high) desc")
    print(f"{'─' * 100}")
    header = f"{'SOC':<12} {'Job Title':<45} {'Occ Group':<25} {'Emp K':>7} {'Disp K':>7} {'Rate%':>6} {'w':>5} {'a_sig':>6} {'S_sig':>6}"
    print(header)
    print("-" * len(header))

    total_emp = 0
    total_disp = 0
    for j in retail_jobs:
        emp = num(j["emp_K"])
        disp = num(j["disp_K_sig_high"])
        rate = (disp / emp * 100) if emp > 0 else 0
        total_emp += emp
        total_disp += disp
        print(f"{j['soc']:<12} {str(j['title'])[:44]:<45} {str(j['occ_group'])[:24]:<25} "
              f"{emp:>7.1f} {disp:>7.1f} {rate:>5.1f}% "
              f"{num(j['w']):>5.2f} {num(j['a_sig']):>6.3f} {num(j['S_sig']):>6.3f}")

    overall_rate = (total_disp / total_emp * 100) if total_emp > 0 else 0
    print("-" * len(header))
    print(f"{'TOTAL':<12} {'':<45} {'':<25} {total_emp:>7.1f} {total_disp:>7.1f} {overall_rate:>5.1f}%")

    # ── Section 2: Subtotals by occupation group ───────────────────────
    print(f"\n{'─' * 100}")
    print("2. SUBTOTALS BY OCCUPATION GROUP (within Retail Trade)")
    print(f"{'─' * 100}")

    by_group = defaultdict(lambda: {
        "count": 0, "emp": 0, "disp": 0,
        "sum_a_sig": 0, "sum_S_sig": 0, "sum_w": 0,
        "sum_E": 0, "sum_T": 0, "sum_R": 0,
        "emp_weighted_a": 0, "emp_weighted_S": 0,
    })

    for j in retail_jobs:
        g = j["occ_group"] or "Unknown"
        emp = num(j["emp_K"])
        disp = num(j["disp_K_sig_high"])
        a_sig = num(j["a_sig"])
        S_sig = num(j["S_sig"])
        w = num(j["w"])
        E = num(j["E"], 1.0)
        T = num(j["T_18mo_high"], 0.5)
        R = num(j["R_high"], 1.0)

        by_group[g]["count"] += 1
        by_group[g]["emp"] += emp
        by_group[g]["disp"] += disp
        by_group[g]["sum_a_sig"] += a_sig
        by_group[g]["sum_S_sig"] += S_sig
        by_group[g]["sum_w"] += w
        by_group[g]["sum_E"] += E
        by_group[g]["sum_T"] += T
        by_group[g]["sum_R"] += R
        by_group[g]["emp_weighted_a"] += emp * a_sig
        by_group[g]["emp_weighted_S"] += emp * S_sig

    header2 = f"{'Occ Group':<30} {'#SOCs':>5} {'Emp K':>8} {'Disp K':>8} {'Rate%':>6} {'Avg w':>6} {'Avg a':>6} {'Avg S':>6} {'E':>5} {'T':>5} {'R':>5}"
    print(header2)
    print("-" * len(header2))

    for g in sorted(by_group.keys()):
        d = by_group[g]
        rate = (d["disp"] / d["emp"] * 100) if d["emp"] > 0 else 0
        avg_a = d["emp_weighted_a"] / d["emp"] if d["emp"] > 0 else d["sum_a_sig"] / d["count"]
        avg_S = d["emp_weighted_S"] / d["emp"] if d["emp"] > 0 else d["sum_S_sig"] / d["count"]
        avg_w = d["sum_w"] / d["count"] if d["count"] > 0 else 0
        avg_E = d["sum_E"] / d["count"] if d["count"] > 0 else 0
        avg_T = d["sum_T"] / d["count"] if d["count"] > 0 else 0
        avg_R = d["sum_R"] / d["count"] if d["count"] > 0 else 0
        print(f"{g[:29]:<30} {d['count']:>5} {d['emp']:>8.1f} {d['disp']:>8.1f} {rate:>5.1f}% "
              f"{avg_w:>6.2f} {avg_a:>6.3f} {avg_S:>6.3f} {avg_E:>5.2f} {avg_T:>5.2f} {avg_R:>5.2f}")

    print("-" * len(header2))
    print(f"{'TOTAL':<30} {len(retail_jobs):>5} {total_emp:>8.1f} {total_disp:>8.1f} {overall_rate:>5.1f}%")

    # ── Section 3: What drives Retail's high displacement rate? ─────────
    print(f"\n{'─' * 100}")
    print("3. DRIVERS OF RETAIL TRADE DISPLACEMENT")
    print(f"{'─' * 100}")

    # Retail sector parameters
    retail_d_max_vals = [num(j["d_max"]) for j in retail_jobs]
    retail_d_max = retail_d_max_vals[0] if retail_d_max_vals else 0  # same for all in sector
    retail_E_vals = [num(j["E"], None) for j in retail_jobs]
    retail_T_vals = [num(j["T_18mo_high"], None) for j in retail_jobs]
    retail_R_vals = [num(j["R_high"], None) for j in retail_jobs]
    retail_a_vals = [num(j["a_sig"]) for j in retail_jobs]
    retail_S_vals = [num(j["S_sig"]) for j in retail_jobs]
    retail_w_vals = [num(j["w"]) for j in retail_jobs]

    # Employment-weighted averages for Retail
    retail_avg_a = sum(num(j["emp_K"]) * num(j["a_sig"]) for j in retail_jobs) / total_emp if total_emp > 0 else 0
    retail_avg_S = sum(num(j["emp_K"]) * num(j["S_sig"]) for j in retail_jobs) / total_emp if total_emp > 0 else 0
    retail_avg_w = sum(num(j["emp_K"]) * num(j["w"]) for j in retail_jobs) / total_emp if total_emp > 0 else 0
    retail_avg_tc_adj_sig = sum(num(j["emp_K"]) * num(j["tc_adj_sig"]) for j in retail_jobs) / total_emp if total_emp > 0 else 0

    # Unweighted averages for E, T, R (they vary by occ group, not by SOC)
    retail_E_clean = [v for v in retail_E_vals if v is not None and v != 0]
    retail_T_clean = [v for v in retail_T_vals if v is not None and v != 0]
    retail_R_clean = [v for v in retail_R_vals if v is not None and v != 0]

    retail_avg_E = sum(retail_E_clean) / len(retail_E_clean) if retail_E_clean else 0
    retail_avg_T = sum(retail_T_clean) / len(retail_T_clean) if retail_T_clean else 0
    retail_avg_R = sum(retail_R_clean) / len(retail_R_clean) if retail_R_clean else 0

    print(f"\nRetail Trade sector parameters:")
    print(f"  d_max              = {retail_d_max:.4f}")
    print(f"  Avg E (unweighted) = {retail_avg_E:.4f}")
    print(f"  Avg T_18mo_high    = {retail_avg_T:.4f}")
    print(f"  Avg R_high         = {retail_avg_R:.4f}")
    print(f"  Emp-weighted avg tc_adj_sig = {retail_avg_tc_adj_sig:.4f}")
    print(f"  Emp-weighted avg w          = {retail_avg_w:.4f}")
    print(f"  Emp-weighted avg a_sig      = {retail_avg_a:.4f}")
    print(f"  Emp-weighted avg S_sig      = {retail_avg_S:.4f}")

    # Decompose: d = d_max * S(a) * E * T * R
    product_E_T_R = retail_avg_E * retail_avg_T * retail_avg_R
    implied_rate = retail_d_max * retail_avg_S * product_E_T_R
    print(f"\n  Implied avg displacement rate = d_max * S * E * T * R")
    print(f"    = {retail_d_max:.4f} * {retail_avg_S:.4f} * {retail_avg_E:.4f} * {retail_avg_T:.4f} * {retail_avg_R:.4f}")
    print(f"    = {implied_rate:.4f} ({implied_rate * 100:.1f}%)")
    print(f"  Actual emp-weighted rate     = {overall_rate:.1f}%")

    # Identify top contributors to total displaced K
    print(f"\n  Top 5 jobs by displaced K (sig/high):")
    for i, j in enumerate(retail_jobs[:5]):
        emp = num(j["emp_K"])
        disp = num(j["disp_K_sig_high"])
        rate = (disp / emp * 100) if emp > 0 else 0
        share = (disp / total_disp * 100) if total_disp > 0 else 0
        print(f"    {i+1}. {j['title']} ({j['soc']})")
        print(f"       Emp={emp:.1f}K, Displaced={disp:.1f}K, Rate={rate:.1f}%, "
              f"Share of total={share:.1f}%, a_sig={num(j['a_sig']):.3f}, w={num(j['w']):.2f}")

    # Employment concentration
    top5_disp = sum(num(j["disp_K_sig_high"]) for j in retail_jobs[:5])
    top10_disp = sum(num(j["disp_K_sig_high"]) for j in retail_jobs[:10])
    print(f"\n  Employment/Displacement Concentration:")
    print(f"    Top 5 jobs account for {top5_disp:.1f}K of {total_disp:.1f}K ({top5_disp/total_disp*100:.1f}%) displaced")
    print(f"    Top 10 jobs account for {top10_disp:.1f}K of {total_disp:.1f}K ({top10_disp/total_disp*100:.1f}%) displaced")

    # ── Section 4: Compare Retail to overall averages ──────────────────
    print(f"\n{'─' * 100}")
    print("4. RETAIL vs. ALL-SECTOR AVERAGES")
    print(f"{'─' * 100}")

    # Compute overall averages across ALL sectors
    by_sector = defaultdict(lambda: {
        "emp": 0, "disp_sig_high": 0,
        "sum_a_emp": 0, "sum_S_emp": 0, "sum_w_emp": 0, "sum_tc_adj_emp": 0,
        "d_max": 0, "E_sum": 0, "T_sum": 0, "R_sum": 0, "count": 0,
    })

    for j in all_jobs:
        sec = j["sector"] or "Unknown"
        emp = num(j["emp_K"])
        disp = num(j["disp_K_sig_high"])
        by_sector[sec]["emp"] += emp
        by_sector[sec]["disp_sig_high"] += disp
        by_sector[sec]["sum_a_emp"] += emp * num(j["a_sig"])
        by_sector[sec]["sum_S_emp"] += emp * num(j["S_sig"])
        by_sector[sec]["sum_w_emp"] += emp * num(j["w"])
        by_sector[sec]["sum_tc_adj_emp"] += emp * num(j["tc_adj_sig"])
        by_sector[sec]["d_max"] = num(j["d_max"])  # same for all jobs in sector
        by_sector[sec]["E_sum"] += num(j["E"], 0)
        by_sector[sec]["T_sum"] += num(j["T_18mo_high"], 0)
        by_sector[sec]["R_sum"] += num(j["R_high"], 0)
        by_sector[sec]["count"] += 1

    total_all_emp = sum(d["emp"] for d in by_sector.values())
    total_all_disp = sum(d["disp_sig_high"] for d in by_sector.values())

    # Sector comparison table
    header3 = f"{'Sector':<30} {'Emp K':>8} {'Disp K':>8} {'Rate%':>6} {'d_max':>6} {'Avg E':>6} {'Avg T':>6} {'Avg R':>6} {'Avg a':>6} {'Avg S':>6}"
    print(header3)
    print("-" * len(header3))

    sector_rows = []
    for sec in sorted(by_sector.keys()):
        d = by_sector[sec]
        rate = (d["disp_sig_high"] / d["emp"] * 100) if d["emp"] > 0 else 0
        avg_a = d["sum_a_emp"] / d["emp"] if d["emp"] > 0 else 0
        avg_S = d["sum_S_emp"] / d["emp"] if d["emp"] > 0 else 0
        avg_E = d["E_sum"] / d["count"] if d["count"] > 0 else 0
        avg_T = d["T_sum"] / d["count"] if d["count"] > 0 else 0
        avg_R = d["R_sum"] / d["count"] if d["count"] > 0 else 0
        marker = " ◀ RETAIL" if sec == "Retail Trade" else ""
        sector_rows.append((sec, d["emp"], d["disp_sig_high"], rate, d["d_max"], avg_E, avg_T, avg_R, avg_a, avg_S))
        print(f"{sec[:29]:<30} {d['emp']:>8.1f} {d['disp_sig_high']:>8.1f} {rate:>5.1f}% "
              f"{d['d_max']:>6.3f} {avg_E:>6.2f} {avg_T:>6.2f} {avg_R:>6.2f} {avg_a:>6.3f} {avg_S:>6.3f}{marker}")

    print("-" * len(header3))
    overall_all_rate = (total_all_disp / total_all_emp * 100) if total_all_emp > 0 else 0
    # Compute employment-weighted overall averages
    all_avg_a = sum(d["sum_a_emp"] for d in by_sector.values()) / total_all_emp if total_all_emp > 0 else 0
    all_avg_S = sum(d["sum_S_emp"] for d in by_sector.values()) / total_all_emp if total_all_emp > 0 else 0
    all_total_count = sum(d["count"] for d in by_sector.values())
    all_avg_E = sum(d["E_sum"] for d in by_sector.values()) / all_total_count if all_total_count > 0 else 0
    all_avg_T = sum(d["T_sum"] for d in by_sector.values()) / all_total_count if all_total_count > 0 else 0
    all_avg_R = sum(d["R_sum"] for d in by_sector.values()) / all_total_count if all_total_count > 0 else 0
    # Emp-weighted d_max
    all_avg_d_max = sum(d["d_max"] * d["emp"] for d in by_sector.values()) / total_all_emp if total_all_emp > 0 else 0

    print(f"{'OVERALL':<30} {total_all_emp:>8.1f} {total_all_disp:>8.1f} {overall_all_rate:>5.1f}% "
          f"{all_avg_d_max:>6.3f} {all_avg_E:>6.2f} {all_avg_T:>6.2f} {all_avg_R:>6.2f} {all_avg_a:>6.3f} {all_avg_S:>6.3f}")

    # ── Direct comparison ──────────────────────────────────────────────
    print(f"\n  DIRECT COMPARISON: Retail Trade vs Overall Average")
    print(f"  {'Parameter':<25} {'Retail':>10} {'Overall':>10} {'Ratio':>8} {'Delta':>10}")
    print(f"  {'-'*65}")

    comparisons = [
        ("d_max", retail_d_max, all_avg_d_max),
        ("Avg E", retail_avg_E, all_avg_E),
        ("Avg T_18mo_high", retail_avg_T, all_avg_T),
        ("Avg R_high", retail_avg_R, all_avg_R),
        ("Emp-wtd a_sig", retail_avg_a, all_avg_a),
        ("Emp-wtd S_sig", retail_avg_S, all_avg_S),
        ("Displacement rate %", overall_rate, overall_all_rate),
    ]

    for name, retail_val, overall_val in comparisons:
        ratio = retail_val / overall_val if overall_val != 0 else float('inf')
        delta = retail_val - overall_val
        print(f"  {name:<25} {retail_val:>10.4f} {overall_val:>10.4f} {ratio:>7.2f}x {delta:>+10.4f}")

    # ── Multiplicative decomposition ───────────────────────────────────
    print(f"\n  MULTIPLICATIVE DECOMPOSITION")
    print(f"  d = d_max * S(a) * E * T * R")
    print(f"  Retail's rate is {overall_rate / overall_all_rate:.2f}x the overall average ({overall_rate:.1f}% vs {overall_all_rate:.1f}%)")
    print(f"  Contribution by factor (ratio of Retail/Overall):")
    for name, retail_val, overall_val in comparisons[:6]:
        ratio = retail_val / overall_val if overall_val != 0 else float('inf')
        direction = "HIGHER" if ratio > 1.05 else ("LOWER" if ratio < 0.95 else "SIMILAR")
        print(f"    {name:<25}: {ratio:.2f}x ({direction})")

    # ── Distribution analysis ──────────────────────────────────────────
    print(f"\n{'─' * 100}")
    print("5. DISTRIBUTION OF AUTONOMY AND DISPLACEMENT WITHIN RETAIL")
    print(f"{'─' * 100}")

    # Autonomy distribution
    a_bins = [(0, 0.1), (0.1, 0.2), (0.2, 0.3), (0.3, 0.4), (0.4, 0.5), (0.5, 0.6), (0.6, 0.7), (0.7, 0.8), (0.8, 1.01)]
    print(f"\n  {'a_sig range':<15} {'#SOCs':>6} {'Emp K':>8} {'Disp K':>8} {'Avg Rate%':>10}")
    print(f"  {'-'*50}")
    for lo, hi in a_bins:
        in_bin = [j for j in retail_jobs if lo <= num(j["a_sig"]) < hi]
        if in_bin:
            emp = sum(num(j["emp_K"]) for j in in_bin)
            disp = sum(num(j["disp_K_sig_high"]) for j in in_bin)
            rate = (disp / emp * 100) if emp > 0 else 0
            print(f"  {lo:.1f} - {hi:.1f}       {len(in_bin):>6} {emp:>8.1f} {disp:>8.1f} {rate:>9.1f}%")

    # w distribution
    print(f"\n  {'w value':<15} {'#SOCs':>6} {'Emp K':>8} {'Disp K':>8}")
    print(f"  {'-'*40}")
    for w_val in [0.25, 0.50, 0.75, 1.00]:
        in_bin = [j for j in retail_jobs if abs(num(j["w"]) - w_val) < 0.01]
        if in_bin:
            emp = sum(num(j["emp_K"]) for j in in_bin)
            disp = sum(num(j["disp_K_sig_high"]) for j in in_bin)
            print(f"  {w_val:.2f}           {len(in_bin):>6} {emp:>8.1f} {disp:>8.1f}")

    # Rank Retail among all sectors
    sector_rates = []
    for sec, d in by_sector.items():
        if d["emp"] > 0:
            rate = d["disp_sig_high"] / d["emp"] * 100
            sector_rates.append((sec, rate, d["emp"], d["disp_sig_high"]))
    sector_rates.sort(key=lambda x: x[1], reverse=True)

    print(f"\n  RETAIL TRADE RANK AMONG ALL SECTORS (by displacement rate):")
    for i, (sec, rate, emp, disp) in enumerate(sector_rates):
        marker = "  <<<" if sec == "Retail Trade" else ""
        print(f"    {i+1:>2}. {sec:<30} {rate:>5.1f}% ({disp:>7.1f}K / {emp:>8.1f}K){marker}")

    wb.close()
    print(f"\n{'=' * 100}")
    print("END OF RETAIL TRADE DEEP DIVE")
    print(f"{'=' * 100}")


if __name__ == "__main__":
    main()
