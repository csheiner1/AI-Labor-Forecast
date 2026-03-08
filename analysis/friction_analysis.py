"""Read-only analysis of friction parameters for Retail Trade and Construction.

Reads the 4H Frictions High tab and computes derived values.
Does NOT modify the workbook.
"""
import math
import os
import openpyxl

WORKBOOK = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                        "jobs-data-v3.xlsx")


def compute_T_18mo(T1, T2, T3, T4, alpha=1.2):
    D = (T1 + T2 + T3 + T4) / 4
    t0 = 1.5 * D - 0.5 * T4 - 0.5
    T_val = 1 / (1 + math.exp(-alpha * (1.5 - t0)))
    return D, t0, T_val


def compute_R(f1, f2, f3):
    F = f1 + f2 + f3
    R = 1 - 0.7 * (F - 3) / 9
    return F, R


def read_frictions_tab(wb, tab_name="4H Frictions High"):
    ws = wb[tab_name]
    rows = []
    for r in range(5, ws.max_row + 1):
        sector = ws.cell(r, 2).value
        occ_group = ws.cell(r, 3).value
        if not sector or not occ_group:
            continue
        T1 = ws.cell(r, 6).value
        T2 = ws.cell(r, 7).value
        T3 = ws.cell(r, 8).value
        T4 = ws.cell(r, 9).value
        f1 = ws.cell(r, 13).value
        f2 = ws.cell(r, 14).value
        f3 = ws.cell(r, 15).value
        E = ws.cell(r, 18).value

        if all(x is None for x in [T1, T2, T3, T4, f1, f2, f3, E]):
            continue

        # Compute derived values
        if None not in (T1, T2, T3, T4):
            D, t0, T_18mo = compute_T_18mo(T1, T2, T3, T4)
        else:
            D, t0, T_18mo = None, None, None

        if None not in (f1, f2, f3):
            F, R = compute_R(f1, f2, f3)
        else:
            F, R = None, None

        # Friction discount = E * T * R
        if E is not None and T_18mo is not None and R is not None:
            friction_discount = E * T_18mo * R
        else:
            friction_discount = None

        rows.append({
            "sector": sector,
            "occ_group": occ_group,
            "T1": T1, "T2": T2, "T3": T3, "T4": T4,
            "D": D, "t0": t0, "T_18mo": T_18mo,
            "f1": f1, "f2": f2, "f3": f3,
            "F": F, "R": R,
            "E": E,
            "friction_discount": friction_discount,
        })
    return rows


def print_table(rows, title):
    print(f"\n{'='*120}")
    print(f"  {title}")
    print(f"{'='*120}")
    hdr = (f"{'Sector':<20} {'Occ Group':<30} "
           f"{'T1':>3} {'T2':>3} {'T3':>3} {'T4':>3} "
           f"{'D':>5} {'t0':>6} {'T18':>5} "
           f"{'f1':>3} {'f2':>3} {'f3':>3} "
           f"{'F':>3} {'R':>5} "
           f"{'E':>5} "
           f"{'E*T*R':>6}")
    print(hdr)
    print("-" * 120)
    for row in rows:
        def fmt(v, w=3, dec=None):
            if v is None:
                return " " * w
            if dec is not None:
                return f"{v:>{w}.{dec}f}"
            if isinstance(v, float):
                return f"{v:>{w}.0f}" if v == int(v) else f"{v:>{w}.1f}"
            return f"{v:>{w}}"

        line = (f"{str(row['sector']):<20} {str(row['occ_group']):<30} "
                f"{fmt(row['T1'], 3)} {fmt(row['T2'], 3)} {fmt(row['T3'], 3)} {fmt(row['T4'], 3)} "
                f"{fmt(row['D'], 5, 2)} {fmt(row['t0'], 6, 3)} {fmt(row['T_18mo'], 5, 3)} "
                f"{fmt(row['f1'], 3)} {fmt(row['f2'], 3)} {fmt(row['f3'], 3)} "
                f"{fmt(row['F'], 3)} {fmt(row['R'], 5, 3)} "
                f"{fmt(row['E'], 5, 2)} "
                f"{fmt(row['friction_discount'], 6, 3)}")
        print(line)


def compute_averages(all_rows):
    """Compute overall averages across all 21 sectors."""
    fields = ["T1", "T2", "T3", "T4", "D", "t0", "T_18mo",
              "f1", "f2", "f3", "F", "R", "E", "friction_discount"]
    avgs = {}
    for field in fields:
        vals = [r[field] for r in all_rows if r[field] is not None]
        avgs[field] = sum(vals) / len(vals) if vals else None
    avgs["sector"] = "ALL SECTORS AVG"
    avgs["occ_group"] = f"(n={len(all_rows)} rows)"
    return avgs


def compute_averages_by_occ_group(all_rows, occ_group):
    """Compute averages for a specific occ_group across all sectors."""
    filtered = [r for r in all_rows if r["occ_group"] == occ_group]
    if not filtered:
        return None
    fields = ["T1", "T2", "T3", "T4", "D", "t0", "T_18mo",
              "f1", "f2", "f3", "F", "R", "E", "friction_discount"]
    avgs = {}
    for field in fields:
        vals = [r[field] for r in filtered if r[field] is not None]
        avgs[field] = sum(vals) / len(vals) if vals else None
    avgs["sector"] = f"AVG ({occ_group})"
    avgs["occ_group"] = f"(n={len(filtered)} sectors)"
    return avgs


def main():
    print(f"Loading workbook: {WORKBOOK}")
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)

    all_rows = read_frictions_tab(wb, "4H Frictions High")
    print(f"Total rows loaded from 4H: {len(all_rows)}")

    # --- RETAIL TRADE ---
    retail_rows = [r for r in all_rows if r["sector"] == "Retail Trade"]
    print_table(retail_rows, "RETAIL TRADE — All Occ Groups (4H Frictions High)")

    # --- CONSTRUCTION ---
    construction_rows = [r for r in all_rows if r["sector"] == "Construction"]
    print_table(construction_rows, "CONSTRUCTION — All Occ Groups (4H Frictions High)")

    # --- OVERALL AVERAGES ---
    overall_avg = compute_averages(all_rows)
    core_avg = compute_averages_by_occ_group(all_rows, "Core")
    g9_avg = compute_averages_by_occ_group(all_rows, "G9 Admin & Office Support")
    g1_avg = compute_averages_by_occ_group(all_rows, "G1 Executive & Management")

    print_table([overall_avg, core_avg, g9_avg, g1_avg],
                "CROSS-SECTOR AVERAGES FOR COMPARISON")

    # --- SPECIFIC DEEP DIVES ---
    print(f"\n{'='*120}")
    print("  DEEP DIVE: Key Questions")
    print(f"{'='*120}")

    # Q1: Retail Core vs G9
    retail_core = next((r for r in retail_rows if r["occ_group"] == "Core"), None)
    retail_g9 = next((r for r in retail_rows if r["occ_group"] == "G9 Admin & Office Support"), None)

    print("\n--- Retail Trade: Core vs G9 Admin & Office Support ---")
    if retail_core:
        print(f"  Core:  E={retail_core['E']}, T(18mo)={retail_core['T_18mo']:.3f}, R={retail_core['R']:.3f}, "
              f"E*T*R={retail_core['friction_discount']:.3f}")
        print(f"         T1={retail_core['T1']}, T2={retail_core['T2']}, T3={retail_core['T3']}, T4={retail_core['T4']} "
              f"-> D={retail_core['D']:.2f}, t0={retail_core['t0']:.3f}")
        print(f"         f1={retail_core['f1']}, f2={retail_core['f2']}, f3={retail_core['f3']} "
              f"-> F={retail_core['F']}, R={retail_core['R']:.3f}")
    if retail_g9:
        print(f"  G9:    E={retail_g9['E']}, T(18mo)={retail_g9['T_18mo']:.3f}, R={retail_g9['R']:.3f}, "
              f"E*T*R={retail_g9['friction_discount']:.3f}")
        print(f"         T1={retail_g9['T1']}, T2={retail_g9['T2']}, T3={retail_g9['T3']}, T4={retail_g9['T4']} "
              f"-> D={retail_g9['D']:.2f}, t0={retail_g9['t0']:.3f}")
        print(f"         f1={retail_g9['f1']}, f2={retail_g9['f2']}, f3={retail_g9['f3']} "
              f"-> F={retail_g9['F']}, R={retail_g9['R']:.3f}")

    # Q2: Construction Core vs G9
    constr_core = next((r for r in construction_rows if r["occ_group"] == "Core"), None)
    constr_g9 = next((r for r in construction_rows if r["occ_group"] == "G9 Admin & Office Support"), None)

    print("\n--- Construction: Core vs G9 Admin & Office Support ---")
    if constr_core:
        print(f"  Core:  E={constr_core['E']}, T(18mo)={constr_core['T_18mo']:.3f}, R={constr_core['R']:.3f}, "
              f"E*T*R={constr_core['friction_discount']:.3f}")
        print(f"         T1={constr_core['T1']}, T2={constr_core['T2']}, T3={constr_core['T3']}, T4={constr_core['T4']} "
              f"-> D={constr_core['D']:.2f}, t0={constr_core['t0']:.3f}")
        print(f"         f1={constr_core['f1']}, f2={constr_core['f2']}, f3={constr_core['f3']} "
              f"-> F={constr_core['F']}, R={constr_core['R']:.3f}")
    if constr_g9:
        print(f"  G9:    E={constr_g9['E']}, T(18mo)={constr_g9['T_18mo']:.3f}, R={constr_g9['R']:.3f}, "
              f"E*T*R={constr_g9['friction_discount']:.3f}")
        print(f"         T1={constr_g9['T1']}, T2={constr_g9['T2']}, T3={constr_g9['T3']}, T4={constr_g9['T4']} "
              f"-> D={constr_g9['D']:.2f}, t0={constr_g9['t0']:.3f}")
        print(f"         f1={constr_g9['f1']}, f2={constr_g9['f2']}, f3={constr_g9['f3']} "
              f"-> F={constr_g9['F']}, R={constr_g9['R']:.3f}")

    # --- FLAGGING ---
    print(f"\n{'='*120}")
    print("  FLAGS & ANOMALIES")
    print(f"{'='*120}")

    # Compare each row against overall averages
    fields_to_check = [
        ("T_18mo", "T(18mo)", "higher = faster adoption"),
        ("R", "R", "higher = less friction resistance"),
        ("E", "E", "higher = more employer willingness"),
        ("friction_discount", "E*T*R", "higher = more displacement"),
    ]

    target_sectors = ["Retail Trade", "Construction"]
    target_rows = [r for r in all_rows if r["sector"] in target_sectors]

    for field, label, desc in fields_to_check:
        # Compute overall stats
        all_vals = [r[field] for r in all_rows if r[field] is not None]
        avg = sum(all_vals) / len(all_vals)
        std = (sum((v - avg)**2 for v in all_vals) / len(all_vals)) ** 0.5
        p75 = sorted(all_vals)[int(0.75 * len(all_vals))]
        p25 = sorted(all_vals)[int(0.25 * len(all_vals))]

        print(f"\n  {label} ({desc})")
        print(f"    Overall: mean={avg:.3f}, std={std:.3f}, p25={p25:.3f}, p75={p75:.3f}")

        for row in target_rows:
            val = row[field]
            if val is None:
                continue
            zscore = (val - avg) / std if std > 0 else 0
            flag = ""
            if abs(zscore) > 1.5:
                flag = " *** OUTLIER (z={:.1f})".format(zscore)
            elif abs(zscore) > 1.0:
                flag = " ** NOTABLE (z={:.1f})".format(zscore)
            if flag or val > p75 or val < p25:
                marker = "HIGH" if val > avg else "LOW"
                print(f"    {row['sector']:20s} {row['occ_group']:30s} = {val:.3f} [{marker}]{flag}")

    # --- Compare Retail and Construction G9 against all other sectors' G9 ---
    print(f"\n{'='*120}")
    print("  G9 Admin & Office Support: Retail & Construction vs All Sectors")
    print(f"{'='*120}")

    g9_rows = [r for r in all_rows if r["occ_group"] == "G9 Admin & Office Support"]
    g9_rows_sorted = sorted(g9_rows, key=lambda r: r["friction_discount"] if r["friction_discount"] else 0, reverse=True)
    print(f"\n  Ranked by E*T*R (friction discount), highest first:")
    for i, row in enumerate(g9_rows_sorted, 1):
        marker = " <<<<" if row["sector"] in target_sectors else ""
        if row["friction_discount"] is not None:
            print(f"    {i:2d}. {row['sector']:25s}  E={row['E']:.2f}  T={row['T_18mo']:.3f}  "
                  f"R={row['R']:.3f}  E*T*R={row['friction_discount']:.3f}{marker}")

    # --- Compare Core rows ---
    print(f"\n{'='*120}")
    print("  Core: Retail & Construction vs All Sectors")
    print(f"{'='*120}")

    core_rows = [r for r in all_rows if r["occ_group"] == "Core"]
    core_rows_sorted = sorted(core_rows, key=lambda r: r["friction_discount"] if r["friction_discount"] else 0, reverse=True)
    print(f"\n  Ranked by E*T*R (friction discount), highest first:")
    for i, row in enumerate(core_rows_sorted, 1):
        marker = " <<<<" if row["sector"] in target_sectors else ""
        if row["friction_discount"] is not None:
            print(f"    {i:2d}. {row['sector']:25s}  E={row['E']:.2f}  T={row['T_18mo']:.3f}  "
                  f"R={row['R']:.3f}  E*T*R={row['friction_discount']:.3f}{marker}")

    wb.close()
    print(f"\n{'='*120}")
    print("  Analysis complete. Workbook was NOT modified.")
    print(f"{'='*120}")


if __name__ == "__main__":
    main()
