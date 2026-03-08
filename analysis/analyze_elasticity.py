"""Analyze E (elasticity) values across all occ groups and sectors.

Reads from 4H Frictions High and 5 Results tabs. Does NOT modify the workbook.
"""
import math
import openpyxl
import os

WORKBOOK = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                        "jobs-data-v3.xlsx")


def compute_T_18mo(T1, T2, T3, T4, alpha=1.2):
    if any(x is None for x in [T1, T2, T3, T4]):
        return None
    D = (T1 + T2 + T3 + T4) / 4
    t0 = 1.5 * D - 0.5 * T4 - 0.5
    return 1 / (1 + math.exp(-alpha * (1.5 - t0)))


def compute_R(f1, f2, f3):
    if any(x is None for x in [f1, f2, f3]):
        return None
    F = f1 + f2 + f3
    return 1 - 0.7 * (F - 3) / 9


def main():
    print(f"Loading workbook: {WORKBOOK}")
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)

    # ── PART 1: Read 4H Frictions High ──────────────────────────────────────
    ws4h = wb["4H Frictions High"]
    print("\n" + "=" * 100)
    print("PART 1: 4H Frictions High — All rows")
    print("=" * 100)

    # First, let's see what the headers actually are
    print("\nHeaders (row 4):")
    for c in range(1, 20):
        val = ws4h.cell(4, c).value
        print(f"  Col {c}: {val}")

    # Read all friction rows
    friction_rows = []
    for r in range(5, ws4h.max_row + 1):
        sector = ws4h.cell(r, 2).value
        occ_group = ws4h.cell(r, 3).value
        if not sector or not occ_group:
            continue

        T1 = ws4h.cell(r, 6).value
        T2 = ws4h.cell(r, 7).value
        T3 = ws4h.cell(r, 8).value
        T4 = ws4h.cell(r, 9).value
        f1 = ws4h.cell(r, 13).value
        f2 = ws4h.cell(r, 14).value
        f3 = ws4h.cell(r, 15).value
        E = ws4h.cell(r, 18).value

        T_18mo = compute_T_18mo(T1, T2, T3, T4)
        R = compute_R(f1, f2, f3)

        friction_rows.append({
            "sector": sector, "occ_group": occ_group,
            "T1": T1, "T2": T2, "T3": T3, "T4": T4,
            "f1": f1, "f2": f2, "f3": f3,
            "E": E, "T_18mo": T_18mo, "R": R,
        })

    # ── PART 1 Detail: G9 Admin & Office Support ────────────────────────────
    print("\n" + "=" * 100)
    print("PART 1: G9 Admin & Office Support — Friction Detail (4H)")
    print("=" * 100)

    g9_rows = [r for r in friction_rows if "G9" in str(r["occ_group"]) or "Admin" in str(r["occ_group"])]
    print(f"\nFound {len(g9_rows)} G9 rows\n")

    hdr = f"{'Sector':<30} {'E':>5} {'T1':>4} {'T2':>4} {'T3':>4} {'T4':>4} {'T_18mo':>7} {'f1':>4} {'f2':>4} {'f3':>4} {'R':>6} {'E*T*R':>7}"
    print(hdr)
    print("-" * len(hdr))

    for r in sorted(g9_rows, key=lambda x: x["sector"]):
        E = r["E"] if r["E"] is not None else "N/A"
        T_18mo = r["T_18mo"]
        R = r["R"]
        if isinstance(E, (int, float)) and T_18mo is not None and R is not None:
            etr = E * T_18mo * R
            etr_s = f"{etr:.4f}"
        else:
            etr_s = "N/A"
        T_18mo_s = f"{T_18mo:.4f}" if T_18mo is not None else "N/A"
        R_s = f"{R:.4f}" if R is not None else "N/A"

        print(f"{r['sector']:<30} {E:>5} {r['T1'] or 'N/A':>4} {r['T2'] or 'N/A':>4} "
              f"{r['T3'] or 'N/A':>4} {r['T4'] or 'N/A':>4} {T_18mo_s:>7} "
              f"{r['f1'] or 'N/A':>4} {r['f2'] or 'N/A':>4} {r['f3'] or 'N/A':>4} "
              f"{R_s:>6} {etr_s:>7}")

    # ── PART 2: G9 from 5 Results ───────────────────────────────────────────
    print("\n" + "=" * 100)
    print("PART 2: G9 Admin & Office Support — Displacement from 5 Results")
    print("=" * 100)

    ws5 = wb["5 Results"]
    # Headers row 1:
    # 1=SOC_Code, 2=Job_Title, 3=Sector, 4=Occupation_Group,
    # 5=Employment_2024_K, 6=Median_Wage,
    # 7=tc_adj_mod, 8=tc_adj_sig, 9=w, 10=a_mod, 11=a_sig, 12=S_mod, 13=S_sig,
    # 14=d_max, 15=E, 16=T_18mo_low, 17=T_18mo_high, 18=R_low, 19=R_high,
    # 20=d_mod_low, 21=d_mod_high, 22=d_sig_low, 23=d_sig_high,
    # 24=displaced_K_mod_low, 25=displaced_K_mod_high,
    # 26=displaced_K_sig_low, 27=displaced_K_sig_high

    # Verify headers
    print("\n5 Results headers:")
    for c in range(1, 28):
        val = ws5.cell(1, c).value
        print(f"  Col {c}: {val}")

    # Collect all results rows
    results_rows = []
    for r in range(2, ws5.max_row + 1):
        soc = ws5.cell(r, 1).value
        if not soc:
            continue
        results_rows.append({
            "soc": soc,
            "title": ws5.cell(r, 2).value,
            "sector": ws5.cell(r, 3).value,
            "occ_group": ws5.cell(r, 4).value,
            "emp_K": ws5.cell(r, 5).value or 0,
            "E": ws5.cell(r, 15).value,
            "displaced_K_sig_high": ws5.cell(r, 27).value or 0,
            "displaced_K_mod_high": ws5.cell(r, 25).value or 0,
        })

    # OCC_GROUP_NORMALIZE (Jobs-tab codes)
    OCC_GROUP_NORMALIZE = {
        "Core": "Core",
        "G1_Exec_Management": "G1 Executive & Management",
        "G2_HR_People": "G2 HR & People Ops",
        "G3_Finance_Accounting": "G3 Finance & Accounting",
        "G4_IT_Digital": "G4 IT & Digital",
        "G5_Marketing_Creative": "G5 Marketing & Creative",
        "G6_Sales_BizDev": "G6 Sales & Business Dev",
        "G7_Legal_Compliance": "G7 Legal & Compliance",
        "G8_Procurement_Supply": "G8 Procurement & Supply Chain",
        "G9_Admin_Office": "G9 Admin & Office Support",
    }
    REVERSE_NORMALIZE = {v: k for k, v in OCC_GROUP_NORMALIZE.items()}

    # G9 results by sector
    g9_results = [r for r in results_rows if r["occ_group"] == "G9_Admin_Office"]
    print(f"\nFound {len(g9_results)} G9 SOCs in 5 Results\n")

    # Aggregate by sector
    from collections import defaultdict
    g9_by_sector = defaultdict(lambda: {"emp_K": 0, "disp_K_sig": 0, "disp_K_mod": 0})
    for r in g9_results:
        s = r["sector"]
        g9_by_sector[s]["emp_K"] += r["emp_K"]
        g9_by_sector[s]["disp_K_sig"] += r["displaced_K_sig_high"]
        g9_by_sector[s]["disp_K_mod"] += r["displaced_K_mod_high"]

    hdr2 = f"{'Sector':<30} {'G9 Emp K':>10} {'G9 Disp K (Sig)':>16} {'G9 Rate %':>10}"
    print(hdr2)
    print("-" * len(hdr2))

    g9_total_emp = 0
    g9_total_disp = 0
    for sector in sorted(g9_by_sector.keys(), key=lambda s: g9_by_sector[s]["disp_K_sig"], reverse=True):
        d = g9_by_sector[sector]
        rate = d["disp_K_sig"] / d["emp_K"] * 100 if d["emp_K"] > 0 else 0
        print(f"{sector:<30} {d['emp_K']:>10.1f} {d['disp_K_sig']:>16.2f} {rate:>9.2f}%")
        g9_total_emp += d["emp_K"]
        g9_total_disp += d["disp_K_sig"]

    g9_rate = g9_total_disp / g9_total_emp * 100 if g9_total_emp > 0 else 0
    print("-" * len(hdr2))
    print(f"{'TOTAL':<30} {g9_total_emp:>10.1f} {g9_total_disp:>16.2f} {g9_rate:>9.2f}%")

    # ── PART 3: All occ groups — E values and displacement ──────────────────
    print("\n" + "=" * 100)
    print("PART 3: E Values and Displacement for ALL Occupation Groups")
    print("=" * 100)

    # Group friction rows by occ_group
    friction_by_group = defaultdict(list)
    for r in friction_rows:
        friction_by_group[r["occ_group"]].append(r)

    # Group results by occ_group
    results_by_group = defaultdict(list)
    for r in results_rows:
        results_by_group[r["occ_group"]].append(r)

    # Ordered group list
    ordered_groups = [
        "Core",
        "G1 Executive & Management",
        "G2 HR & People Ops",
        "G3 Finance & Accounting",
        "G4 IT & Digital",
        "G5 Marketing & Creative",
        "G6 Sales & Business Dev",
        "G7 Legal & Compliance",
        "G8 Procurement & Supply Chain",
        "G9 Admin & Office Support",
    ]

    for group in ordered_groups:
        fric_rows = friction_by_group.get(group, [])
        jobs_code = REVERSE_NORMALIZE.get(group, group)
        res_rows = results_by_group.get(jobs_code, [])

        total_emp = sum(r["emp_K"] for r in res_rows)
        total_disp_sig = sum(r["displaced_K_sig_high"] for r in res_rows)
        total_disp_mod = sum(r["displaced_K_mod_high"] for r in res_rows)

        print(f"\n{'─' * 100}")
        print(f"  {group} (code: {jobs_code})")
        print(f"  {len(fric_rows)} sector entries in 4H | {len(res_rows)} SOCs in Results")
        print(f"  Total: {total_emp:.1f}K employed, {total_disp_sig:.1f}K displaced (Sig High), {total_disp_mod:.1f}K (Mod High)")
        if total_emp > 0:
            print(f"  Displacement rate: {total_disp_sig/total_emp*100:.2f}% (Sig), {total_disp_mod/total_emp*100:.2f}% (Mod)")
        print(f"{'─' * 100}")

        # Show E values by sector
        e_values = set()
        hdr3 = f"  {'Sector':<30} {'E':>5} {'T_18mo':>7} {'R':>6} {'E*T*R':>7}"
        print(hdr3)
        print("  " + "-" * (len(hdr3) - 2))
        for fr in sorted(fric_rows, key=lambda x: x["sector"]):
            E = fr["E"]
            T = fr["T_18mo"]
            R = fr["R"]
            if isinstance(E, (int, float)):
                e_values.add(E)
            if isinstance(E, (int, float)) and T is not None and R is not None:
                etr = E * T * R
                etr_s = f"{etr:.4f}"
            else:
                etr_s = "N/A"
            T_s = f"{T:.4f}" if T is not None else "N/A"
            R_s = f"{R:.4f}" if R is not None else "N/A"
            E_s = f"{E}" if E is not None else "N/A"
            print(f"  {fr['sector']:<30} {E_s:>5} {T_s:>7} {R_s:>6} {etr_s:>7}")

        # Flag analysis
        if e_values:
            unique_e = sorted(e_values)
            print(f"\n  >>> Unique E values: {unique_e}")
            if len(unique_e) == 1:
                print(f"  >>> ALL sectors use E={unique_e[0]} for {group}")
            for ev in unique_e:
                if ev == 1.0:
                    sectors_1 = [fr["sector"] for fr in fric_rows if fr["E"] == 1.0]
                    print(f"  >>> E=1.0 (INELASTIC, no Jevons) in {len(sectors_1)} sectors: {', '.join(sectors_1[:5])}{'...' if len(sectors_1)>5 else ''}")
                elif ev == 0.25:
                    sectors_025 = [fr["sector"] for fr in fric_rows if fr["E"] == 0.25]
                    print(f"  >>> E=0.25 (ELASTIC, strong Jevons) in {len(sectors_025)} sectors: {', '.join(sectors_025[:5])}{'...' if len(sectors_025)>5 else ''}")
                elif ev < 1.0:
                    sectors_mid = [fr["sector"] for fr in fric_rows if fr["E"] == ev]
                    print(f"  >>> E={ev} (PARTIAL elasticity) in {len(sectors_mid)} sectors: {', '.join(sectors_mid[:5])}{'...' if len(sectors_mid)>5 else ''}")

    # ── PART 4: Detailed Assessment per Occ Group ───────────────────────────
    print("\n" + "=" * 100)
    print("PART 4: ELASTICITY ASSESSMENT — Does Each E Value Make Sense?")
    print("=" * 100)

    # Build a summary dict for quick reference
    e_summary = {}
    for group in ordered_groups:
        fric_rows = friction_by_group.get(group, [])
        e_by_sector = {}
        for fr in fric_rows:
            e_by_sector[fr["sector"]] = fr["E"]
        e_summary[group] = e_by_sector

    assessments = {
        "Core": {
            "question": "If AI automates core production tasks, does demand for output expand?",
            "analysis": (
                "Core is sector-specific (nurses, teachers, welders, etc.). "
                "Elasticity depends heavily on sector. Manufacturing may see Jevons (cheaper production -> more output), "
                "but Healthcare (demand is insurance-constrained) and Education (demand is enrollment-constrained) may not. "
                "Government is clearly inelastic. "
                "A uniform E across all Core rows would be wrong — sector variation is essential."
            ),
        },
        "G1 Executive & Management": {
            "question": "If AI makes management more efficient, do companies need MORE managers?",
            "analysis": (
                "No — management is a coordination overhead, not a revenue-generating activity. "
                "Cheaper management doesn't create demand for more management. "
                "E=1.0 (inelastic) makes sense for almost all sectors. "
                "Possible exception: Consulting/Advisory firms where 'management' IS the product."
            ),
        },
        "G2 HR & People Ops": {
            "question": "If AI automates recruiting/onboarding, does the company do MORE recruiting?",
            "analysis": (
                "Mostly no. HR is an overhead function — hiring volume is driven by business growth, not HR efficiency. "
                "However, if AI makes recruiting cheaper, companies might afford to be more selective (more candidates per hire), "
                "and could invest more in employee experience. Mild Jevons at most. "
                "E=1.0 is mostly defensible, but E=0.85 might be more accurate for high-growth sectors."
            ),
        },
        "G3 Finance & Accounting": {
            "question": "If AI automates bookkeeping/reporting, do companies want MORE financial analysis?",
            "analysis": (
                "Bookkeeping: No — it's compliance-driven, fixed by transaction volume. E=1.0 makes sense. "
                "Financial analysis/FP&A: Maybe mild Jevons — cheaper analysis could lead to more frequent reporting, "
                "more granular budgeting. But the core demand is transaction-driven. "
                "E=1.0 is reasonable for most sectors. Financial Services sector might be E=0.75 since analysis IS the product."
            ),
        },
        "G4 IT & Digital": {
            "question": "If AI automates IT support and dev tasks, does the company need more IT?",
            "analysis": (
                "YES — this is a strong Jevons candidate. Companies consistently want more software, more automation, "
                "more digital transformation. If AI makes dev cheaper, companies build more. "
                "Tech companies especially: cheaper dev = more products = more dev. "
                "E=0.50 or even E=0.25 could be justified for Tech sector. "
                "Other sectors: E=0.50-0.75 — they'd invest more in digital but IT isn't their core product."
            ),
        },
        "G5 Marketing & Creative": {
            "question": "If AI makes content creation cheaper, does the company produce more content?",
            "analysis": (
                "YES — strong Jevons effect. Content marketing is already volume-constrained by cost. "
                "Cheaper content = more A/B tests, more personalization, more channels. "
                "Social media, advertising, and content marketing all scale with reduced cost. "
                "E=0.25-0.50 is defensible for most sectors. "
                "Exception: Government/Healthcare where marketing is limited by regulation."
            ),
        },
        "G6 Sales & Business Dev": {
            "question": "If AI automates sales outreach, does the company do more outreach?",
            "analysis": (
                "YES — sales teams are already volume-constrained. Cheaper outreach = more prospects contacted, "
                "more follow-ups, more personalized pitches. SDR functions would massively scale. "
                "E=0.25-0.50 is defensible for most commercial sectors. "
                "Exception: Government (procurement-driven, not sales-driven)."
            ),
        },
        "G7 Legal & Compliance": {
            "question": "If AI automates legal research/contract review, do companies do more legal work?",
            "analysis": (
                "Partially. Contract review volume is driven by deal volume, not lawyer efficiency. "
                "But: cheaper legal review could lead to more contracts reviewed, more compliance audits, "
                "more proactive legal risk assessment. "
                "Law firms: mild Jevons (cheaper research = more research done per case). "
                "Other sectors: E=1.0 is mostly right since legal is a cost center. "
                "E=0.85 might be more accurate for sectors with heavy contract/compliance loads."
            ),
        },
        "G8 Procurement & Supply Chain": {
            "question": "If AI automates procurement/logistics, does the company do more procurement?",
            "analysis": (
                "Partially. Procurement volume is driven by production needs, not procurement efficiency. "
                "But: cheaper procurement analysis could lead to more vendor evaluation, more frequent rebidding, "
                "more sophisticated demand planning. "
                "E=0.75-1.0 for most sectors. Retail/Manufacturing might benefit from E=0.75 (more optimization). "
                "Staffing: E=0.50 — matching/placement IS the product."
            ),
        },
        "G9 Admin & Office Support": {
            "question": "If AI automates admin tasks (scheduling, filing, correspondence), does the company need more admin?",
            "analysis": (
                "NO — admin is pure overhead. Cheaper admin doesn't create demand for more admin. "
                "Companies already minimize admin headcount. AI would just accelerate that trend. "
                "E=1.0 (fully inelastic) makes sense for virtually all sectors. "
                "No plausible Jevons effect — nobody says 'scheduling is cheaper, let's schedule more meetings.'"
            ),
        },
    }

    for group in ordered_groups:
        e_vals = e_summary.get(group, {})
        assessment = assessments.get(group, {"question": "?", "analysis": "No assessment provided."})

        unique_e = sorted(set(v for v in e_vals.values() if v is not None))

        print(f"\n{'━' * 100}")
        print(f"  {group}")
        print(f"{'━' * 100}")
        print(f"  QUESTION: {assessment['question']}")
        print(f"  CURRENT E VALUES: {unique_e if unique_e else 'NONE SET'}")

        # Show per-sector breakdown if there's variation
        if len(unique_e) > 1:
            print(f"  VARIATION BY SECTOR:")
            for sector in sorted(e_vals.keys()):
                print(f"    {sector:<30} E={e_vals[sector]}")
        elif len(unique_e) == 1:
            print(f"  UNIFORM: E={unique_e[0]} across all {len(e_vals)} sectors")

        print(f"\n  ASSESSMENT:")
        for line in assessment["analysis"].split(". "):
            if line.strip():
                print(f"    {line.strip()}.")

        # Verdict
        if group == "Core":
            print(f"\n  VERDICT: Requires per-sector review (Core is heterogeneous)")
        elif unique_e == [1.0]:
            if group in ["G1 Executive & Management", "G9 Admin & Office Support"]:
                print(f"\n  VERDICT: E=1.0 is CORRECT — {group} is pure overhead, no Jevons effect")
            elif group in ["G4 IT & Digital", "G5 Marketing & Creative", "G6 Sales & Business Dev"]:
                print(f"\n  VERDICT: E=1.0 is WRONG — strong Jevons effect expected. Should be E=0.25-0.50")
                print(f"           This is OVER-COUNTING displacement by 2-4x for this group!")
            elif group in ["G2 HR & People Ops", "G3 Finance & Accounting", "G7 Legal & Compliance"]:
                print(f"\n  VERDICT: E=1.0 is MOSTLY CORRECT but could be E=0.85 for some sectors")
            elif group in ["G8 Procurement & Supply Chain"]:
                print(f"\n  VERDICT: E=1.0 is SLIGHTLY HIGH — E=0.75-1.0 depending on sector")
        elif unique_e:
            print(f"\n  VERDICT: Mixed E values — review per-sector assignments above")

    # ── Final Summary Table ─────────────────────────────────────────────────
    print("\n" + "=" * 100)
    print("SUMMARY: Impact of E on Displacement by Occ Group (Significant + High Friction)")
    print("=" * 100)

    hdr_sum = f"{'Occ Group':<35} {'Emp K':>8} {'Disp K':>8} {'Rate%':>7} {'E':>12} {'E Correct?':>20}"
    print(hdr_sum)
    print("-" * len(hdr_sum))

    verdict_map = {
        "Core": "Per-sector review",
        "G1_Exec_Management": "YES (overhead)",
        "G2_HR_People": "Mostly (E=0.85?)",
        "G3_Finance_Accounting": "Mostly (E=0.85?)",
        "G4_IT_Digital": "NO -> E=0.25-0.50",
        "G5_Marketing_Creative": "NO -> E=0.25-0.50",
        "G6_Sales_BizDev": "NO -> E=0.25-0.50",
        "G7_Legal_Compliance": "Mostly (E=0.85?)",
        "G8_Procurement_Supply": "Slightly high",
        "G9_Admin_Office": "YES (overhead)",
    }

    grand_emp = 0
    grand_disp = 0
    for group in ordered_groups:
        jobs_code = REVERSE_NORMALIZE.get(group, group)
        res_rows = results_by_group.get(jobs_code, [])
        total_emp = sum(r["emp_K"] for r in res_rows)
        total_disp = sum(r["displaced_K_sig_high"] for r in res_rows)
        rate = total_disp / total_emp * 100 if total_emp > 0 else 0

        fric_rows_g = friction_by_group.get(group, [])
        unique_e = sorted(set(fr["E"] for fr in fric_rows_g if fr["E"] is not None))
        e_str = ",".join(str(e) for e in unique_e) if unique_e else "N/A"

        verdict = verdict_map.get(jobs_code, "?")

        print(f"{group:<35} {total_emp:>8.1f} {total_disp:>8.1f} {rate:>6.2f}% {e_str:>12} {verdict:>20}")
        grand_emp += total_emp
        grand_disp += total_disp

    grand_rate = grand_disp / grand_emp * 100 if grand_emp > 0 else 0
    print("-" * len(hdr_sum))
    print(f"{'TOTAL':<35} {grand_emp:>8.1f} {grand_disp:>8.1f} {grand_rate:>6.2f}%")

    # ── Estimate impact of correcting E ─────────────────────────────────────
    print("\n" + "=" * 100)
    print("ESTIMATED IMPACT: What if we correct E for groups with Jevons effect?")
    print("=" * 100)

    proposed_changes = {
        "G4_IT_Digital": {"current": 1.0, "proposed": 0.50, "rationale": "Strong Jevons — cheaper IT = more digital investment"},
        "G5_Marketing_Creative": {"current": 1.0, "proposed": 0.50, "rationale": "Strong Jevons — cheaper content = more content"},
        "G6_Sales_BizDev": {"current": 1.0, "proposed": 0.50, "rationale": "Strong Jevons — cheaper outreach = more outreach"},
        "G8_Procurement_Supply": {"current": 1.0, "proposed": 0.75, "rationale": "Mild Jevons — some expansion of procurement analysis"},
    }

    total_reduction = 0
    for jobs_code, change in proposed_changes.items():
        group_name = OCC_GROUP_NORMALIZE.get(jobs_code, jobs_code)
        res_rows = results_by_group.get(jobs_code, [])
        current_disp = sum(r["displaced_K_sig_high"] for r in res_rows)
        # displacement is proportional to E, so new_disp = current_disp * (proposed/current)
        ratio = change["proposed"] / change["current"]
        new_disp = current_disp * ratio
        reduction = current_disp - new_disp

        print(f"\n  {group_name} ({jobs_code})")
        print(f"    Current E: {change['current']} -> Proposed E: {change['proposed']}")
        print(f"    Rationale: {change['rationale']}")
        print(f"    Current displaced (Sig High): {current_disp:.1f}K")
        print(f"    Proposed displaced:           {new_disp:.1f}K")
        print(f"    Reduction:                    {reduction:.1f}K ({reduction/current_disp*100:.0f}%)" if current_disp > 0 else "    Reduction: 0K")
        total_reduction += reduction

    print(f"\n  TOTAL REDUCTION if E corrected: {total_reduction:.1f}K fewer displaced workers")
    print(f"  Current total (Sig High):       {grand_disp:.1f}K")
    print(f"  Revised total (Sig High):       {grand_disp - total_reduction:.1f}K")
    print(f"  Percentage change:              {-total_reduction/grand_disp*100:.1f}%" if grand_disp > 0 else "")

    wb.close()
    print("\nDone. Workbook was NOT modified.")


if __name__ == "__main__":
    main()
